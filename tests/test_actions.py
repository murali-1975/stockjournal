"""
Tests for src/actions.py
========================

Validates the exit recommendation engine:
    - Rule 1 (Cheat Exit - Interpretation B)
    - Rule 2 (Tranch Exit - LTP near Stop Loss)
    - Overwrite/append logic for Action Tracker sheets
"""

import os
import tempfile
import unittest
from unittest.mock import patch
import pandas as pd
import openpyxl
import datetime

from src.actions import (
    generate_action_recommendations,
    write_recommendations_to_excel,
    HEADERS
)


class TestActions(unittest.TestCase):
    """Test suite for recommendations rules and writer logic."""

    def test_generate_sell_recommendations_cheat(self):
        """Rule 1: Underperforming Cheat Satellite stocks (return < 5%) should be recommended by holding period."""
        # Create a mock Current_Portfolio dataframe
        data = {
            'Symbol': ['CHEAT1', 'CHEAT2', 'CHEAT3', 'CHEAT4', 'CHEAT5', 'CHEAT6', 'CORE1', 'SAT_OK'],
            'Cap': ['Small Cap'] * 8,
            'TF_Sector': ['Industrial'] * 8,
            'TF_Classification': ['Satellite', 'Satellite', 'Satellite', 'Satellite', 'Satellite', 'Satellite', 'Core', 'Satellite'],
            'Latest_Tranche': ['Cheat 1', 'Cheat 2', 'Cheat 1', 'Cheat 2', 'Cheat 1', 'Cheat 2', 'Cheat 1', 'Cheat 1'],
            'Current_Quantity': [100] * 8,
            'Average_Buy_Price': [10.0] * 8,
            'Trend': ['▲'] * 8,
            'SL': [9.0] * 8,
            'LTP_SL_Diff_Pct': [0.20] * 8, # not close to SL
            'Invested_Value': [1000.0] * 8,
            'LTP': [10.0] * 8,
            'Prev_Day_Close': [10.0] * 8,
            'Prev_Week_Close': [10.0] * 8,
            'EMA9': [10.0] * 8,
            'EMA10': [10.0] * 8,
            'EMA11': [10.0] * 8,
            'EMA21': [10.0] * 8,
            'Current_Value': [1000.0] * 8,
            'Unrealized_PnL': [0.0] * 8,
            # CHEAT1-6 returns are:
            # CHEAT1: -10% (underperforming, held 200d) -> SELL
            # CHEAT2: +1%  (underperforming, held 150d) -> SELL
            # CHEAT3: +4%  (underperforming, held 10d)  -> SELL
            # CHEAT4: -2%  (underperforming, held 5d)   -> SELL
            # CHEAT5: +2%  (underperforming, held 2d)   -> SELL
            # CHEAT6: +3%  (underperforming, held 1d)   -> (Not in top 5, held 1d) -> HOLD
            # CORE1: -10%  (but it is CORE, not Satellite!) -> HOLD
            # SAT_OK: +10% (Satellite Cheat but return is >= 5%) -> HOLD
            'Return_Pct': [-0.10, 0.01, 0.04, -0.02, 0.02, 0.03, -0.10, 0.10],
            'XIRR': [0.0] * 8,
            'Holding_Period': [200, 150, 10, 5, 2, 1, 300, 250],
            'Split_Info': [''] * 8,
            'Adj_Required': ['No'] * 8
        }
        df = pd.DataFrame(data)
        recs = generate_action_recommendations(df)
        recs = recs[recs['Action (Recommended) '] == 'SELL']

        self.assertFalse(recs.empty)
        self.assertEqual(len(recs), 5)
        
        symbols = recs['Stock '].tolist()
        self.assertIn('CHEAT1', symbols)
        self.assertIn('CHEAT2', symbols)
        self.assertIn('CHEAT3', symbols)
        self.assertIn('CHEAT4', symbols)
        self.assertIn('CHEAT5', symbols)
        self.assertNotIn('CHEAT6', symbols)
        self.assertNotIn('CORE1', symbols)
        self.assertNotIn('SAT_OK', symbols)

        recs_dict = recs.set_index('Stock ').to_dict('index')
        self.assertEqual(recs_dict['CHEAT1']['Reason'], 'Cheat holding period long and red (200 days, -10.00% return)')
        self.assertEqual(recs_dict['CHEAT2']['Reason'], 'Cheat holding period long and less return (150 days, 1.00% return)')

    def test_generate_sell_recommendations_tranch(self):
        """Rule 2: Tranch stocks with LTP close to SL (< 5% difference) should be recommended."""
        data = {
            'Symbol': ['TRANCH1', 'TRANCH2', 'TRANCH3'],
            'Cap': ['Small Cap'] * 3,
            'TF_Sector': ['Industrial'] * 3,
            'TF_Classification': ['Satellite', 'Satellite', 'Core'],
            'Latest_Tranche': ['Tranch 1', 'Tranch 2', 'Tranch 1'],
            'Current_Quantity': [100] * 3,
            'Average_Buy_Price': [10.0] * 3,
            'Trend': ['▲'] * 3,
            'SL': [9.5] * 3,
            # TRANCH1 is close to SL (3% difference) -> SELL
            # TRANCH2 is far from SL (10% difference) -> HOLD
            # TRANCH3 is close to SL (3%) but is CORE -> HOLD
            'LTP_SL_Diff_Pct': [0.03, 0.10, 0.03], 
            'Invested_Value': [1000.0] * 3,
            'LTP': [10.0] * 3,
            'Prev_Day_Close': [10.0] * 3,
            'Prev_Week_Close': [10.0] * 3,
            'EMA9': [10.0] * 3,
            'EMA10': [10.0] * 3,
            'EMA11': [10.0] * 3,
            'EMA21': [10.0] * 3,
            'Current_Value': [1000.0] * 3,
            'Unrealized_PnL': [0.0] * 3,
            'Return_Pct': [0.10] * 3,
            'XIRR': [0.0] * 3,
            'Holding_Period': [50] * 3,
            'Split_Info': [''] * 3,
            'Adj_Required': ['No'] * 3
        }
        df = pd.DataFrame(data)
        recs = generate_action_recommendations(df)

        self.assertFalse(recs.empty)
        self.assertEqual(len(recs), 1)
        symbols = recs['Stock '].tolist()
        self.assertEqual(symbols, ['TRANCH1'])
        self.assertEqual(recs.iloc[0]['Reason'], 'Tranch LTP close to stop loss (< 5% difference)')

    def test_generate_add_recommendations_cheat_and_tranch(self):
        """Rule 1 and Rule 2: Enforces Cheat ADD (positive returns + holding days) and Tranch ADD (momentum bands)."""
        data = {
            'Symbol': ['CHEAT_ADD', 'TRANCH_MOM1', 'TRANCH_MOM2', 'TRANCH_HOLD'],
            'Cap': ['Small Cap'] * 4,
            'TF_Sector': ['Industrial'] * 4,
            'TF_Classification': ['Satellite', 'Satellite', 'Satellite', 'Satellite'],
            'Latest_Tranche': ['Cheat 1', 'Tranch 1', 'Tranch 2', 'Tranch 1'],
            'Current_Quantity': [100] * 4,
            'Average_Buy_Price': [10.0] * 4,
            'Trend': ['▲'] * 4,
            'SL': [9.0] * 4,
            'LTP_SL_Diff_Pct': [0.20] * 4, # safe from SL
            'Invested_Value': [1000.0] * 4,
            'LTP': [10.9, 10.3, 11.5, 9.0],
            'Prev_Day_Close': [10.0] * 4,
            # CHEAT_ADD: return is +9%, held 13d -> ADD
            # TRANCH_MOM1: LTP 10.3 > PWC 10.0 (3.00% above) -> ADD [0 to 5%]
            # TRANCH_MOM2: LTP 11.5 > PWC 10.0 (15.00% above) -> ADD [> 5%]
            # TRANCH_HOLD: LTP 9.0 < PWC 10.0 -> HOLD
            'Prev_Week_Close': [10.0, 10.0, 10.0, 10.0],
            'EMA9': [10.0] * 4,
            'EMA10': [10.0] * 4,
            'EMA11': [10.0] * 4,
            'EMA21': [10.0] * 4,
            'Current_Value': [1000.0] * 4,
            'Unrealized_PnL': [90.0, 30.0, 150.0, -100.0],
            'Return_Pct': [0.09, 0.03, 0.15, -0.10],
            'XIRR': [0.0] * 4,
            'Holding_Period': [13, 50, 60, 70],
            'Split_Info': [''] * 4,
            'Adj_Required': ['No'] * 4
        }
        df = pd.DataFrame(data)
        recs = generate_action_recommendations(df)

        self.assertFalse(recs.empty)
        # Should have exactly 3 ADD recommendations
        self.assertEqual(len(recs), 3)
        self.assertEqual(set(recs['Action (Recommended) ']), {'ADD'})

        recs_dict = recs.set_index('Stock ').to_dict('index')
        self.assertEqual(recs_dict['CHEAT_ADD']['Reason'], 'Cheat positive return (13 days, 9.00% return, 9.00% return from prev week close)')
        self.assertEqual(recs_dict['TRANCH_MOM1']['Reason'], 'Tranch LTP above previous week close [0 to 5%] (3.00% above)')
        self.assertEqual(recs_dict['TRANCH_MOM2']['Reason'], 'Tranch LTP above previous week close [> 5%] (15.00% above)')
        self.assertNotIn('TRANCH_HOLD', recs_dict)

    def test_generate_recommendations_precedence(self):
        """Precedence test: SELL recommendations must completely override ADD recommendations for the same stock."""
        data = {
            'Symbol': ['CHEAT_CONFLICT', 'TRANCH_CONFLICT'],
            'Cap': ['Small Cap'] * 2,
            'TF_Sector': ['Industrial'] * 2,
            'TF_Classification': ['Satellite', 'Satellite'],
            'Latest_Tranche': ['Cheat 1', 'Tranch 1'],
            'Current_Quantity': [100] * 2,
            'Average_Buy_Price': [10.0] * 2,
            'Trend': ['▲'] * 2,
            # CHEAT_CONFLICT: return +2% (positive, ADD candidate) but underperforming and long hold (150 days) -> SELL overrides ADD!
            # TRANCH_CONFLICT: LTP 10.1 > PWC 9.5 (ADD momentum candidate) but LTP close to SL (SL 10.0, Diff 0.99%) -> SELL overrides ADD!
            'SL': [9.0, 10.0],
            'LTP_SL_Diff_Pct': [0.20, 0.0099],
            'Invested_Value': [1000.0] * 2,
            'LTP': [10.2, 10.1],
            'Prev_Day_Close': [10.0] * 2,
            'Prev_Week_Close': [10.0, 9.5],
            'EMA9': [10.0] * 2,
            'EMA10': [10.0] * 2,
            'EMA11': [10.0] * 2,
            'EMA21': [10.0] * 2,
            'Current_Value': [1020.0, 1010.0],
            'Unrealized_PnL': [20.0, 10.0],
            'Return_Pct': [0.02, 0.01],
            'XIRR': [0.0] * 2,
            'Holding_Period': [150, 50],
            'Split_Info': [''] * 2,
            'Adj_Required': ['No'] * 2
        }
        df = pd.DataFrame(data)
        recs = generate_action_recommendations(df)

        self.assertFalse(recs.empty)
        self.assertEqual(len(recs), 2)
        # Verify both recommendations are SELL, and no ADD got generated for these conflicted symbols
        self.assertEqual(set(recs['Action (Recommended) ']), {'SELL'})
        
        recs_dict = recs.set_index('Stock ').to_dict('index')
        self.assertIn('CHEAT_CONFLICT', recs_dict)
        self.assertIn('TRANCH_CONFLICT', recs_dict)

    def test_write_recommendations_overwrite_append(self):
        """Verifies same-day overwrite and multi-date append in Action Tracker."""
        fd, path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        
        try:
            # Create a dummy Excel workbook with headers in 'Action Tracker'
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Action Tracker'
            for c_idx, name in enumerate(HEADERS, 1):
                ws.cell(row=1, column=c_idx, value=name)
                
            # Put an old recommendation in the sheet
            old_date = '20-05-2026'
            ws.cell(row=2, column=1, value=old_date)
            ws.cell(row=2, column=2, value='Satellite')
            ws.cell(row=2, column=3, value='OLDSTOCK')
            ws.cell(row=2, column=7, value='SELL')
            ws.cell(row=2, column=8, value='Old Reason')

            # Put a same-day recommendation in the sheet
            today_str = datetime.date.today().strftime('%d-%m-%Y')
            ws.cell(row=3, column=1, value=today_str)
            ws.cell(row=3, column=3, value='DUPLICATE')
            ws.cell(row=3, column=7, value='SELL')
            ws.cell(row=3, column=8, value='Should Be Replaced')

            wb.save(path)
            wb.close()

            # Now generate new recommendations for today: one SELL and one ADD
            new_recs = pd.DataFrame([
                {
                    'Date': today_str,
                    'TF_Classification': 'Satellite',
                    'Stock ': 'NEWSTOCK',
                    'Sector': 'Finance',
                    'TF Classification ': 'Satellite',
                    'Latest Tranch or Cheat': 'Cheat 1',
                    'Action (Recommended) ': 'SELL',
                    'Reason': 'New Reason',
                    'Action Taken ': '',
                    'Remarks': ''
                },
                {
                    'Date': today_str,
                    'TF_Classification': 'Satellite',
                    'Stock ': 'ADDSTOCK',
                    'Sector': 'Industrial',
                    'TF Classification ': 'Satellite',
                    'Latest Tranch or Cheat': 'Tranch 2',
                    'Action (Recommended) ': 'ADD',
                    'Reason': 'Add Reason',
                    'Action Taken ': '',
                    'Remarks': ''
                }
            ])

            # Write it
            write_recommendations_to_excel(path, new_recs)

            # Reload sheet and check
            wb = openpyxl.load_workbook(path)
            ws = wb['Action Tracker']

            # Total rows: 1 header row, 1 OLDSTOCK row, 2 new recommendation rows (NEWSTOCK, ADDSTOCK) = 4 rows total
            self.assertEqual(ws.max_row, 4)

            # Validate OLDSTOCK is still there
            self.assertEqual(ws.cell(row=2, column=1).value, old_date)
            self.assertEqual(ws.cell(row=2, column=3).value, 'OLDSTOCK')

            # Validate DUPLICATE is gone
            # Since the rows are sorted in combine/write, let's verify row-by-row
            rows_data = []
            for r in range(2, ws.max_row + 1):
                rows_data.append({
                    'Stock': ws.cell(row=r, column=3).value,
                    'Action': ws.cell(row=r, column=7).value,
                    'Reason': ws.cell(row=r, column=8).value,
                    'RowIndex': r
                })

            # OLDSTOCK should be first (unmodified old date row)
            self.assertEqual(rows_data[0]['Stock'], 'OLDSTOCK')

            # Find SELL and ADD rows
            sell_row_info = [r for r in rows_data if r['Stock'] == 'NEWSTOCK'][0]
            add_row_info = [r for r in rows_data if r['Stock'] == 'ADDSTOCK'][0]

            self.assertEqual(sell_row_info['Action'], 'SELL')
            self.assertEqual(add_row_info['Action'], 'ADD')

            # Check cell font and fill styles on SELL cell
            sell_cell = ws.cell(row=sell_row_info['RowIndex'], column=7)
            self.assertTrue(str(sell_cell.font.color.rgb).endswith('A93226'))
            self.assertTrue(str(sell_cell.fill.start_color.rgb).endswith('FADBD8'))

            # Check cell font and fill styles on ADD cell
            add_cell = ws.cell(row=add_row_info['RowIndex'], column=7)
            self.assertTrue(str(add_cell.font.color.rgb).endswith('375623'))
            self.assertTrue(str(add_cell.fill.start_color.rgb).endswith('E2EFDA'))

            wb.close()
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_generate_recommendations_grouping_and_sorting(self):
        """Verifies the exact grouping by SELL/ADD, then Cheat/Tranch, and Cheat sorted by holding period descending."""
        data = {
            'Symbol': ['CHEAT_S_SHORT', 'CHEAT_S_LONG', 'TRANCH_S', 'CHEAT_A_SHORT', 'CHEAT_A_LONG', 'TRANCH_A'],
            'Cap': ['Small Cap'] * 6,
            'TF_Sector': ['Industrial'] * 6,
            'TF_Classification': ['Satellite'] * 6,
            'Latest_Tranche': ['Cheat 1', 'Cheat 2', 'Tranch 1', 'Cheat 1', 'Cheat 2', 'Tranch 2'],
            'Current_Quantity': [100] * 6,
            'Average_Buy_Price': [10.0] * 6,
            'Trend': ['▲'] * 6,
            'SL': [9.0] * 6,
            'LTP_SL_Diff_Pct': [0.20, 0.20, 0.03, 0.20, 0.20, 0.20], # TRANCH_S is close to SL -> SELL
            'Invested_Value': [1000.0] * 6,
            'LTP': [10.0, 10.0, 9.2, 11.0, 11.0, 12.0],
            'Prev_Day_Close': [10.0] * 6,
            'Prev_Week_Close': [10.0] * 6,
            'EMA9': [10.0] * 6,
            'EMA10': [10.0] * 6,
            'EMA11': [10.0] * 6,
            'EMA21': [10.0] * 6,
            'Current_Value': [1000.0, 1000.0, 920.0, 1100.0, 1100.0, 1200.0],
            'Unrealized_PnL': [0.0, 0.0, -80.0, 100.0, 100.0, 200.0],
            # CHEAT_S_SHORT: Return 0% (SELL Cheat candidate, held 50 days)
            # CHEAT_S_LONG: Return 0% (SELL Cheat candidate, held 150 days)
            # TRANCH_S: LTP close to stop loss (< 5% difference) -> SELL Tranch
            # CHEAT_A_SHORT: Return +10% (ADD Cheat candidate, held 10 days)
            # CHEAT_A_LONG: Return +10% (ADD Cheat candidate, held 20 days)
            # TRANCH_A: LTP 12 > PWC 10 (ADD Tranch candidate)
            'Return_Pct': [0.0, 0.0, -0.08, 0.10, 0.10, 0.20],
            'XIRR': [0.0] * 6,
            'Holding_Period': [50, 150, 40, 10, 20, 30],
            'Split_Info': [''] * 6,
            'Adj_Required': ['No'] * 6
        }
        df = pd.DataFrame(data)
        recs = generate_action_recommendations(df)

        self.assertFalse(recs.empty)
        self.assertEqual(len(recs), 6)

        # Retrieve symbols in order returned
        recs_ordered = recs.to_dict('records')
        ordered_symbols = [r['Stock '] for r in recs_ordered]

        # Expected exact order:
        # 1. SELL + Cheat (150 days) -> CHEAT_S_LONG
        # 2. SELL + Cheat (50 days) -> CHEAT_S_SHORT
        # 3. SELL + Tranch -> TRANCH_S
        # 4. ADD + Cheat (20 days) -> CHEAT_A_LONG
        # 5. ADD + Cheat (10 days) -> CHEAT_A_SHORT
        # 6. ADD + Tranch -> TRANCH_A
        expected_symbols = [
            'CHEAT_S_LONG',
            'CHEAT_S_SHORT',
            'TRANCH_S',
            'CHEAT_A_LONG',
            'CHEAT_A_SHORT',
            'TRANCH_A'
        ]

        self.assertEqual(ordered_symbols, expected_symbols)

    @patch('os.path.exists')
    @patch('pandas.read_excel')
    @patch('yfinance.download')
    def test_generate_recommendations_buy_logic(self, mock_yf_download, mock_read_excel, mock_exists):
        """Rule 3: Enforces Watchlist BUY logic and computes RSI, ADX, and volume parameters."""
        mock_exists.return_value = True
        portfolio_df = pd.DataFrame(columns=['Symbol', 'TF_Classification'])
        
        watchlist_data = {
            'Stock': ['BUY_OK', 'BUY_FAIL_RSI', 'BUY_FAIL_ADX', 'BUY_FAIL_VOL'],
            'Color': ['Blue', 'Green', 'Blue', 'Green'],
            'Date': ['16-05-2026'] * 4
        }
        mock_read_excel.return_value = pd.DataFrame(watchlist_data)
        
        idx = pd.MultiIndex.from_product([
            ['Open', 'High', 'Low', 'Close', 'Volume'],
            ['BUY_OK.NS', 'BUY_FAIL_RSI.NS', 'BUY_FAIL_ADX.NS', 'BUY_FAIL_VOL.NS']
        ])
        
        dates = pd.date_range(end='2026-05-29', periods=53, freq='W')
        hist_df = pd.DataFrame(index=dates, columns=idx)
        n_periods = len(dates)
        
        for sym in ['BUY_OK.NS', 'BUY_FAIL_RSI.NS', 'BUY_FAIL_ADX.NS', 'BUY_FAIL_VOL.NS']:
            hist_df[('Open', sym)] = [10.0 + i*0.2 for i in range(n_periods)]
            hist_df[('High', sym)] = [10.5 + i*0.2 for i in range(n_periods)]
            hist_df[('Low', sym)] = [9.5 + i*0.2 for i in range(n_periods)]
            hist_df[('Close', sym)] = [10.2 + i*0.2 for i in range(n_periods)]
            hist_df[('Volume', sym)] = [10000] * n_periods
            
        hist_df[('Close', 'BUY_FAIL_RSI.NS')] = [100.0 - i*1.0 for i in range(n_periods)]
        
        hist_df[('High', 'BUY_FAIL_ADX.NS')] = [10.0] * n_periods
        hist_df[('Low', 'BUY_FAIL_ADX.NS')] = [10.0] * n_periods
        hist_df[('Close', 'BUY_FAIL_ADX.NS')] = [10.0] * n_periods
        
        hist_df[('Volume', 'BUY_FAIL_VOL.NS')] = [10000] * (n_periods - 1) + [1000]
        hist_df[('Volume', 'BUY_OK.NS')] = [10000] * (n_periods - 1) + [50000]
        
        mock_yf_download.return_value = hist_df
        
        recs = generate_action_recommendations(portfolio_df, filepath='dummy_path.xlsx')
        
        self.assertFalse(recs.empty)
        recs_dict = recs.set_index('Stock ').to_dict('index')
        self.assertIn('BUY_OK', recs_dict)
        self.assertEqual(recs_dict['BUY_OK']['Action (Recommended) '], 'BUY')
        self.assertNotIn('BUY_FAIL_RSI', recs_dict)
        self.assertNotIn('BUY_FAIL_ADX', recs_dict)
        self.assertNotIn('BUY_FAIL_VOL', recs_dict)

    def test_generate_recommendations_filters(self):
        """Filters test: Enforces filters by action type (all, add, buy, sell)."""
        data = {
            'Symbol': ['CONFLICT_1', 'CONFLICT_2'],
            'Cap': ['Small Cap'] * 2,
            'TF_Sector': ['Industrial'] * 2,
            'TF_Classification': ['Satellite', 'Satellite'],
            'Latest_Tranche': ['Cheat 1', 'Tranch 1'],
            'Current_Quantity': [100] * 2,
            'Average_Buy_Price': [10.0] * 2,
            'Trend': ['▲'] * 2,
            'SL': [9.0, 10.0],
            'LTP_SL_Diff_Pct': [0.20, 0.0099],
            'Invested_Value': [1000.0] * 2,
            'LTP': [10.2, 10.1],
            'Prev_Day_Close': [10.0] * 2,
            'Prev_Week_Close': [10.0, 9.5],
            'EMA9': [10.0] * 2,
            'EMA10': [10.0] * 2,
            'EMA11': [10.0] * 2,
            'EMA21': [10.0] * 2,
            'Current_Value': [1020.0, 1010.0],
            'Unrealized_PnL': [20.0, 10.0],
            'Return_Pct': [0.02, 0.01],
            'XIRR': [0.0] * 2,
            'Holding_Period': [150, 60],
            'Split_Info': [''] * 2,
            'Adj_Required': ['No'] * 2
        }
        df = pd.DataFrame(data)
        
        # Filter all: should have both SELL rows
        recs_all = generate_action_recommendations(df, rec_filter='all')
        self.assertEqual(len(recs_all), 2)
        self.assertEqual(set(recs_all['Action (Recommended) ']), {'SELL'})
        
        # Filter sell: should have both SELL rows
        recs_sell = generate_action_recommendations(df, rec_filter='sell')
        self.assertEqual(len(recs_sell), 2)
        
        # Filter add: should have 0 rows (since both conflicts resolved to SELL)
        recs_add = generate_action_recommendations(df, rec_filter='add')
        self.assertEqual(len(recs_add), 0)

    def test_write_recommendations_preserves_other_actions(self):
        """Preserves other action types for the same date when filtering is active."""
        import tempfile
        import shutil

        # Set up a temp file
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "test_wb.xlsx")
            
            # Create a blank workbook with Action Tracker
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Action Tracker"
            
            # Write headers
            for c_idx, col_name in enumerate(HEADERS, 1):
                ws.cell(row=1, column=c_idx, value=col_name)
                
            # Write two initial actions for today: one SELL and one ADD
            today_str = datetime.date.today().strftime('%d-%m-%Y')
            
            ws.cell(row=2, column=1, value=today_str)
            ws.cell(row=2, column=2, value='Satellite')
            ws.cell(row=2, column=3, value='OLDSELL')
            ws.cell(row=2, column=5, value='Satellite')
            ws.cell(row=2, column=6, value='Tranch 1')
            ws.cell(row=2, column=7, value='SELL')
            ws.cell(row=2, column=8, value='Old Sell Reason')
            
            ws.cell(row=3, column=1, value=today_str)
            ws.cell(row=3, column=2, value='Satellite')
            ws.cell(row=3, column=3, value='OLDADD')
            ws.cell(row=3, column=5, value='Satellite')
            ws.cell(row=3, column=6, value='Tranch 2')
            ws.cell(row=3, column=7, value='ADD')
            ws.cell(row=3, column=8, value='Old Add Reason')
            
            wb.save(path)
            wb.close()
            
            # 1. Now write a new BUY action using rec_filter='buy'
            new_recs_buy = pd.DataFrame([
                {
                    'Date': today_str,
                    'TF_Classification': 'Satellite',
                    'Stock ': 'NEWBUY',
                    'Sector': 'Energy',
                    'TF Classification ': 'Satellite',
                    'Latest Tranch or Cheat': 'Watchlist',
                    'Action (Recommended) ': 'BUY',
                    'Reason': 'New Buy Reason',
                    'Action Taken ': '',
                    'Remarks': ''
                }
            ])
            
            # Force standard openpyxl by patching/mocking xlwings or letting it fall back
            # (or just call standard openpyxl since we are in unit tests and might not have Excel running)
            with patch('sys.platform', 'linux'): # Force openpyxl fallback
                write_recommendations_to_excel(path, new_recs_buy, rec_filter='buy')
                
            # Reload sheet and check
            wb = openpyxl.load_workbook(path)
            ws = wb['Action Tracker']
            
            # Verify we have 3 recommendations now: OLDSELL, OLDADD, and NEWBUY
            self.assertEqual(ws.max_row, 4) # 1 header + 3 data rows
            
            stocks = [ws.cell(row=r, column=3).value for r in range(2, 5)]
            actions = [ws.cell(row=r, column=7).value for r in range(2, 5)]
            
            self.assertIn('OLDSELL', stocks)
            self.assertIn('OLDADD', stocks)
            self.assertIn('NEWBUY', stocks)
            
            wb.close()
            
            # 2. Now overwrite/update the ADD action using rec_filter='add'
            new_recs_add = pd.DataFrame([
                {
                    'Date': today_str,
                    'TF_Classification': 'Satellite',
                    'Stock ': 'NEWADD',
                    'Sector': 'Tech',
                    'TF Classification ': 'Satellite',
                    'Latest Tranch or Cheat': 'Tranch 2',
                    'Action (Recommended) ': 'ADD',
                    'Reason': 'New Add Reason',
                    'Action Taken ': '',
                    'Remarks': ''
                }
            ])
            
            with patch('sys.platform', 'linux'):
                write_recommendations_to_excel(path, new_recs_add, rec_filter='add')
                
            # Reload sheet and check
            wb = openpyxl.load_workbook(path)
            ws = wb['Action Tracker']
            
            # Verify we have 3 recommendations: OLDSELL (preserved), NEWBUY (preserved), and NEWADD (overwrote OLDADD)
            self.assertEqual(ws.max_row, 4)
            
            stocks_after = [ws.cell(row=r, column=3).value for r in range(2, 5)]
            actions_after = [ws.cell(row=r, column=7).value for r in range(2, 5)]
            
            self.assertIn('OLDSELL', stocks_after)
            self.assertIn('NEWBUY', stocks_after)
            self.assertIn('NEWADD', stocks_after)
            self.assertNotIn('OLDADD', stocks_after)
            
            wb.close()

    @patch('os.path.exists')
    @patch('yfinance.download')
    @patch('src.actions.pd.read_excel')
    def test_generate_recommendations_buy_latest_date_only(self, mock_read_excel, mock_yf_download, mock_exists):
        """Only stocks on the absolute latest date should be evaluated as BUY candidates."""
        mock_exists.return_value = True
        portfolio_df = pd.DataFrame(columns=['Symbol', 'TF_Classification'])
        
        # Mock watchlists:
        # 'OLD_STOCK' on 10-05-2026 (an older date)
        # 'NEW_STOCK' on 20-05-2026 (the latest date)
        watchlist_data = {
            'Stock': ['OLD_STOCK', 'NEW_STOCK'],
            'Color': ['Blue', 'Blue'],
            'Date': ['10-05-2026', '20-05-2026']
        }
        mock_read_excel.return_value = pd.DataFrame(watchlist_data)
        
        # Mock historical data for both so we don't crash
        idx = pd.MultiIndex.from_product([
            ['Open', 'High', 'Low', 'Close', 'Volume'],
            ['NEW_STOCK.NS'] # We only want yfinance to be called for NEW_STOCK!
        ])
        dates = pd.date_range(end='2026-05-29', periods=53, freq='W')
        hist_df = pd.DataFrame(index=dates, columns=idx)
        n_periods = len(dates)
        
        hist_df[('Open', 'NEW_STOCK.NS')] = [10.0 + i*0.2 for i in range(n_periods)]
        hist_df[('High', 'NEW_STOCK.NS')] = [10.5 + i*0.2 for i in range(n_periods)]
        hist_df[('Low', 'NEW_STOCK.NS')] = [9.5 + i*0.2 for i in range(n_periods)]
        hist_df[('Close', 'NEW_STOCK.NS')] = [10.2 + i*0.2 for i in range(n_periods)]
        hist_df[('Volume', 'NEW_STOCK.NS')] = [10000] * (n_periods - 1) + [50000] # Breakout volume
        
        mock_yf_download.return_value = hist_df
        
        recs = generate_action_recommendations(portfolio_df, filepath='dummy_path.xlsx')
        
        # Verify that only NEW_STOCK gets processed and OLD_STOCK is ignored
        recs_dict = recs.set_index('Stock ').to_dict('index')
        self.assertIn('NEW_STOCK', recs_dict)
        self.assertNotIn('OLD_STOCK', recs_dict)
        
        # Verify yf.download was called with NEW_STOCK.NS but NOT OLD_STOCK.NS
        called_args = mock_yf_download.call_args[0][0]
        self.assertIn('NEW_STOCK.NS', called_args)
        self.assertNotIn('OLD_STOCK.NS', called_args)


if __name__ == '__main__':
    unittest.main()
