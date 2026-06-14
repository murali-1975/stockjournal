"""
Tests for src/excel_writer.py
==============================

Validates that the Excel writer:
    - Creates all expected sheets
    - Applies INR currency formatting to monetary columns
    - Applies percentage formatting to PnL % columns
    - Converts large ID numbers to native integers (no scientific notation)
"""

import os
import tempfile
import unittest

import pandas as pd
from openpyxl import load_workbook

from src.excel_writer import save_workbook


class TestSaveWorkbook(unittest.TestCase):
    """Test suite for the save_workbook function."""

    def setUp(self):
        """Create sample DataFrames and a local output path to prevent Excel Protected View."""
        # Generate a unique local output filename based on the test name
        test_name = self.id().split('.')[-1]
        self.output_path = os.path.abspath(f'test_output_{test_name}.xlsx')
        if os.path.exists(self.output_path):
            try:
                os.remove(self.output_path)
            except Exception:
                pass

        self.raw_df = pd.DataFrame({
            'Trade Date': ['2025-01-01'],
            'Symbol': ['RELIANCE'],
            'Trade Type': ['buy'],
            'Quantity': [10],
            'Price': [2500.0],
            'Order ID': ['1200000047930385'],
            'Trade ID': ['405271862'],
        })

        self.grouped_df = pd.DataFrame({
            'Trade Date': ['2025-01-01'],
            'Symbol': ['RELIANCE'],
            'Trade Type': ['buy'],
            'Total_Quantity': [10],
            'Average_Price': [2500.0],
            'Total_Value': [25000.0],
            'Tranches/Cheat': ['Tranch 1'],
        })

        self.portfolio_df = pd.DataFrame({
            'Symbol': ['RELIANCE'],
            'Cap': ['Large Cap'],
            'TF_Sector': ['Technology'],
            'TF_Classification': ['Core'],
            'Latest_Tranche': ['Tranch 1'],
            'Current_Quantity': [10],
            'Average_Buy_Price': [2500.0],
            'SL': [2250.0],
            'Invested_Value': [25000.0],
            'LTP': [2600.0],
            'EMA9': [2590.0],
            'EMA10': [2585.0],
            'EMA11': [2580.0],
            'EMA21': [2550.0],
            'Current_Value': [26000.0],
            'Unrealized_PnL': [1000.0],
            'Holding_Period': [30],
        })

        self.overall_df = pd.DataFrame({
            'Symbol': ['RELIANCE'],
            'Cap': ['Large Cap'],
            'TF_Sector': ['Technology'],
            'TF_Classification': ['Core'],
            'Latest_Tranche': ['Tranch 1'],
            'Total_Buy_Quantity': [10],
            'Total_Buy_Value': [25000.0],
            'Average_Buy_Price': [2500.0],
            'Total_Sell_Quantity': [0],
            'Total_Sell_Value': [0.0],
            'Average_Sell_Price': [0.0],
            'Current_Quantity': [10],
            'Invested_Value': [25000.0],
            'LTP': [2600.0],
            'Current_Value': [26000.0],
            'Realized_PnL': [0.0],
            'Unrealized_PnL': [1000.0],
            'Total_PnL': [1000.0],
            'Total_PnL_Percentage': [0.04],
            'Holding_Period': [30],
        })

    def tearDown(self):
        """Clean up the test workbook."""
        if os.path.exists(self.output_path):
            try:
                os.remove(self.output_path)
            except Exception:
                pass

    def test_sheets_created(self):
        """All four expected sheets should be created in the workbook."""
        save_workbook(self.raw_df, self.grouped_df, self.portfolio_df, self.overall_df, self.output_path)
        wb = load_workbook(self.output_path)
        expected_sheets = {'Raw_Tradebook', 'Transaction', 'Current_Portfolio', 'Overall_Portfolio'}
        self.assertTrue(expected_sheets.issubset(set(wb.sheetnames)))
        wb.close()

    def test_id_column_is_integer(self):
        """Order ID cells should contain native integers, not strings."""
        save_workbook(self.raw_df, self.grouped_df, self.portfolio_df, self.overall_df, self.output_path)
        wb = load_workbook(self.output_path)
        ws = wb['Raw_Tradebook']

        # Find Order ID column index
        header_row = [cell.value for cell in ws[1]]
        oid_col = header_row.index('Order ID') + 1

        cell_value = ws.cell(row=2, column=oid_col).value
        self.assertIsInstance(cell_value, int)
        self.assertEqual(cell_value, 1200000047930385)
        wb.close()

    def test_currency_format_applied(self):
        """Monetary columns in Transaction sheet should have INR currency format."""
        save_workbook(self.raw_df, self.grouped_df, self.portfolio_df, self.overall_df, self.output_path)
        wb = load_workbook(self.output_path)
        ws = wb['Transaction']

        header_row = [cell.value for cell in ws[1]]
        tv_col = header_row.index('Total_Value') + 1

        fmt = ws.cell(row=2, column=tv_col).number_format
        self.assertIn('₹', fmt)
        wb.close()

    def test_percentage_format_applied(self):
        """Total_PnL_Percentage in Overall_Portfolio should have percentage format."""
        save_workbook(self.raw_df, self.grouped_df, self.portfolio_df, self.overall_df, self.output_path)
        wb = load_workbook(self.output_path)
        ws = wb['Overall_Portfolio']

        header_row = [cell.value for cell in ws[1]]
        pct_col = header_row.index('Total_PnL_Percentage') + 1

        fmt = ws.cell(row=2, column=pct_col).number_format
        self.assertIn('%', fmt)
        wb.close()

    def test_append_mode_preserves_custom_sheets(self):
        """Running save_workbook twice should not destroy custom user sheets."""
        # First run: create the workbook
        save_workbook(self.raw_df, self.grouped_df, self.portfolio_df, self.overall_df, self.output_path)

        # Manually add a custom sheet
        wb = load_workbook(self.output_path)
        wb.create_sheet('My_Notes')
        wb['My_Notes']['A1'] = 'User custom data'
        wb.save(self.output_path)
        wb.close()

        # Second run: should replace data sheets but keep My_Notes
        save_workbook(self.raw_df, self.grouped_df, self.portfolio_df, self.overall_df, self.output_path)

        wb = load_workbook(self.output_path)
        self.assertIn('My_Notes', wb.sheetnames)
        self.assertEqual(wb['My_Notes']['A1'].value, 'User custom data')
        wb.close()

    def test_core_watchlist_formatting(self):
        """Core stocks in Current_Portfolio and Dashboard should be formatted based on Core_Watchlist latest month trends strictly."""
        import os
        from openpyxl import Workbook
        
        # Add TCS (Core) to portfolio_df to test that its background is cleared if it is only in the past month
        tcs_row = pd.Series({
            'Symbol': 'TCS',
            'Cap': 'Large Cap',
            'TF_Sector': 'Technology',
            'TF_Classification': 'Core',
            'Latest_Tranche': 'Tranch 1',
            'Current_Quantity': 10,
            'Average_Buy_Price': 3000.0,
            'SL': 2700.0,
            'Invested_Value': 30000.0,
            'LTP': 3100.0,
            'EMA9': 3090.0,
            'EMA10': 3085.0,
            'EMA11': 3080.0,
            'EMA21': 3050.0,
            'Current_Value': 31000.0,
            'Unrealized_PnL': 1000.0,
            'Holding_Period': 30,
        })
        portfolio_test_df = pd.concat([self.portfolio_df, pd.DataFrame([tcs_row])], ignore_index=True)
        
        # Create empty workbook with Core_Watchlist manually containing multiple months
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        ws_cw = wb.create_sheet('Core_Watchlist')
        ws_cw.append(['Month', 'Company', 'Company Name', 'Sector', 'Theme', 'Market Cap (MC)', 'Trend Status'])
        # RELIANCE is in the latest month
        ws_cw.append(['2026-05-01', 'RELIANCE', 'Reliance Industries', 'Energy', 'Industrial', 'Large Cap', 'Strong Trend- Green'])
        # TCS is only in a past month (should be filtered out/not colored)
        ws_cw.append(['2026-04-01', 'TCS', 'Tata Consultancy Services', 'Technology', 'Industrial', 'Large Cap', 'Medium Trend'])
        wb.save(self.output_path)
        wb.close()
        
        # Run save_workbook which appends/overwrites data sheets and builds Dashboard
        save_workbook(self.raw_df, self.grouped_df, portfolio_test_df, self.overall_df, self.output_path)
        
        # 1. Read back and verify styling on 'Current_Portfolio' sheet
        wb = load_workbook(self.output_path)
        ws_cp = wb['Current_Portfolio']
        
        header = [cell.value for cell in ws_cp[1]]
        symbol_col_idx = header.index('Symbol') + 1
        
        # Find row for RELIANCE and TCS in Current_Portfolio
        reliance_row, tcs_row_idx = None, None
        for r in range(2, ws_cp.max_row + 1):
            val = ws_cp.cell(row=r, column=symbol_col_idx).value
            if val == 'RELIANCE':
                reliance_row = r
            elif val == 'TCS':
                tcs_row_idx = r
                
        self.assertIsNotNone(reliance_row)
        self.assertIsNotNone(tcs_row_idx)
        
        reliance_cell = ws_cp.cell(row=reliance_row, column=symbol_col_idx)
        tcs_cell = ws_cp.cell(row=tcs_row_idx, column=symbol_col_idx)
        
        # RELIANCE should be green
        self.assertIn(reliance_cell.fill.start_color.rgb, ['0000B050', 'FF00B050'])
        self.assertTrue(reliance_cell.font.bold)
        
        # TCS should be cleared (no background fill)
        self.assertIn(tcs_cell.fill.fill_type, [None, 'none'])
        self.assertFalse(tcs_cell.font.bold)
        
        # 2. Read back and verify formula insertion on 'Dashboard' sheet
        self.assertIn('Dashboard', wb.sheetnames)
        ws_db = wb['Dashboard']
        
        # Scan Dashboard sheet Column A for the headers
        sat_tb_row = None
        core_tb_row = None
        sat_dl_row = None
        core_dl_row = None
        for r in range(1, ws_db.max_row + 1):
            val = ws_db.cell(row=r, column=1).value
            if val == 'Top 5 Gainers / Bottom 5 Losers (Satellite)':
                sat_tb_row = r
            elif val == 'Top 5 Gainers / Bottom 5 Losers (Core)':
                core_tb_row = r
            elif val == 'Top 10 underperforming Satellite Stocks (LTP < Prev Day Close)':
                sat_dl_row = r
            elif val == 'Top 10 underperforming Core Stocks (LTP < Prev Day Close)':
                core_dl_row = r
                
        self.assertIsNotNone(sat_tb_row, "Satellite table header not found on Dashboard")
        self.assertIsNotNone(core_tb_row, "Core table header not found on Dashboard")
        self.assertIsNotNone(sat_dl_row, "Satellite daily losers header not found on Dashboard")
        self.assertIsNotNone(core_dl_row, "Core daily losers header not found on Dashboard")
        
        # Check formula in the first data cell under headers
        sat_formula = ws_db.cell(row=sat_tb_row + 2, column=1).value
        core_formula = ws_db.cell(row=core_tb_row + 2, column=1).value
        sat_dl_formula = ws_db.cell(row=sat_dl_row + 2, column=1).value
        core_dl_formula = ws_db.cell(row=core_dl_row + 2, column=1).value
        
        self.assertIsNotNone(sat_formula)
        self.assertIsNotNone(core_formula)
        self.assertIsNotNone(sat_dl_formula)
        self.assertIsNotNone(core_dl_formula)
        
        self.assertTrue(sat_formula.startswith('='))
        if 'LET' in sat_formula:
            self.assertIn('Current_Portfolio', sat_formula)
        else:
            self.assertIn('Gainers/Losers (Satellite)', sat_formula)
            
        self.assertTrue(core_formula.startswith('='))
        if 'LET' in core_formula:
            self.assertIn('Current_Portfolio', core_formula)
        else:
            self.assertIn('Gainers/Losers (Core)', core_formula)

        self.assertTrue(sat_dl_formula.startswith('='))
        if 'LET' in sat_dl_formula:
            self.assertIn('Current_Portfolio', sat_dl_formula)
        else:
            self.assertIn('Daily Losers (Satellite)', sat_dl_formula)

        self.assertTrue(core_dl_formula.startswith('='))
        if 'LET' in core_dl_formula:
            self.assertIn('Current_Portfolio', core_dl_formula)
        else:
            self.assertIn('Daily Losers (Core)', core_dl_formula)
        
        wb.close()


if __name__ == '__main__':
    unittest.main()
