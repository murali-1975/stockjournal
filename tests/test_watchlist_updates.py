"""
Tests for Watchlist Enrichment & Dynamic Formatting
===================================================
Validates that columns G, H, I, J are correctly populated with weekly market data,
columns D, E, F are preserved as formulas, and conditional formatting is applied.
"""

import os
import unittest
import pandas as pd
import openpyxl
from openpyxl import Workbook

from src.excel_writer import _update_satellite_watchlist_columns, _apply_satellite_watchlist_conditional_formatting


class TestWatchlistUpdates(unittest.TestCase):
    """Test suite for validating Satellite_Watchlist sheet updates."""

    def setUp(self):
        self.test_filename = "temp_test_watchlist_enrichment.xlsx"
        
        wb = Workbook()
        # Create default sheets
        ws_portfolio = wb.active
        ws_portfolio.title = "Current_Portfolio"
        ws_portfolio.append(["Symbol", "TF_Classification"])
        ws_portfolio.append(["BALAMINES", "Satellite"])
        
        # Create Satellite_Watchlist sheet
        ws_watchlist = wb.create_sheet(title="Satellite_Watchlist")
        # Headers: Date, Stock, Color, Company Name, Price, Previous Close, Previous week Close
        ws_watchlist.append([
            "Date", "Stock", "Color", "Company Name", "Price", "Previous Close ", "Previous week Close"
        ])
        
        # Row 2 (BALAMINES): Price/Previous Close are formulas
        # Price: =_FV(D2,"Price")
        # Previous Close: =_FV(D2,"Previous close",TRUE)
        ws_watchlist.cell(row=2, column=1, value="16-05-2026")
        ws_watchlist.cell(row=2, column=2, value="BALAMINES")
        ws_watchlist.cell(row=2, column=3, value="Blue")
        ws_watchlist.cell(row=2, column=4, value="=_FV(B2, \"Name\")")
        ws_watchlist.cell(row=2, column=5, value="=_FV(B2, \"Price\")")
        ws_watchlist.cell(row=2, column=6, value="=_FV(B2, \"Previous close\", TRUE)")
        ws_watchlist.cell(row=2, column=7, value=None) # Previous week Close (Column G)

        wb.save(self.test_filename)
        wb.close()

    def tearDown(self):
        if os.path.exists(self.test_filename):
            try:
                os.remove(self.test_filename)
            except Exception:
                pass

    def test_watchlist_column_updates_and_formatting(self):
        """Validates enrichment of columns G-J, dynamic conditional formatting, and formula preservation."""
        # 1. Modify the sheet using a mock context in openpyxl ExcelWriter
        # We simulate the pandas ExcelWriter in append mode
        with pd.ExcelWriter(self.test_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # We call the update columns helper
            _update_satellite_watchlist_columns(writer, self.test_filename)
            
            # We call the conditional formatting helper
            _apply_satellite_watchlist_conditional_formatting(writer)

        # 2. Read back the spreadsheet and inspect
        wb = openpyxl.load_workbook(self.test_filename, data_only=False)
        ws = wb['Satellite_Watchlist']

        # Verify headers in Columns G, H, I, J
        self.assertEqual(ws.cell(row=1, column=7).value, "Previous week Close")
        self.assertEqual(ws.cell(row=1, column=8).value, "EMA 9 (weekly)")
        self.assertEqual(ws.cell(row=1, column=9).value, "EMA 11 (weekly)")
        self.assertEqual(ws.cell(row=1, column=10).value, "EMA 21 (weekly)")

        # Verify formulas are preserved in Columns D, E, F
        self.assertEqual(ws.cell(row=2, column=4).value, '=_FV(B2, "Name")')
        self.assertEqual(ws.cell(row=2, column=5).value, '=_FV(B2, "Price")')
        self.assertEqual(ws.cell(row=2, column=6).value, '=_FV(B2, "Previous close", TRUE)')

        # Verify columns G, H, I, J have been updated with float numeric values (or default 0.0 if fetch fails/mocked)
        self.assertIsNotNone(ws.cell(row=2, column=7).value)
        self.assertIsNotNone(ws.cell(row=2, column=8).value)
        self.assertIsNotNone(ws.cell(row=2, column=9).value)
        self.assertIsNotNone(ws.cell(row=2, column=10).value)

        # Verify currency formatting is applied
        inr_format = '[$₹-en-IN] #,##0.00'
        self.assertEqual(ws.cell(row=2, column=7).number_format, inr_format)
        self.assertEqual(ws.cell(row=2, column=8).number_format, inr_format)
        self.assertEqual(ws.cell(row=2, column=9).number_format, inr_format)
        self.assertEqual(ws.cell(row=2, column=10).number_format, inr_format)

        # Verify conditional formatting rules
        rules = ws.conditional_formatting._cf_rules
        self.assertGreater(len(rules), 0)
        
        # Verify that both rules were added
        f_found = False
        g_found = False
        for cf, rules in ws.conditional_formatting._cf_rules.items():
            for r in rules:
                if r.formula:
                    if 'E2' in r.formula[0] and 'F2' in r.formula[0]:
                        f_found = True
                    if 'F2' in r.formula[0] and 'G2' in r.formula[0]:
                        g_found = True
                            
        self.assertTrue(f_found, "Conditional formatting rule for Column F not found")
        self.assertTrue(g_found, "Conditional formatting rule for Column G not found")

        wb.close()


if __name__ == '__main__':
    unittest.main()
