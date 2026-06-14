"""
Tests for src/dashboard.py
==========================

Validates the generation and formatting of the Excel Dashboard Worksheet.
"""

import unittest
import pandas as pd
import openpyxl

from src.dashboard import create_dashboard


class TestDashboard(unittest.TestCase):
    """Test suite for the create_dashboard function."""

    def setUp(self):
        """Create sample Portfolio and Overall DataFrames for testing."""
        self.wb = openpyxl.Workbook()
        
        # Current Portfolio Mock
        self.portfolio_df = pd.DataFrame({
            'Symbol': ['TCS', 'INFY'],
            'Cap': ['Large Cap', 'Large Cap'],
            'TF_Sector': ['IT', 'IT'],
            'TF_Classification': ['Core', 'Satellite'],
            'Latest_Tranche': ['Tranch 1', 'Cheat 1'],
            'Current_Quantity': [10.0, 5.0],
            'Average_Buy_Price': [3000.0, 1500.0],
            'Invested_Value': [30000.0, 7500.0],
            'LTP': [3200.0, 1600.0],
            'EMA9': [3150.0, 1580.0],
            'EMA10': [3140.0, 1570.0],
            'EMA11': [3130.0, 1560.0],
            'EMA21': [3100.0, 1500.0],
            'Current_Value': [32000.0, 8000.0],
            'Unrealized_PnL': [2000.0, 500.0],
            'Holding_Period': [30, 45],
            'Split_Info': ['', ''],
            'Adj_Required': ['No', 'No'],
            'SL': [2900.0, 1400.0]
        })
        # Add derived columns needed for dashboard logic
        self.portfolio_df['LTP_SL_Diff'] = self.portfolio_df['LTP'] - self.portfolio_df['SL']
        self.portfolio_df['LTP_SL_Diff_Pct'] = self.portfolio_df['LTP_SL_Diff'] / self.portfolio_df['SL']

        # Overall Portfolio Mock
        self.overall_df = pd.DataFrame({
            'Symbol': ['TCS', 'INFY', 'WIPRO'],
            'Cap': ['Large Cap', 'Large Cap', 'Large Cap'],
            'TF_Sector': ['IT', 'IT', 'IT'],
            'Realized_PnL': [1000.0, -500.0, 200.0],
            'Unrealized_PnL': [2000.0, 500.0, 0.0],
            'Split_Info': ['', '', '2:1 Split on 2024-01-01'],
            'Adj_Required': ['No', 'No', 'Yes']
        })

    def test_create_dashboard(self):
        """Dashboard should render without exceptions and create the required sheet."""
        
        # Add a mock watchlist dataframe
        watchlist_df = pd.DataFrame({
            'Stock': ['INFY', 'TCS', 'NEWSTOCK'],
            'Color': ['BLUE', 'GREEN', 'RED'],
            'Date': ['01-05-2026', '02-05-2026', '03-05-2026'],
            'Price': [1700.0, 3300.0, 500.0],
            'Previous week Close': [1600.0, 3200.0, 450.0]
        })
        
        # Execute the main function with the watchlist
        create_dashboard(self.wb, self.portfolio_df, self.overall_df, watchlist_df=watchlist_df)
        
        # Verify the sheet was created
        self.assertIn('Dashboard', self.wb.sheetnames)
        
        ws = self.wb['Dashboard']
        
        # Verify title is correct
        title_cell_value = ws['A1'].value
        self.assertEqual(title_cell_value, '📊 Portfolio Dashboard')
        
        # Verify column widths are set
        self.assertEqual(ws.column_dimensions['A'].width, 24)
        
        # We don't thoroughly test every cell location because formatting can change often,
        # but the function shouldn't raise any errors with sample data.

    def test_create_dashboard_empty_data(self):
        """Dashboard should gracefully handle empty DataFrames."""
        # Execute with empty data (and None for watchlist_df)
        create_dashboard(self.wb, pd.DataFrame(), pd.DataFrame())
        
        self.assertIn('Dashboard', self.wb.sheetnames)
        ws = self.wb['Dashboard']
        self.assertEqual(ws['A1'].value, '📊 Portfolio Dashboard')

    def test_corporate_actions_filter(self):
        """Corporate actions should only include current portfolio stocks."""
        # Add a split to a current portfolio stock to see if it's picked up
        self.portfolio_df.loc[0, 'Split_Info'] = '1:1 Bonus'
        create_dashboard(self.wb, self.portfolio_df, self.overall_df)
        ws = self.wb['Dashboard']
        
        # Check all cells to ensure WIPRO (which has a split but is only in overall_df) is NOT in the dashboard
        found_wipro = False
        found_tcs_split = False
        for row in ws.iter_rows(values_only=True):
            if row:
                if 'WIPRO' in row:
                    found_wipro = True
                if '1:1 Bonus' in row:
                    found_tcs_split = True
        
        self.assertFalse(found_wipro, "WIPRO from overall_df should not appear in corporate actions.")
        self.assertTrue(found_tcs_split, "TCS split from current portfolio should appear.")

    def test_stop_loss_excludes_core(self):
        """Stop loss dashboard should exclude Core stocks."""
        create_dashboard(self.wb, self.portfolio_df, self.overall_df)
        ws = self.wb['Dashboard']
        
        # The stop loss section is titled '⚠️ Stocks Nearest to Stop Loss'
        sl_section_started = False
        core_stock_found = False
        satellite_stock_found = False
        
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == '⚠️ Stocks Nearest to Stop Loss':
                sl_section_started = True
                
            if sl_section_started and row and row[0]:
                if row[0] == 'TCS': # TCS is Core
                    core_stock_found = True
                if row[0] == 'INFY': # INFY is Satellite
                    satellite_stock_found = True
                    
        self.assertFalse(core_stock_found, "TCS (Core) should not be in the Stop Loss table.")
        self.assertTrue(satellite_stock_found, "INFY (Satellite) should be in the Stop Loss table.")

    def test_top_bottom_split(self):
        """Top 5/Bottom 5 should be split into Core and Satellite tables."""
        create_dashboard(self.wb, self.portfolio_df, self.overall_df)
        ws = self.wb['Dashboard']
        
        satellite_header_found = False
        core_header_found = False
        
        in_satellite_section = False
        in_core_section = False
        
        for row in ws.iter_rows(values_only=True):
            # Check for headers
            if row and row[0] == 'Top 5 Gainers / Bottom 5 Losers (Satellite)':
                satellite_header_found = True
                in_satellite_section = True
                in_core_section = False
            elif row and row[0] == 'Top 5 Gainers / Bottom 5 Losers (Core)':
                core_header_found = True
                in_core_section = True
                in_satellite_section = False
            elif row and row[0] and isinstance(row[0], str) and ('Nearest to Stop Loss' in row[0] or 'Corporate Actions' in row[0] or 'Top 5 Cheat' in row[0]):
                in_satellite_section = False
                in_core_section = False
                
            # Check data rows
            if row and len(row) >= 2:
                classification = row[1]
                if classification == 'Satellite':
                    self.assertFalse(in_core_section, "Satellite stock found in Core section")
                elif classification == 'Core':
                    self.assertFalse(in_satellite_section, "Core stock found in Satellite section")
                    
        self.assertTrue(satellite_header_found, "Satellite Top/Bottom table header missing")
        self.assertTrue(core_header_found, "Core Top/Bottom table header missing")

    def test_movers_tables(self):
        """Test that movers tables use Prev_Week_Close / Previous week Close columns correctly."""
        # Add Prev_Week_Close to portfolio_df
        self.portfolio_df['Prev_Week_Close'] = [3100.0, 1500.0]
        self.portfolio_df['Prior_Week_Close'] = [3050.0, 1450.0]
        
        watchlist_df = pd.DataFrame({
            'Stock': ['INFY', 'TCS', 'NEWSTOCK'],
            'Color': ['BLUE', 'GREEN', 'RED'],
            'Date': ['01-05-2026', '02-05-2026', '03-05-2026'],
            'Price': [1700.0, 3300.0, 500.0],
            'Previous week Close': [1600.0, 3200.0, 450.0]
        })
        
        create_dashboard(self.wb, self.portfolio_df, self.overall_df, watchlist_df=watchlist_df)
        ws = self.wb['Dashboard']
        
        found_port_movers = False
        found_watchlist_movers = False
        
        for row in ws.iter_rows(values_only=True):
            if row:
                if 'Top 10 Movers (Current Portfolio)' in row:
                    found_port_movers = True
                if 'Top 10 Movers (Watchlist Only)' in row:
                    found_watchlist_movers = True
                for val in row:
                    if val and isinstance(val, str) and val.startswith('Error:'):
                        self.fail(f"Movers table has error: {val}")
                    
        self.assertTrue(found_port_movers)
        self.assertTrue(found_watchlist_movers)

    def test_performance_column(self):
        """Test that the Performance column is correctly generated in the Advancing/Declining table."""
        # Execute the main function
        create_dashboard(self.wb, self.portfolio_df, self.overall_df)
        ws = self.wb['Dashboard']
        
        # Scan Dashboard sheet for the table header row
        header_row_idx = None
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=4).value # col_start is 4 (Column D)
            if val == 'TF Classfication':
                header_row_idx = r
                break
                
        self.assertIsNotNone(header_row_idx, "Advancing/Declining table header row not found")
        
        # Verify the headers
        headers = [ws.cell(row=header_row_idx, column=c).value for c in range(4, 8)]
        self.assertEqual(headers, ['TF Classfication', 'Advancing', 'Declining', 'Performance'])
        
        # Verify the 4 sub-metrics data rows
        expected_rows = [
            ('Core (Previous month close)', 'AA'),
            ('Satellite (Previous week close)', 'O'),
            ('Core (Previous Close)', 'N'),
            ('Satellite (Previous Close)', 'N')
        ]
        
        for idx, (label, baseline_col) in enumerate(expected_rows):
            r = header_row_idx + 1 + idx
            
            # Check label in column D (4)
            self.assertEqual(ws.cell(row=r, column=4).value, label)
            
            # Check formula in column G (7)
            perf_formula = ws.cell(row=r, column=7).value
            self.assertIsNotNone(perf_formula)
            self.assertTrue(perf_formula.startswith('='))
            self.assertIn('LET(curr,', perf_formula)
            self.assertIn(f'${baseline_col}$1000', perf_formula) # should reference baseline close column
            
            # Check format in column G (7)
            self.assertEqual(ws.cell(row=r, column=7).number_format, '0.00%;[Red]-0.00%')

    def test_classification_allocation_columns(self):
        """Test that Core & Satellite Distribution table includes Realized and Un-Realized columns with correct formulas."""
        create_dashboard(self.wb, self.portfolio_df, self.overall_df)
        ws = self.wb['Dashboard']
        
        # Scan Dashboard sheet Column A (1) for the table section title
        title_row_idx = None
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val == 'Core & Satellite Distribution':
                title_row_idx = r
                break
                
        self.assertIsNotNone(title_row_idx, "Core & Satellite Distribution section title not found")
        
        # Header is in the next row
        header_row_idx = title_row_idx + 1
        headers = [ws.cell(row=header_row_idx, column=c).value for c in range(1, 7)]
        self.assertEqual(headers, ['Classification', 'Invested (₹)', '% of Total', 'Returns', 'Realized (₹)', 'Un-Realized (₹)'])
        
        # Check data rows: Core and Satellite
        expected_classes = ['Core', 'Satellite']
        for idx, classification in enumerate(expected_classes):
            r = header_row_idx + 1 + idx
            self.assertEqual(ws.cell(row=r, column=1).value, classification)
            
            # Realized (Col 5 / E)
            realized_formula = ws.cell(row=r, column=5).value
            self.assertEqual(realized_formula, f'=SUMIF(Overall_Portfolio!$D$2:$D$1000, "{classification}", Overall_Portfolio!$P$2:$P$1000)')
            self.assertEqual(ws.cell(row=r, column=5).number_format, '[$₹-en-IN] #,##0.00')
            
            # Un-Realized (Col 6 / F)
            unrealized_formula = ws.cell(row=r, column=6).value
            self.assertEqual(unrealized_formula, f'=SUMIF(Current_Portfolio!$D$2:$D$1000, "{classification}", Current_Portfolio!$U$2:$U$1000)')
            self.assertEqual(ws.cell(row=r, column=6).number_format, '[$₹-en-IN] #,##0.00')
            
        # Check TOTAL row (row after the data rows)
        total_row_idx = header_row_idx + 1 + len(expected_classes)
        self.assertEqual(ws.cell(row=total_row_idx, column=1).value, 'TOTAL')
        
        realized_total = ws.cell(row=total_row_idx, column=5).value
        self.assertEqual(realized_total, f'=SUM(E{header_row_idx+1}:E{total_row_idx-1})')
        self.assertEqual(ws.cell(row=total_row_idx, column=5).number_format, '[$₹-en-IN] #,##0.00')
        
        unrealized_total = ws.cell(row=total_row_idx, column=6).value
        self.assertEqual(unrealized_total, f'=SUM(F{header_row_idx+1}:F{total_row_idx-1})')
        self.assertEqual(ws.cell(row=total_row_idx, column=6).number_format, '[$₹-en-IN] #,##0.00')

    def test_cap_allocation_and_pnl_columns(self):
        """Test that Cap Allocation & PnL Breakdown table includes merged columns and correct formulas."""
        create_dashboard(self.wb, self.portfolio_df, self.overall_df)
        ws = self.wb['Dashboard']
        
        # Scan Dashboard sheet Column A (1) for the table section title
        title_row_idx = None
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val == 'Cap Allocation & PnL Breakdown':
                title_row_idx = r
                break
                
        self.assertIsNotNone(title_row_idx, "Cap Allocation & PnL Breakdown section title not found")
        
        # Header is in the next row
        header_row_idx = title_row_idx + 1
        headers = [ws.cell(row=header_row_idx, column=c).value for c in range(1, 8)]
        self.assertEqual(headers, ['Cap', 'Invested (₹)', '% of Total', 'Returns', 'Realized (₹)', 'Un-Realized (₹)', 'Total PnL (₹)'])
        
        # Check data rows: Cap category (sorted alphabetically)
        expected_caps = ['Large Cap']
        for idx, cap in enumerate(expected_caps):
            r = header_row_idx + 1 + idx
            self.assertEqual(ws.cell(row=r, column=1).value, cap)
            
            # Invested (Col 2 / B)
            invested_formula = ws.cell(row=r, column=2).value
            self.assertEqual(invested_formula, f'=SUMIF(Current_Portfolio!$B$2:$B$1000, "{cap}", Current_Portfolio!$L$2:$L$1000)')
            self.assertEqual(ws.cell(row=r, column=2).number_format, '[$₹-en-IN] #,##0.00')
            
            # % of Total (Col 3 / C)
            total_row_idx = header_row_idx + 1 + len(expected_caps)
            pct_formula = ws.cell(row=r, column=3).value
            self.assertEqual(pct_formula, f'=B{r}/$B${total_row_idx}')
            self.assertEqual(ws.cell(row=r, column=3).number_format, '0.00%')
            
            # Returns (Col 4 / D)
            returns_formula = ws.cell(row=r, column=4).value
            self.assertEqual(returns_formula, f'=IF(B{r}>0, (SUMIF(Current_Portfolio!$B$2:$B$1000, "{cap}", Current_Portfolio!$T$2:$T$1000) - B{r}) / B{r}, 0)')
            self.assertEqual(ws.cell(row=r, column=4).number_format, '0.00%')
            
            # Realized (Col 5 / E)
            realized_formula = ws.cell(row=r, column=5).value
            self.assertEqual(realized_formula, f'=SUMIF(Overall_Portfolio!$B$2:$B$1000, "{cap}", Overall_Portfolio!$P$2:$P$1000)')
            self.assertEqual(ws.cell(row=r, column=5).number_format, '[$₹-en-IN] #,##0.00')
            
            # Un-Realized (Col 6 / F)
            unrealized_formula = ws.cell(row=r, column=6).value
            self.assertEqual(unrealized_formula, f'=SUMIF(Current_Portfolio!$B$2:$B$1000, "{cap}", Current_Portfolio!$U$2:$U$1000)')
            self.assertEqual(ws.cell(row=r, column=6).number_format, '[$₹-en-IN] #,##0.00')
            
            # Total PnL (Col 7 / G)
            total_pnl_formula = ws.cell(row=r, column=7).value
            self.assertEqual(total_pnl_formula, f'=E{r}+F{r}')
            self.assertEqual(ws.cell(row=r, column=7).number_format, '[$₹-en-IN] #,##0.00')
            
        # Check TOTAL row
        total_row_idx = header_row_idx + 1 + len(expected_caps)
        self.assertEqual(ws.cell(row=total_row_idx, column=1).value, 'TOTAL')
        
        self.assertEqual(ws.cell(row=total_row_idx, column=2).value, f'=SUM(B{header_row_idx+1}:B{total_row_idx-1})')
        self.assertEqual(ws.cell(row=total_row_idx, column=2).number_format, '[$₹-en-IN] #,##0.00')
        
        self.assertEqual(ws.cell(row=total_row_idx, column=3).value, 1.0)
        self.assertEqual(ws.cell(row=total_row_idx, column=3).number_format, '0.00%')
        
        self.assertEqual(ws.cell(row=total_row_idx, column=4).value, f'=IF(B{total_row_idx}>0, (SUM(Current_Portfolio!$T$2:$T$1000) - B{total_row_idx}) / B{total_row_idx}, 0)')
        self.assertEqual(ws.cell(row=total_row_idx, column=4).number_format, '0.00%')
        
        self.assertEqual(ws.cell(row=total_row_idx, column=5).value, f'=SUM(E{header_row_idx+1}:E{total_row_idx-1})')
        self.assertEqual(ws.cell(row=total_row_idx, column=5).number_format, '[$₹-en-IN] #,##0.00')
        
        self.assertEqual(ws.cell(row=total_row_idx, column=6).value, f'=SUM(F{header_row_idx+1}:F{total_row_idx-1})')
        self.assertEqual(ws.cell(row=total_row_idx, column=6).number_format, '[$₹-en-IN] #,##0.00')
        
        self.assertEqual(ws.cell(row=total_row_idx, column=7).value, f'=SUM(G{header_row_idx+1}:G{total_row_idx-1})')
        self.assertEqual(ws.cell(row=total_row_idx, column=7).number_format, '[$₹-en-IN] #,##0.00')

    def test_sector_allocation_and_pnl_columns(self):
        """Test that Sector Allocation & PnL Breakdown table includes merged columns and correct formulas."""
        create_dashboard(self.wb, self.portfolio_df, self.overall_df)
        ws = self.wb['Dashboard']
        
        # Scan Dashboard sheet Column A (1) for the table section title
        title_row_idx = None
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val == 'Sector Allocation & PnL Breakdown':
                title_row_idx = r
                break
                
        self.assertIsNotNone(title_row_idx, "Sector Allocation & PnL Breakdown section title not found")
        
        # Header is in the next row
        header_row_idx = title_row_idx + 1
        headers = [ws.cell(row=header_row_idx, column=c).value for c in range(1, 9)]
        self.assertEqual(headers, ['Sector', 'Core (₹)', 'Satellite (₹)', 'Total Invested (₹)', '% of Total', 'Realized (₹)', 'Un-Realized (₹)', 'Total PnL (₹)'])
        
        # Check data rows: Sector category
        expected_sectors = ['IT']
        for idx, sector in enumerate(expected_sectors):
            r = header_row_idx + 1 + idx
            self.assertEqual(ws.cell(row=r, column=1).value, sector)
            
            # Core (Col 2 / B)
            core_formula = ws.cell(row=r, column=2).value
            self.assertEqual(core_formula, f'=SUMIFS(Current_Portfolio!$L$2:$L$1000, Current_Portfolio!$C$2:$C$1000, "{sector}", Current_Portfolio!$D$2:$D$1000, "*Core*")')
            self.assertEqual(ws.cell(row=r, column=2).number_format, '[$₹-en-IN] #,##0.00')
            
            # Satellite (Col 3 / C)
            sat_formula = ws.cell(row=r, column=3).value
            self.assertEqual(sat_formula, f'=SUMIFS(Current_Portfolio!$L$2:$L$1000, Current_Portfolio!$C$2:$C$1000, "{sector}", Current_Portfolio!$D$2:$D$1000, "*Satellite*")')
            self.assertEqual(ws.cell(row=r, column=3).number_format, '[$₹-en-IN] #,##0.00')
            
            # Total Invested (Col 4 / D)
            total_inv_formula = ws.cell(row=r, column=4).value
            self.assertEqual(total_inv_formula, f'=B{r}+C{r}')
            self.assertEqual(ws.cell(row=r, column=4).number_format, '[$₹-en-IN] #,##0.00')
            
            # % of Total (Col 5 / E)
            total_row_idx = header_row_idx + 1 + len(expected_sectors)
            pct_formula = ws.cell(row=r, column=5).value
            self.assertEqual(pct_formula, f'=D{r}/$D${total_row_idx}')
            self.assertEqual(ws.cell(row=r, column=5).number_format, '0.00%')
            
            # Realized (Col 6 / F)
            realized_formula = ws.cell(row=r, column=6).value
            self.assertEqual(realized_formula, f'=SUMIF(Overall_Portfolio!$C$2:$C$1000, "{sector}", Overall_Portfolio!$P$2:$P$1000)')
            self.assertEqual(ws.cell(row=r, column=6).number_format, '[$₹-en-IN] #,##0.00')
            
            # Un-Realized (Col 7 / G)
            unrealized_formula = ws.cell(row=r, column=7).value
            self.assertEqual(unrealized_formula, f'=SUMIF(Current_Portfolio!$C$2:$C$1000, "{sector}", Current_Portfolio!$U$2:$U$1000)')
            self.assertEqual(ws.cell(row=r, column=7).number_format, '[$₹-en-IN] #,##0.00')
            
            # Total PnL (Col 8 / H)
            total_pnl_formula = ws.cell(row=r, column=8).value
            self.assertEqual(total_pnl_formula, f'=F{r}+G{r}')
            self.assertEqual(ws.cell(row=r, column=8).number_format, '[$₹-en-IN] #,##0.00')
            
        # Check TOTAL row
        total_row_idx = header_row_idx + 1 + len(expected_sectors)
        self.assertEqual(ws.cell(row=total_row_idx, column=1).value, 'TOTAL')
        
        self.assertEqual(ws.cell(row=total_row_idx, column=2).value, f'=SUM(B{header_row_idx+1}:B{total_row_idx-1})')
        self.assertEqual(ws.cell(row=total_row_idx, column=2).number_format, '[$₹-en-IN] #,##0.00')
        
        self.assertEqual(ws.cell(row=total_row_idx, column=3).value, f'=SUM(C{header_row_idx+1}:C{total_row_idx-1})')
        self.assertEqual(ws.cell(row=total_row_idx, column=3).number_format, '[$₹-en-IN] #,##0.00')
        
        self.assertEqual(ws.cell(row=total_row_idx, column=4).value, f'=SUM(D{header_row_idx+1}:D{total_row_idx-1})')
        self.assertEqual(ws.cell(row=total_row_idx, column=4).number_format, '[$₹-en-IN] #,##0.00')
        
        self.assertEqual(ws.cell(row=total_row_idx, column=5).value, 1.0)
        self.assertEqual(ws.cell(row=total_row_idx, column=5).number_format, '0.00%')
        
        self.assertEqual(ws.cell(row=total_row_idx, column=6).value, f'=SUM(F{header_row_idx+1}:F{total_row_idx-1})')
        self.assertEqual(ws.cell(row=total_row_idx, column=6).number_format, '[$₹-en-IN] #,##0.00')
        
        self.assertEqual(ws.cell(row=total_row_idx, column=7).value, f'=SUM(G{header_row_idx+1}:G{total_row_idx-1})')
        self.assertEqual(ws.cell(row=total_row_idx, column=7).number_format, '[$₹-en-IN] #,##0.00')
        
        self.assertEqual(ws.cell(row=total_row_idx, column=8).value, f'=SUM(H{header_row_idx+1}:H{total_row_idx-1})')
        self.assertEqual(ws.cell(row=total_row_idx, column=8).number_format, '[$₹-en-IN] #,##0.00')

if __name__ == '__main__':
    unittest.main()
