"""
Tests for Satellite Stock Color-Coding
=====================================
Validates that satellite stocks in Current_Portfolio are formatted with colors
matching the latest color codes in the Satellite_Watchlist sheet.
"""

import os
import unittest
import pandas as pd
import openpyxl
from openpyxl import Workbook

from src.excel_writer import _apply_satellite_watchlist_formatting


class TestSatelliteColors(unittest.TestCase):
    """Test suite for verifying the satellite stock color coding logic."""

    def setUp(self):
        self.test_filename = "temp_test_watchlist.xlsx"
        
        # 1. Create a dummy workbook with a Satellite_Watchlist sheet
        wb = Workbook()
        # openpyxl creates a default sheet
        ws_default = wb.active
        ws_default.title = "Current_Portfolio"
        
        # Add headers to Current_Portfolio
        ws_default.append(["Symbol", "TF_Classification", "Current_Quantity"])
        # Add test data
        ws_default.append(["APOLLO", "Satellite", 10])       # Should be colored ORANGE (latest date)
        ws_default.append(["ASTRAMICRO", "Satellite", 5])    # Should be colored GREEN
        ws_default.append(["ACUTAAS", "Core", 20])           # Core, should NOT be colored
        ws_default.append(["GAEL", "Satellite", 15])         # Satellite, but not in watchlist (no color)

        # Add Satellite_Watchlist sheet
        ws_watchlist = wb.create_sheet(title="Satellite_Watchlist")
        ws_watchlist.append(["Date", "Stock", "Color"])
        ws_watchlist.append(["15-05-2026", "APOLLO", "Blue"])       # Older entry
        ws_watchlist.append(["16-05-2026", "APOLLO", "Orange"])     # Newest entry (should win)
        ws_watchlist.append(["16-05-2026", "ASTRAMICRO", "Green"])  # Only entry

        wb.save(self.test_filename)
        wb.close()

    def tearDown(self):
        if os.path.exists(self.test_filename):
            try:
                os.remove(self.test_filename)
            except Exception:
                pass

    def test_satellite_color_formatting(self):
        """Validates that correct background and font colors are applied based on latest date."""
        df_portfolio = pd.DataFrame([
            {"Symbol": "APOLLO", "TF_Classification": "Satellite", "Current_Quantity": 10},
            {"Symbol": "ASTRAMICRO", "TF_Classification": "Satellite", "Current_Quantity": 5},
            {"Symbol": "ACUTAAS", "TF_Classification": "Core", "Current_Quantity": 20},
            {"Symbol": "GAEL", "TF_Classification": "Satellite", "Current_Quantity": 15}
        ])

        # We must load and modify the sheet using ExcelWriter in append mode
        with pd.ExcelWriter(self.test_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_portfolio.to_excel(writer, sheet_name='Current_Portfolio', index=False)
            
            # Apply the color formatting function we wrote
            _apply_satellite_watchlist_formatting(writer, df_portfolio, self.test_filename)

        # Read back the saved file and verify cell formatting
        wb = openpyxl.load_workbook(self.test_filename)
        ws = wb['Current_Portfolio']

        # APOLLO (Row 2, Column 1)
        apollo_cell = ws.cell(row=2, column=1)
        self.assertEqual(apollo_cell.value, "APOLLO")
        self.assertIsNotNone(apollo_cell.fill)
        self.assertEqual(apollo_cell.fill.start_color.rgb, "00FCE4D6")  # Orange BG (Light shade)
        self.assertEqual(apollo_cell.font.color.rgb, "00C65911")        # Orange Font (Dark shade)
        self.assertTrue(apollo_cell.font.bold)

        # ASTRAMICRO (Row 3, Column 1)
        astramicro_cell = ws.cell(row=3, column=1)
        self.assertEqual(astramicro_cell.value, "ASTRAMICRO")
        self.assertIsNotNone(astramicro_cell.fill)
        self.assertEqual(astramicro_cell.fill.start_color.rgb, "00E2EFDA")  # Green BG (Light shade)
        self.assertEqual(astramicro_cell.font.color.rgb, "00375623")        # Green Font (Dark shade)
        self.assertTrue(astramicro_cell.font.bold)

        # ACUTAAS (Row 4, Column 1 - Core)
        acutaas_cell = ws.cell(row=4, column=1)
        self.assertEqual(acutaas_cell.value, "ACUTAAS")
        # Fill is either None or openpyxl default (fill_type is None)
        self.assertTrue(acutaas_cell.fill is None or acutaas_cell.fill.fill_type is None)
        # Font is default (not bold)
        self.assertFalse(acutaas_cell.font.bold)

        # GAEL (Row 5, Column 1 - Satellite, not in watchlist)
        gael_cell = ws.cell(row=5, column=1)
        self.assertEqual(gael_cell.value, "GAEL")
        self.assertTrue(gael_cell.fill is None or gael_cell.fill.fill_type is None)
        self.assertFalse(gael_cell.font.bold)

        wb.close()


if __name__ == '__main__':
    unittest.main()
