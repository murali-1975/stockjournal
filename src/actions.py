"""
Actions Module
===============

Handles generation of SELL/ADD/BUY action recommendations and updating the
'Action Tracker' sheet inside the master workbook.
"""

import os
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Exact headers as found in the user's excel template (including trailing spaces)
HEADERS = [
    'Date',
    'TF_Classification',
    'Stock ',
    'Sector',
    'TF Classification ',
    'Latest Tranch or Cheat',
    'Action (Recommended) ',
    'Reason',
    'Action Taken ',
    'Remarks'
]


def load_portfolio_and_tracker(filepath: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Loads 'Current_Portfolio' and 'Action Tracker' worksheets from Excel.
    If 'Action Tracker' doesn't exist, returns an empty DataFrame with proper headers.
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Workbook {filepath} does not exist.")

    # Load Current_Portfolio
    try:
        portfolio_df = pd.read_excel(filepath, sheet_name='Current_Portfolio')
    except ValueError:
        raise ValueError("Worksheet 'Current_Portfolio' not found in the Excel file.")

    # Load Action Tracker
    try:
        tracker_df = pd.read_excel(filepath, sheet_name='Action Tracker')
        # Ensure all columns are present
        for col in HEADERS:
            if col not in tracker_df.columns:
                tracker_df[col] = None
        tracker_df = tracker_df[HEADERS]
    except ValueError:
        tracker_df = pd.DataFrame(columns=HEADERS)

    return portfolio_df, tracker_df


def generate_action_recommendations(portfolio_df: pd.DataFrame, filepath: str = None, rec_filter: str = 'all') -> pd.DataFrame:
    """
    Applies SELL and ADD recommendation rules to Satellite stocks:
    
    SELL RULES:
    Rule 1 (Cheat exits):
        - Latest_Tranche starts with 'Cheat'
        - Return_Pct < 5% (i.e. < 0.05)
        - Sort by Holding_Period descending
        - Take top 5 longest-held
        - Reasons:
            - If Return_Pct < 0: 'Cheat holding period long and red ({holding} days, {ret_val_pct:.2f}% return)'
            - If Return_Pct >= 0: 'Cheat holding period long and less return ({holding} days, {ret_val_pct:.2f}% return)'

    Rule 2 (Tranch exits):
        - Latest_Tranche starts with 'Tranch'
        - LTP_SL_Diff_Pct < 5% (i.e. < 0.05)
        - Reason: 'Tranch LTP close to stop loss (< 5% difference)'

    ADD RULES (Precedence: SELL always overrides ADD):
    Rule 1 (Cheat additions):
        - Latest_Tranche starts with 'Cheat'
        - Return_Pct > 0 (positive returns)
        - Reason: 'Cheat positive return ({holding} days, {ret_val_pct:.2f}% return)'

    Rule 2 (Tranch additions):
        - Latest_Tranche starts with 'Tranch'
        - LTP > Prev_Week_Close
        - Grouped by Momentum:
            - LTP > 0 to 5% above close: 'Tranch LTP above previous week close [0 to 5%] ({pct_above:.2f}% above)'
            - LTP > 5% above close: 'Tranch LTP above previous week close [> 5%] ({pct_above:.2f}% above)'
    """
    # Filter for active Satellite stocks
    if 'TF_Classification' in portfolio_df.columns:
        satellite = portfolio_df[portfolio_df['TF_Classification'] == 'Satellite'].copy()
    else:
        satellite = pd.DataFrame(columns=portfolio_df.columns)

    if satellite.empty and (not filepath or not os.path.exists(filepath)):
        return pd.DataFrame(columns=HEADERS)

    # Recalculate dynamic values from formulas using static columns if file lookup is provided
    if not satellite.empty and filepath and os.path.exists(filepath):
        from src.market_api import fetch_market_data_from_yahoo
        from src.data_io import load_price_updates

        # Fetch unique symbols in active satellite portfolio
        symbols = satellite['Symbol'].tolist()
        classifications = {sym: 'Satellite' for sym in symbols}

        # 1. Fetch live prices from Yahoo Finance
        try:
            market_data = fetch_market_data_from_yahoo(symbols, classifications=classifications, fetch_info=False)
        except Exception as e:
            print(f"Warning: Failed fetching live market data from Yahoo Finance: {e}")
            market_data = {}

        # 2. Load local price updates as secondary fallback
        price_updates = load_price_updates(filepath)

        def resolve_ltp(row):
            sym = str(row['Symbol']).strip()
            # Try Yahoo Finance first
            ltp_val = market_data.get(sym, {}).get('LTP', 0.0)
            if ltp_val > 0.0:
                return ltp_val
            # Try local price updates next
            ltp_val = price_updates.get(sym, 0.0)
            if ltp_val > 0.0:
                return ltp_val
            # Try static row LTP
            try:
                val = float(row['LTP'])
                if pd.notna(val) and val > 0.0:
                    return val
            except (ValueError, TypeError):
                pass
            return 0.0

        def resolve_prev_week_close(row):
            sym = str(row['Symbol']).strip()
            # Try Yahoo Finance first
            pwc_val = market_data.get(sym, {}).get('Prev_Week_Close', 0.0)
            if pwc_val > 0.0:
                return pwc_val
            # Try static row Prev_Week_Close
            try:
                val = float(row['Prev_Week_Close'])
                if pd.notna(val) and val > 0.0:
                    return val
            except (ValueError, TypeError):
                pass
            return 0.0

        satellite['LTP'] = satellite.apply(resolve_ltp, axis=1)
        satellite['Prev_Week_Close'] = satellite.apply(resolve_prev_week_close, axis=1)
        
        satellite['Current_Value'] = (satellite['Current_Quantity'] * satellite['LTP']).round(2)
        satellite['Unrealized_PnL'] = (satellite['Current_Value'] - satellite['Invested_Value']).round(2)
        
        satellite['Return_Pct'] = satellite.apply(
            lambda r: r['Unrealized_PnL'] / r['Invested_Value'] if r['Invested_Value'] > 0 else 0.0,
            axis=1
        )
        
        satellite['LTP_SL_Diff_Pct'] = satellite.apply(
            lambda r: (r['LTP'] - r['SL']) / r['LTP'] if r['LTP'] > 0 else 0.0,
            axis=1
        )

    recommendations = []
    today_str = datetime.date.today().strftime('%d-%m-%Y')
    sell_symbols = set()

    # ==================== 1. SELL CALCULATIONS ====================
    cheat_sell = pd.DataFrame()
    tranch_sell = pd.DataFrame()
    if not satellite.empty and 'Latest_Tranche' in satellite.columns:
        # --- SELL Rule 1: Cheat Exits (Interpretation B) ---
        cheat_stocks = satellite[satellite['Latest_Tranche'].str.startswith('Cheat', na=False)].copy()
        cheat_underperforming = cheat_stocks[cheat_stocks['Return_Pct'] < 0.05].copy()
        cheat_sell = cheat_underperforming.sort_values(by='Holding_Period', ascending=False).head(5)

    for _, row in cheat_sell.iterrows():
        symbol = str(row['Symbol']).strip()
        sector = str(row.get('TF_Sector', '')).strip()
        tf_class = str(row.get('TF_Classification', '')).strip()
        tranche = str(row.get('Latest_Tranche', '')).strip()
        ret = float(row.get('Return_Pct', 0.0))
        holding = int(row.get('Holding_Period', 0))
        ret_val_pct = ret * 100
        sell_symbols.add(symbol)

        if ret < 0:
            reason = f"Cheat holding period long and red ({holding} days, {ret_val_pct:.2f}% return)"
        else:
            reason = f"Cheat holding period long and less return ({holding} days, {ret_val_pct:.2f}% return)"

        recommendations.append({
            'Date': today_str,
            'TF_Classification': tf_class,
            'Stock ': symbol,
            'Sector': sector,
            'TF Classification ': tf_class,
            'Latest Tranch or Cheat': tranche,
            'Action (Recommended) ': 'SELL',
            'Reason': reason,
            'Action Taken ': '',
            'Remarks': '',
            '_Holding_Period': holding
        })

    # --- SELL Rule 2: Tranch Exits ---
    if not satellite.empty and 'Latest_Tranche' in satellite.columns:
        tranch_stocks = satellite[satellite['Latest_Tranche'].str.startswith('Tranch', na=False)].copy()
        tranch_sell = tranch_stocks[tranch_stocks['LTP_SL_Diff_Pct'] < 0.05].copy()

    for _, row in tranch_sell.iterrows():
        symbol = str(row['Symbol']).strip()
        sector = str(row.get('TF_Sector', '')).strip()
        tf_class = str(row.get('TF_Classification', '')).strip()
        tranche = str(row.get('Latest_Tranche', '')).strip()
        sell_symbols.add(symbol)

        reason = "Tranch LTP close to stop loss (< 5% difference)"

        try:
            holding = int(row.get('Holding_Period', 0))
        except (ValueError, TypeError):
            holding = 0

        recommendations.append({
            'Date': today_str,
            'TF_Classification': tf_class,
            'Stock ': symbol,
            'Sector': sector,
            'TF Classification ': tf_class,
            'Latest Tranch or Cheat': tranche,
            'Action (Recommended) ': 'SELL',
            'Reason': reason,
            'Action Taken ': '',
            'Remarks': '',
            '_Holding_Period': holding
        })

    # ==================== 2. ADD CALCULATIONS ====================

    # --- ADD Rule 1: Cheat Additions ---
    # Candidates are Cheat stocks NOT already recommended for SELL
    cheat_add = pd.DataFrame()
    if not satellite.empty and 'Latest_Tranche' in satellite.columns:
        cheat_add_candidates = cheat_stocks[~cheat_stocks['Symbol'].str.strip().isin(sell_symbols)].copy()
        cheat_add = cheat_add_candidates[cheat_add_candidates['Return_Pct'] > 0]

    for _, row in cheat_add.iterrows():
        symbol = str(row['Symbol']).strip()
        sector = str(row.get('TF_Sector', '')).strip()
        tf_class = str(row.get('TF_Classification', '')).strip()
        tranche = str(row.get('Latest_Tranche', '')).strip()
        ret = float(row.get('Return_Pct', 0.0))
        holding = int(row.get('Holding_Period', 0))
        ret_val_pct = ret * 100
        ltp = float(row.get('LTP', 0.0))
        pwc = float(row.get('Prev_Week_Close', 0.0))

        if pwc > 0:
            pct_above_val = ((ltp - pwc) / pwc) * 100
            reason = f"Cheat positive return ({holding} days, {ret_val_pct:.2f}% return, {pct_above_val:.2f}% return from prev week close)"
        else:
            reason = f"Cheat positive return ({holding} days, {ret_val_pct:.2f}% return)"

        recommendations.append({
            'Date': today_str,
            'TF_Classification': tf_class,
            'Stock ': symbol,
            'Sector': sector,
            'TF Classification ': tf_class,
            'Latest Tranch or Cheat': tranche,
            'Action (Recommended) ': 'ADD',
            'Reason': reason,
            'Action Taken ': '',
            'Remarks': '',
            '_Holding_Period': holding
        })

    # --- ADD Rule 2: Tranch Additions (Momentum) ---
    # Candidates are Tranch stocks NOT already recommended for SELL
    tranch_add = pd.DataFrame()
    if not satellite.empty and 'Latest_Tranche' in satellite.columns:
        tranch_add_candidates = tranch_stocks[~tranch_stocks['Symbol'].str.strip().isin(sell_symbols)].copy()
        tranch_add = tranch_add_candidates[tranch_add_candidates['LTP'] > tranch_add_candidates['Prev_Week_Close']].copy()

    for _, row in tranch_add.iterrows():
        symbol = str(row['Symbol']).strip()
        sector = str(row.get('TF_Sector', '')).strip()
        tf_class = str(row.get('TF_Classification', '')).strip()
        tranche = str(row.get('Latest_Tranche', '')).strip()
        ltp = float(row.get('LTP', 0.0))
        pwc = float(row.get('Prev_Week_Close', 0.0))
        
        try:
            holding = int(row.get('Holding_Period', 0))
        except (ValueError, TypeError):
            holding = 0

        if pwc > 0:
            pct_above = (ltp - pwc) / pwc
            pct_above_val = pct_above * 100
            
            group_str = "0 to 5%" if pct_above <= 0.05 else "> 5%"
            reason = f"Tranch LTP above previous week close [{group_str}] ({pct_above_val:.2f}% above)"

            recommendations.append({
                'Date': today_str,
                'TF_Classification': tf_class,
                'Stock ': symbol,
                'Sector': sector,
                'TF Classification ': tf_class,
                'Latest Tranch or Cheat': tranche,
                'Action (Recommended) ': 'ADD',
                'Reason': reason,
                'Action Taken ': '',
                'Remarks': '',
                '_Holding_Period': holding
            })

    # ==================== 3. BUY CALCULATIONS ====================
    def _calculate_rsi(df, period=14):
        delta = df['Close'].diff()
        gain = delta.where(delta > 0, 0.0)
        loss = -delta.where(delta < 0, 0.0)
        avg_gain = gain.ewm(alpha=1/period, min_periods=period).mean()
        avg_loss = loss.ewm(alpha=1/period, min_periods=period).mean()
        rs = avg_gain / avg_loss
        return 100 - (100 / (1 + rs))

    def _calculate_adx(df, period=14):
        high = df['High']
        low = df['Low']
        close = df['Close']
        tr1 = high - low
        tr2 = (high - close.shift(1)).abs()
        tr3 = (low - close.shift(1)).abs()
        tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
        atr = tr.ewm(alpha=1/period, min_periods=period).mean()
        
        up_move = high.diff()
        down_move = low.shift(1) - low
        
        plus_dm = pd.Series(0.0, index=df.index)
        minus_dm = pd.Series(0.0, index=df.index)
        
        plus_dm_mask = (up_move > down_move) & (up_move > 0)
        plus_dm[plus_dm_mask] = up_move[plus_dm_mask]
        
        minus_dm_mask = (down_move > up_move) & (down_move > 0)
        minus_dm[minus_dm_mask] = down_move[minus_dm_mask]
        
        plus_di = 100 * (plus_dm.ewm(alpha=1/period, min_periods=period).mean() / atr)
        minus_di = 100 * (minus_dm.ewm(alpha=1/period, min_periods=period).mean() / atr)
        
        dx = 100 * ((plus_di - minus_di).abs() / (plus_di + minus_di))
        return dx.ewm(alpha=1/period, min_periods=period).mean()

    # Identify BUY Candidates from Satellite_Watchlist
    green_blue_candidates = []
    watchlist_df = pd.DataFrame()
    if filepath and os.path.exists(filepath):
        try:
            watchlist_df = pd.read_excel(filepath, sheet_name='Satellite_Watchlist')
            if not watchlist_df.empty:
                watchlist_df = watchlist_df.dropna(subset=['Stock', 'Color'])
                watchlist_df['Stock'] = watchlist_df['Stock'].astype(str).str.strip()
                watchlist_df['Color'] = watchlist_df['Color'].astype(str).str.strip().str.upper()
                
                watchlist_df['Date'] = pd.to_datetime(watchlist_df['Date'], dayfirst=True, errors='coerce')
                watchlist_df = watchlist_df.dropna(subset=['Date'])
                
                if not watchlist_df.empty:
                    latest_date = watchlist_df['Date'].max()
                    latest_date_df = watchlist_df[watchlist_df['Date'] == latest_date]
                    
                    # Keep latest color for each stock on the latest date
                    latest_watchlist_colors = latest_date_df.drop_duplicates(subset=['Stock']).set_index('Stock')['Color'].to_dict()
                    
                    # Current portfolio symbols to filter out
                    current_portfolio_symbols = set(portfolio_df['Symbol'].dropna().astype(str).str.strip().tolist())
                    
                    # Candidates are watchlisted Green/Blue stocks ON THE LATEST DATE that are NOT currently in portfolio
                    green_blue_candidates = [
                        sym for sym, col in latest_watchlist_colors.items()
                        if col in ['GREEN', 'BLUE'] and sym not in current_portfolio_symbols
                    ]
        except Exception as e:
            print(f"Warning: Failed loading candidates from Satellite_Watchlist: {e}")

    if green_blue_candidates:
        import yfinance as yf
        symbols_ns = [sym + '.NS' for sym in green_blue_candidates]
        try:
            print(f"Downloading weekly historical data for {len(green_blue_candidates)} BUY candidate stocks...")
            hist_data = yf.download(symbols_ns, period="1y", interval="1wk", progress=False)
            
            for sym, ns_sym in zip(green_blue_candidates, symbols_ns):
                try:
                    # Squeeze columns or extract for the symbol
                    if len(green_blue_candidates) == 1:
                        df_sym = hist_data.copy()
                        if isinstance(df_sym.columns, pd.MultiIndex):
                            df_sym.columns = df_sym.columns.get_level_values(0)
                    else:
                        # Extract the columns for this symbol
                        df_sym = pd.DataFrame()
                        for col in ['Open', 'High', 'Low', 'Close', 'Volume']:
                            if col in hist_data and ns_sym in hist_data[col]:
                                df_sym[col] = hist_data[col][ns_sym]
                    
                    df_sym = df_sym.dropna(subset=['Close'])
                    if len(df_sym) < 21: # Need at least 21 data points for rolling 20-week volume
                        continue
                    
                    # Squeeze columns if they contain multi-index values
                    if isinstance(df_sym.columns, pd.MultiIndex):
                        df_sym.columns = df_sym.columns.get_level_values(0)
                    
                    # Compute indicators
                    rsi_series = _calculate_rsi(df_sym)
                    adx_series = _calculate_adx(df_sym)
                    
                    # Latest parameters
                    latest_close = float(df_sym['Close'].iloc[-1].item() if hasattr(df_sym['Close'].iloc[-1], 'item') else df_sym['Close'].iloc[-1])
                    prev_close = float(df_sym['Close'].iloc[-2].item() if hasattr(df_sym['Close'].iloc[-2], 'item') else df_sym['Close'].iloc[-2])
                    
                    latest_rsi = float(rsi_series.iloc[-1].item() if hasattr(rsi_series.iloc[-1], 'item') else rsi_series.iloc[-1])
                    latest_adx = float(adx_series.iloc[-1].item() if hasattr(adx_series.iloc[-1], 'item') else adx_series.iloc[-1])
                    
                    latest_vol = float(df_sym['Volume'].iloc[-1].item() if hasattr(df_sym['Volume'].iloc[-1], 'item') else df_sym['Volume'].iloc[-1])
                    avg_vol = df_sym['Volume'].rolling(window=20).mean().iloc[-1]
                    latest_avg_vol = float(avg_vol.item() if hasattr(avg_vol, 'item') else avg_vol)
                    
                    # BUY Conditions check:
                    # 1. LTP > previous weekly close
                    # 2. ADX > 25
                    # 3. RSI >= 50
                    # 4. Weekly volume > average weekly volume (20w)
                    if (latest_close > prev_close) and (latest_rsi >= 50.0) and (latest_adx > 25.0) and (latest_vol > latest_avg_vol):
                        vol_pct_above = ((latest_vol - latest_avg_vol) / latest_avg_vol) * 100
                        reason = f"Buy breakout (RSI: {latest_rsi:.1f}, ADX: {latest_adx:.1f}, Vol: +{vol_pct_above:.0f}% above average)"
                        
                        recommendations.append({
                            'Date': today_str,
                            'TF_Classification': 'Satellite',
                            'Stock ': sym,
                            'Sector': '',
                            'TF Classification ': 'Satellite',
                            'Latest Tranch or Cheat': 'Watchlist',
                            'Action (Recommended) ': 'BUY',
                            'Reason': reason,
                            'Action Taken ': '',
                            'Remarks': '',
                            '_Holding_Period': 0
                        })
                except Exception as sym_err:
                    print(f"Warning: Failed processing buy indicators for {sym}: {sym_err}")
        except Exception as e:
            print(f"Warning: Failed historical download for BUY candidates: {e}")

    if not recommendations:
        return pd.DataFrame(columns=HEADERS)

    # Compile DataFrame and sort
    recs_df = pd.DataFrame(recommendations)
    
    # Sort criteria helper columns
    # 1. SELL (0), ADD (1), BUY (2)
    recs_df['Action_Sort'] = recs_df['Action (Recommended) '].map({'SELL': 0, 'ADD': 1, 'BUY': 2}).fillna(3)
    
    # 2. Cheat (0) then Tranch (1)
    recs_df['Tranche_Sort'] = recs_df['Latest Tranch or Cheat'].apply(
        lambda x: 0 if str(x).strip().startswith('Cheat') else 1
    )
    
    # 3. Within Cheat, sort by holding period descending (ascending -Holding_Period)
    def get_holding_sort(r):
        val = r.get('_Holding_Period', 0)
        try:
            h_val = float(val)
        except (ValueError, TypeError):
            h_val = 0.0
        # If it's a Cheat stock, sort descending
        if r['Tranche_Sort'] == 0:
            return -h_val
        return 0.0

    recs_df['Cheat_Holding_Sort'] = recs_df.apply(get_holding_sort, axis=1)

    # 4. Perform multi-level sort
    recs_df = recs_df.sort_values(by=['Action_Sort', 'Tranche_Sort', 'Cheat_Holding_Sort', 'Stock '])

    # Drop sort helper columns
    recs_df = recs_df.drop(columns=['Action_Sort', 'Tranche_Sort', 'Cheat_Holding_Sort', '_Holding_Period'], errors='ignore')

    # Filter by action type if requested
    if rec_filter and rec_filter != 'all':
        recs_df = recs_df[recs_df['Action (Recommended) '].str.upper() == rec_filter.upper()]

    return recs_df


def write_recommendations_to_excel(filepath: str, new_recs_df: pd.DataFrame, rec_filter: str = 'all') -> None:
    """
    Saves and styles recommendations in the 'Action Tracker' sheet of the Excel workbook.
    Overwrites rows matching today's date (preventing same-day duplicates) and appends
    for new dates. Applies professional styling.
    
    Tries xlwings first to preserve Excel STOCKS rich data types, falling back to openpyxl
    if xlwings is not installed or fails.
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Workbook {filepath} does not exist.")

    # 1. Fetch latest colors from Satellite_Watchlist
    latest_colors = {}
    try:
        watchlist_df = pd.read_excel(filepath, sheet_name='Satellite_Watchlist')
        if not watchlist_df.empty:
            watchlist_df = watchlist_df.dropna(subset=['Stock', 'Color'])
            watchlist_df['Stock'] = watchlist_df['Stock'].astype(str).str.strip()
            watchlist_df['Color'] = watchlist_df['Color'].astype(str).str.strip().str.upper()
            
            watchlist_df['Date'] = pd.to_datetime(watchlist_df['Date'], dayfirst=True, errors='coerce')
            watchlist_df = watchlist_df.sort_values(by='Date', ascending=False)
            
            latest_colors = watchlist_df.drop_duplicates(subset=['Stock']).set_index('Stock')['Color'].to_dict()
    except Exception:
        pass

    COLOR_MAP_RGB = {
        'BLUE':   {'bg': (221, 235, 247), 'font': (31, 78, 120)},
        'ORANGE': {'bg': (252, 228, 214), 'font': (198, 89, 17)},
        'GREEN':  {'bg': (226, 239, 218), 'font': (55, 86, 35)},
        'RED':    {'bg': (250, 219, 216), 'font': (169, 50, 38)},
        'PINK':   {'bg': (250, 219, 216), 'font': (169, 50, 38)},
        'YELLOW': {'bg': (255, 242, 204), 'font': (127, 96, 0)},
        'PURPLE': {'bg': (225, 213, 231), 'font': (96, 73, 122)}
    }

    COLOR_MAP_HEX = {
        'BLUE':   {'bg': 'DDEBF7', 'font': '1F4E78'},
        'ORANGE': {'bg': 'FCE4D6', 'font': 'C65911'},
        'GREEN':  {'bg': 'E2EFDA', 'font': '375623'},
        'RED':    {'bg': 'FADBD8', 'font': 'A93226'},
        'PINK':   {'bg': 'FADBD8', 'font': 'A93226'},
        'YELLOW': {'bg': 'FFF2CC', 'font': '7F6000'},
        'PURPLE': {'bg': 'E1D5E7', 'font': '60497A'}
    }

    use_xlwings = False
    try:
        import xlwings as xw
        # Ensure we are on Windows (xlwings requires a native Excel app installation)
        import sys
        if sys.platform == 'win32':
            use_xlwings = True
    except ImportError:
        pass

    if use_xlwings:
        print("Using xlwings to update 'Action Tracker' (preserving STOCKS data types)...")
        app = None
        try:
            import xlwings as xw
            app = xw.App(visible=False)
            app.display_alerts = False
            wb = app.books.open(os.path.abspath(filepath))
            
            # Access or create Action Tracker sheet
            sheet_names = [wb.sheets[i].name for i in range(len(wb.sheets))]
            if 'Action Tracker' in sheet_names:
                ws = wb.sheets['Action Tracker']
            else:
                # Add after the first sheet (usually Dashboard)
                ws = wb.sheets.add('Action Tracker', after=wb.sheets[0])
            
            # Read existing rows
            existing_rows = []
            # Find last row in column A
            last_row = ws.range('A' + str(ws.cells.last_cell.row)).end('up').row
            if last_row >= 2:
                # Read all values from column A to J
                headers_raw = ws.range('A1:J1').value
                headers = [str(h) for h in headers_raw]
                col_indices = {name: i for i, name in enumerate(headers) if name in HEADERS}
                
                data_range = ws.range(f'A2:J{last_row}').value
                if last_row == 2:
                    data_range = [data_range]
                
                for row_vals in data_range:
                    if row_vals:
                        row_dict = {}
                        for col_name in HEADERS:
                            idx = col_indices.get(col_name)
                            row_dict[col_name] = row_vals[idx] if idx is not None and idx < len(row_vals) else None
                        existing_rows.append(row_dict)
            
            existing_df = pd.DataFrame(existing_rows, columns=HEADERS)
            
            # Same-Day Overwrite Logic
            today_str = datetime.date.today().strftime('%d-%m-%Y')
            if not existing_df.empty:
                if rec_filter == 'buy':
                    existing_df = existing_df[~((existing_df['Date'] == today_str) & (existing_df['Action (Recommended) '].astype(str).str.strip().str.upper() == 'BUY'))]
                elif rec_filter == 'add':
                    existing_df = existing_df[~((existing_df['Date'] == today_str) & (existing_df['Action (Recommended) '].astype(str).str.strip().str.upper() == 'ADD'))]
                elif rec_filter == 'sell':
                    existing_df = existing_df[~((existing_df['Date'] == today_str) & (existing_df['Action (Recommended) '].astype(str).str.strip().str.upper() == 'SELL'))]
                else:
                    existing_df = existing_df[existing_df['Date'] != today_str]
            
            if not new_recs_df.empty:
                combined_df = pd.concat([existing_df, new_recs_df], ignore_index=True)
            else:
                combined_df = existing_df
            
            # Clear everything on sheet
            ws.clear()
            
            # Write headers
            ws.range('A1').value = HEADERS
            
            # Write combined data rows
            if not combined_df.empty:
                # Replace nan with empty string
                combined_df = combined_df.fillna('')
                ws.range('A2').value = combined_df.values.tolist()
            
            last_write_row = len(combined_df) + 1
            
            # Header styling
            header_range = ws.range('A1:J1')
            header_range.color = (47, 84, 150) # RGB for #2F5496
            header_range.font.bold = True
            header_range.font.color = (255, 255, 255)
            header_range.api.HorizontalAlignment = -4108 # Center
            header_range.api.VerticalAlignment = -4108 # Center
            
            # Borders: Apply thin borders
            full_range = ws.range(f'A1:J{last_write_row}')
            for border_id in [7, 8, 9, 10, 11, 12]:
                try:
                    full_range.api.Borders(border_id).LineStyle = 1 # xlContinuous
                    full_range.api.Borders(border_id).Weight = 2 # xlThin
                except Exception:
                    pass
            
            # Freeze Panes 'A2'
            try:
                ws.activate()
                app.api.ActiveWindow.ScrollRow = 1
                app.api.ActiveWindow.ScrollColumn = 1
                app.api.ActiveWindow.FreezePanes = False
                ws.range('A2').select()
                app.api.ActiveWindow.FreezePanes = True
            except Exception:
                pass
            
            # Styling alignments row-by-row
            center_cols = ['A', 'B', 'C', 'E', 'F', 'G']
            for col_let in center_cols:
                try:
                    ws.range(f'{col_let}2:{col_let}{last_write_row}').api.HorizontalAlignment = -4108 # Center
                except Exception:
                    pass
            
            left_cols = ['D', 'H', 'I', 'J']
            for col_let in left_cols:
                try:
                    ws.range(f'{col_let}2:{col_let}{last_write_row}').api.HorizontalAlignment = -4131 # Left
                except Exception:
                    pass
            
            # Color Recommended Actions & Stocks
            for r in range(2, last_write_row + 1):
                act_cell = ws.range((r, 7))
                act_val = str(act_cell.value).strip().upper()
                if act_val == 'SELL':
                    act_cell.color = (250, 219, 216) # Soft Red #FADBD8
                    act_cell.font.color = (169, 50, 38) # Dark Red #A93226
                    act_cell.font.bold = True
                elif act_val == 'ADD':
                    act_cell.color = (226, 239, 218) # Soft Green #E2EFDA
                    act_cell.font.color = (55, 86, 35) # Dark Green #375623
                    act_cell.font.bold = True
                elif act_val == 'BUY':
                    act_cell.color = (221, 235, 247) # Soft Blue #DDEBF7
                    act_cell.font.color = (31, 78, 120) # Dark Blue #1F4E78
                    act_cell.font.bold = True

                # Dynamic color-coding for stock symbols
                stock_cell = ws.range((r, 3))
                stock_val = str(stock_cell.value).strip()
                if stock_val in latest_colors:
                    color_name = latest_colors[stock_val]
                    if color_name in COLOR_MAP_RGB:
                        bg_rgb = COLOR_MAP_RGB[color_name]['bg']
                        font_rgb = COLOR_MAP_RGB[color_name]['font']
                        stock_cell.color = bg_rgb
                        stock_cell.font.color = font_rgb
                        stock_cell.font.bold = True
                else:
                    stock_cell.font.color = (139, 0, 0) # Dark Red
                    stock_cell.font.bold = True
            
            # Auto-fit columns
            ws.autofit()
            
            wb.save()
            wb.close()
            app.quit()
            print("Action Tracker sheet successfully updated via xlwings!")
            return
        except Exception as e:
            print(f"Failed using xlwings for Action Tracker: {e}. Falling back to openpyxl...")
            if app:
                try:
                    app.quit()
                except Exception:
                    pass

    # --- FALLBACK: Standard openpyxl Writer Block ---
    wb = openpyxl.load_workbook(filepath)

    # 1. Access or create Action Tracker sheet
    if 'Action Tracker' in wb.sheetnames:
        ws = wb['Action Tracker']
    else:
        ws = wb.create_sheet('Action Tracker')

    # Read existing rows
    existing_rows = []
    if ws.max_row >= 2:
        # Check headers first
        headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        col_indices = {name: i + 1 for i, name in enumerate(headers) if name in HEADERS}
        
        # Load data rows
        for r in range(2, ws.max_row + 1):
            row_data = {}
            for col_name in HEADERS:
                c_idx = col_indices.get(col_name)
                row_data[col_name] = ws.cell(row=r, column=c_idx).value if c_idx else None
            existing_rows.append(row_data)

    existing_df = pd.DataFrame(existing_rows, columns=HEADERS)

    # 2. Same-Day Overwrite Logic
    today_str = datetime.date.today().strftime('%d-%m-%Y')
    
    if not existing_df.empty:
        if rec_filter == 'buy':
            existing_df = existing_df[~((existing_df['Date'] == today_str) & (existing_df['Action (Recommended) '].astype(str).str.strip().str.upper() == 'BUY'))]
        elif rec_filter == 'add':
            existing_df = existing_df[~((existing_df['Date'] == today_str) & (existing_df['Action (Recommended) '].astype(str).str.strip().str.upper() == 'ADD'))]
        elif rec_filter == 'sell':
            existing_df = existing_df[~((existing_df['Date'] == today_str) & (existing_df['Action (Recommended) '].astype(str).str.strip().str.upper() == 'SELL'))]
        else:
            existing_df = existing_df[existing_df['Date'] != today_str]

    # Combine existing non-today rows with new recommendations
    if not new_recs_df.empty:
        combined_df = pd.concat([existing_df, new_recs_df], ignore_index=True)
    else:
        combined_df = existing_df

    # 3. Write data to the worksheet
    # Clear the entire sheet first to clean old contents & old conditional formats
    ws.delete_rows(1, ws.max_row + 1)

    # Write headers
    for c_idx, col_name in enumerate(HEADERS, 1):
        ws.cell(row=1, column=c_idx, value=col_name)

    # Write data rows
    for r_idx, row_dict in enumerate(combined_df.to_dict('records'), 2):
        for c_idx, col_name in enumerate(HEADERS, 1):
            val = row_dict.get(col_name, '')
            ws.cell(row=r_idx, column=c_idx, value=val if pd.notna(val) else '')

    # 4. Sheet Styling (Borders, Headers, Centering, Fills)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    sell_fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
    sell_font = Font(color='A93226', bold=True)

    add_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    add_font = Font(color='375623', bold=True)

    buy_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    buy_font = Font(color='1F4E78', bold=True)

    # Style Header Row
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Style Data Rows
    ws.freeze_panes = 'A2'
    for row in range(2, ws.max_row + 1):
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = thin_border
            
            # Align centered for code fields, dates, actions
            col_name = HEADERS[col_idx - 1]
            if col_name in ['Date', 'TF_Classification', 'Stock ', 'TF Classification ', 'Latest Tranch or Cheat', 'Action (Recommended) ']:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

            # Dynamic styling for recommended actions
            if col_name == 'Action (Recommended) ':
                val = str(cell.value).strip().upper()
                if val == 'SELL':
                    cell.fill = sell_fill
                    cell.font = sell_font
                elif val == 'ADD':
                    cell.fill = add_fill
                    cell.font = add_font
                elif val == 'BUY':
                    cell.fill = buy_fill
                    cell.font = buy_font

            # Dynamic styling for stock symbols based on watchlist colors
            elif col_name == 'Stock ':
                sym = str(cell.value).strip()
                if sym in latest_colors:
                    color_name = latest_colors[sym]
                    if color_name in COLOR_MAP_HEX:
                        bg_hex = COLOR_MAP_HEX[color_name]['bg']
                        font_hex = COLOR_MAP_HEX[color_name]['font']
                        cell.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type='solid')
                        cell.font = Font(color=font_hex, bold=True)
                else:
                    cell.font = Font(color='8B0000', bold=True)

    # Dynamic column widths auto-fitting
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    wb.save(filepath)
    wb.close()


def execute_recommendation_bypass(filepath: str, rec_filter: str = 'all') -> None:
    """
    Executes the recommendations pipeline directly:
    1. Reads the portfolio state.
    2. Runs criteria to identify new recommendations.
    3. Overwrites today's rows / appends new rows in the workbook.
    """
    print("=" * 60)
    print(f"  RUNNING RECOMMENDATIONS PIPELINE (BYPASS MODE - FILTER: {rec_filter.upper()})")
    print("=" * 60)
    print(f"Reading state from {filepath}...")
    
    portfolio_df, _ = load_portfolio_and_tracker(filepath)
    print(f"Loaded portfolio: {len(portfolio_df)} total active holdings.")

    print("Evaluating recommendation exit/entry criteria rules...")
    recs_df = generate_action_recommendations(portfolio_df, filepath, rec_filter=rec_filter)

    if not recs_df.empty:
        print(f"Found {len(recs_df)} recommended exit/entry/buy actions:")
        for idx, row in recs_df.iterrows():
            print(f" - {row['Stock ']}: {row['Action (Recommended) ']} ({row['Reason']})")
    else:
        print("No stocks matched the exit/entry/buy recommendation criteria.")

    print(f"Writing updates to 'Action Tracker' sheet in {filepath}...")
    write_recommendations_to_excel(filepath, recs_df, rec_filter=rec_filter)
    print("[SUCCESS] Recommendations successfully compiled!")
    print("=" * 60)
