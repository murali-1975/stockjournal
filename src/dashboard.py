"""
Dashboard Module
================

Creates a 'Dashboard' sheet with pre-computed summary tables that are
ready for the user to chart manually in Excel.

Tables included:
    1. Portfolio KPIs
    2. Cap-wise Allocation
    3. Core & Satellite Distribution
    4. Sector-wise Allocation
    5. Top 5 Gainers / Bottom 5 Losers
    6. Realized vs Unrealized PnL by Cap
    7. Stocks Nearest to Stop Loss
    8. Tranche Distribution
    9. Holding Period Distribution
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd


# ─── Styling Constants ──────────────────────────────────────────────
TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='2F5496')
SECTION_FONT = Font(name='Calibri', bold=True, size=12, color='2F5496')
HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
LABEL_FONT = Font(name='Calibri', bold=True, size=11)
VALUE_FONT = Font(name='Calibri', size=11)
GREEN_FONT = Font(name='Calibri', bold=True, color='147A1E')
RED_FONT = Font(name='Calibri', bold=True, color='C00000')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
INR_FMT = '[$₹-en-IN] #,##0.00'
PCT_FMT = '0.00%'

COLOR_MAP = {
    'BLUE':   {'bg': 'DDEBF7', 'font': '1F4E78'},
    'ORANGE': {'bg': 'FCE4D6', 'font': 'C65911'},
    'GREEN':  {'bg': 'E2EFDA', 'font': '375623'},
    'RED':    {'bg': 'FADBD8', 'font': 'A93226'},
    'PINK':   {'bg': 'FADBD8', 'font': 'A93226'},
    'YELLOW': {'bg': 'FFF2CC', 'font': '7F6000'},
    'PURPLE': {'bg': 'E1D5E7', 'font': '60497A'}
}

def _extract_portfolio_styles(wb) -> dict:
    """
    Extracts the exact font and fill styles applied to the 'Symbol' column in 'Current_Portfolio'.
    """
    portfolio_styles = {}
    if 'Current_Portfolio' in wb.sheetnames:
        ws = wb['Current_Portfolio']
        symbol_col = -1
        for col_idx in range(1, ws.max_column + 1):
            if str(ws.cell(row=1, column=col_idx).value).strip().upper() == 'SYMBOL':
                symbol_col = col_idx
                break
                
        if symbol_col > 0:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=symbol_col)
                symbol = str(cell.value).strip().upper()
                if symbol:
                    # Manually build new Font and PatternFill objects to completely avoid openpyxl StyleProxy reference chains
                    f = cell.font
                    new_font = None
                    if f:
                        new_font = Font(
                            name=f.name,
                            size=f.size,
                            bold=f.bold,
                            italic=f.italic,
                            charset=f.charset,
                            color=f.color,
                            underline=f.underline,
                            strike=f.strike,
                            vertAlign=f.vertAlign,
                            scheme=f.scheme
                        )
                    
                    fill = cell.fill
                    new_fill = None
                    if fill and getattr(fill, 'fill_type', None) is not None:
                        new_fill = PatternFill(
                            fill_type=fill.fill_type,
                            start_color=fill.start_color,
                            end_color=fill.end_color
                        )
                    
                    portfolio_styles[symbol] = {
                        'font': new_font,
                        'fill': new_fill
                    }
    return portfolio_styles

def _get_portfolio_style(symbol: str, portfolio_styles: dict) -> tuple:
    """
    Returns (font, fill) for a given symbol by looking it up in portfolio_styles.
    If not found or no custom fill, returns (LABEL_FONT, None).
    """
    if symbol in portfolio_styles:
        st = portfolio_styles[symbol]
        if st['fill'] and getattr(st['fill'], 'fill_type', None):
            return st['font'], st['fill']
    return LABEL_FONT, None

def _find_price_update_columns_from_workbook(wb) -> tuple:
    """
    Locates the Symbol and LTP columns in the given Price_Update sheet.
    Returns (symbol_col_letter, ltp_col_letter) or (None, None) if not found.
    """
    from openpyxl.utils import get_column_letter
    try:
        if 'Price_Update' not in wb.sheetnames:
            return None, None
        ws = wb['Price_Update']
        symbol_col_letter = None
        ltp_col_letter = None
        
        # Check the first 10 rows and 100 columns
        for row_idx in range(1, 11):
            for col_idx in range(1, 100):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    header = str(cell_val).strip().upper()
                    if header in ['SYMBOL', 'STOCK SYMBOL', 'STOCK_SYMBOL']:
                        symbol_col_letter = get_column_letter(col_idx)
                    elif header in ['LTP', 'LAST TRADED PRICE', 'LAST_TRADED_PRICE', 'PRICE']:
                        ltp_col_letter = get_column_letter(col_idx)
            if symbol_col_letter and ltp_col_letter:
                break
        return symbol_col_letter, ltp_col_letter
    except Exception:
        return None, None


def create_dashboard(wb, portfolio_df: pd.DataFrame, overall_df: pd.DataFrame, raw_df: pd.DataFrame = None, benchmark_returns: dict = None, watchlist_df: pd.DataFrame = None, latest_core_trends: dict = None) -> None:
    """
    Creates (or replaces) a 'Dashboard' sheet with pre-computed summary tables.

    Args:
        wb:           An open openpyxl Workbook instance.
        portfolio_df: The Current Portfolio DataFrame.
        overall_df:   The Overall Portfolio DataFrame.
    """
    # Force numeric types on essential columns to prevent 'object' dtype errors during calculations
    num_cols_p = ['Invested_Value', 'Current_Value', 'Unrealized_PnL', 'LTP', 'SL', 
                  'LTP_SL_Diff', 'LTP_SL_Diff_Pct', 'Holding_Period', 'Average_Buy_Price']
    for col in num_cols_p:
        if col in portfolio_df.columns:
            portfolio_df[col] = pd.to_numeric(portfolio_df[col], errors='coerce').fillna(0)

    num_cols_o = ['Realized_PnL', 'Unrealized_PnL', 'Total_PnL', 'Total_PnL_Percentage', 
                  'Invested_Value', 'Current_Value', 'Total_Buy_Value', 'Average_Buy_Price',
                  'Total_Sell_Value', 'Average_Sell_Price']
    for col in num_cols_o:
        if col in overall_df.columns:
            overall_df[col] = pd.to_numeric(overall_df[col], errors='coerce').fillna(0)

    # Dynamic columns lookup from Price_Update sheet
    price_cols = _find_price_update_columns_from_workbook(wb)

    if 'Dashboard' in wb.sheetnames:
        del wb['Dashboard']

    ws = wb.create_sheet('Dashboard', 0)

    # Column widths
    widths = {'A': 24, 'B': 18, 'C': 18, 'D': 24, 'E': 18, 'F': 18, 'G': 18, 'H': 18,
              'I': 4,  # spacer
              'J': 24, 'K': 18, 'L': 18, 'M': 18, 'N': 18, 'O': 18, 'P': 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── Title ────────────────────────────────────────────────────────
    ws.merge_cells('A1:H1')
    ws['A1'].value = '📊 Portfolio Dashboard'
    ws['A1'].font = Font(name='Calibri', bold=True, size=18, color='2F5496')
    ws['A1'].alignment = Alignment(horizontal='center')

    if raw_df is not None and not raw_df.empty and 'Trade Date' in raw_df.columns:
        try:
            max_date = pd.to_datetime(raw_df['Trade Date']).max()
            if pd.notna(max_date):
                ws.merge_cells('J1:N1')
                ws['J1'].value = f'🕒 Last Transacted Date: {max_date.strftime("%d %b %Y")}'
                ws['J1'].font = Font(name='Calibri', italic=True, size=12, color='595959', bold=True)
                ws['J1'].alignment = Alignment(horizontal='right')
        except Exception:
            pass

    # Extract colors from watchlist for consistent Satellite styling
    latest_colors = {}
    if watchlist_df is not None and not watchlist_df.empty:
        try:
            wdf = watchlist_df.dropna(subset=['Stock', 'Color']).copy()
            wdf['Stock'] = wdf['Stock'].astype(str).str.strip().str.upper()
            wdf['Color'] = wdf['Color'].astype(str).str.strip().str.upper()
            wdf['Date'] = pd.to_datetime(wdf['Date'], dayfirst=True, errors='coerce')
            wdf = wdf.dropna(subset=['Date'])
            if not wdf.empty:
                latest_date = wdf['Date'].max()
                wdf = wdf[wdf['Date'] == latest_date]
                latest_colors = wdf.drop_duplicates(subset=['Stock']).set_index('Stock')['Color'].to_dict()
        except Exception:
            pass

    latest_core_trends = latest_core_trends or {}

    # ══════════════════════════════════════════════════════════════════
    #  LEFT SIDE (Columns A-H) — KPIs, Allocation Tables, Movers, etc.
    # ══════════════════════════════════════════════════════════════════

    portfolio_styles = _extract_portfolio_styles(wb)

    row = 3
    _write_performance_kpis(ws, row, col_start=4)
    row = _write_kpi_table(ws, portfolio_df, overall_df, row)
    row += 2
    row = _write_cap_allocation(ws, portfolio_df, overall_df, row, col_start=1)
    row += 2
    row = _write_classification_allocation(ws, portfolio_df, row, col_start=1)
    row += 2
    row = _write_sector_allocation(ws, portfolio_df, overall_df, row, col_start=1)
    row += 2
    row = _write_top_bottom_table(ws, portfolio_df, portfolio_styles, row, classification_filter='Satellite')
    row += 2
    row = _write_top_bottom_table(ws, portfolio_df, portfolio_styles, row, classification_filter='Core')
    row += 2
    row = _write_nearest_sl_table(ws, portfolio_df, portfolio_styles, row)
    row += 2
    row = _write_corporate_actions(ws, portfolio_df, overall_df, portfolio_styles, row)
    row += 2
    row = _write_top_cheats_table(ws, portfolio_df, portfolio_styles, row, classification_filter='Satellite')
    row += 2
    row = _write_top_cheats_table(ws, portfolio_df, portfolio_styles, row, classification_filter='Core')
    row += 2
    row = _write_daily_losers_table(ws, portfolio_df, portfolio_styles, row, classification_filter='Satellite')
    row += 2
    row = _write_daily_losers_table(ws, portfolio_df, portfolio_styles, row, classification_filter='Core')

    # ══════════════════════════════════════════════════════════════════
    #  RIGHT SIDE (Columns J-P) — Benchmark, Distributions, Movers
    # ══════════════════════════════════════════════════════════════════

    rrow = 3
    if benchmark_returns:
        rrow = _write_benchmark_returns_table(ws, benchmark_returns, rrow, price_cols, col_start=10)
        rrow += 2

    rrow = _write_tranche_distribution(ws, portfolio_df, rrow, col_start=10)
    rrow += 2
    rrow = _write_holding_distribution(ws, portfolio_df, rrow, col_start=10)
    rrow += 2
    rrow = _write_portfolio_movers_table(ws, portfolio_df, watchlist_df, portfolio_styles, rrow, col_start=10)
    rrow += 2
    rrow = _write_watchlist_movers_table(ws, watchlist_df, portfolio_df, latest_colors, rrow, col_start=10)
    rrow += 2

    print("Dashboard sheet created.")


# ═══════════════════════════════════════════════════════════════════
#  Helper: Write a styled table header row
# ═══════════════════════════════════════════════════════════════════

def _styled_header(ws, row, col_start, headers):
    """Writes a styled header row and returns the next row."""
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center')
        c.border = THIN_BORDER
    return row + 1


def _section_title(ws, row, col, title):
    """Writes a section title and returns the next row."""
    c = ws.cell(row=row, column=col, value=title)
    c.font = SECTION_FONT
    return row + 1


def _data_cell(ws, row, col, value, fmt=None, font=None):
    """Writes a formatted data cell."""
    c = ws.cell(row=row, column=col, value=value)
    c.border = THIN_BORDER
    if fmt:
        c.number_format = fmt
    if font:
        c.font = font
    return c


def _apply_dynamic_pnl_color(ws, cell_range):
    """Applies conditional formatting to color positive values green and negative values red."""
    from openpyxl.formatting.rule import CellIsRule
    ws.conditional_formatting.add(cell_range, CellIsRule(operator='greaterThan', formula=['0'], font=GREEN_FONT))
    ws.conditional_formatting.add(cell_range, CellIsRule(operator='lessThan', formula=['0'], font=RED_FONT))


# ═══════════════════════════════════════════════════════════════════
#  Table 1: Portfolio KPIs
# ═══════════════════════════════════════════════════════════════════

def _write_kpi_table(ws, portfolio_df, overall_df, row):
    """Writes KPIs. Returns next free row."""
    row = _section_title(ws, row, 1, 'Portfolio Summary')
    row = _styled_header(ws, row, 1, ['Metric', 'Value'])

    start_row = row
    kpis = [
        ('Total Invested Value', f"=SUM(Current_Portfolio!$L$2:$L$1000)", INR_FMT),
        ('Total Current Value', f"=SUM(Current_Portfolio!$T$2:$T$1000)", INR_FMT),
        ('Unrealized PnL', f"=B{start_row+1}-B{start_row}", INR_FMT), # B6 - B5
        ('Realized PnL', f"=SUM(Overall_Portfolio!$P$2:$P$1000)", INR_FMT),
        ('Combined PnL', f"=B{start_row+2}+B{start_row+3}", INR_FMT), # B7 + B8
        ('Combined PnL %', f"=IF(B{start_row}>0, B{start_row+4}/B{start_row}, 0)", PCT_FMT), # B9 / B5
        ('Core Allocation (%)', f"=IF(B{start_row}>0, SUMIF(Current_Portfolio!$D$2:$D$1000, \"Core\", Current_Portfolio!$L$2:$L$1000) / B{start_row}, 0)", PCT_FMT),
        ('Satellite Allocation (%)', f"=IF(B{start_row}>0, SUMIF(Current_Portfolio!$D$2:$D$1000, \"Satellite\", Current_Portfolio!$L$2:$L$1000) / B{start_row}, 0)", PCT_FMT),
        ('Active Holdings', f"=COUNTA(Current_Portfolio!$A$2:$A$1000)", '0'),
        ('Total Stocks Traded', f"=COUNTA(Overall_Portfolio!$A$2:$A$1000)", '0'),
    ]

    for label, formula, fmt in kpis:
        _data_cell(ws, row, 1, label, font=LABEL_FONT)
        _data_cell(ws, row, 2, formula, fmt=fmt)
        if 'PnL' in label:
            _apply_dynamic_pnl_color(ws, f"B{row}")
        row += 1

    return row


# ═══════════════════════════════════════════════════════════════════
#  Table 1b: Performance Metrics (Win/Loss, Risk/Reward)
# ═══════════════════════════════════════════════════════════════════

def _write_performance_kpis(ws, start_row, col_start):
    """Writes Win/Loss and Risk/Reward metrics. Returns next free row."""
    row = _section_title(ws, start_row, col_start, 'Trading Performance')
    row = _styled_header(ws, row, col_start, ['Metric', 'Realized (Closed)', 'Unrealized (Open)'])
    
    col_e = get_column_letter(col_start + 1)
    col_f = get_column_letter(col_start + 2)
    
    metrics = [
        ('Winning Trades', 
         f"=COUNTIF(Overall_Portfolio!$P$2:$P$1000, \">0\")", 
         f"=COUNTIF(Current_Portfolio!$U$2:$U$1000, \">0\")", '0'),
        ('Losing Trades', 
         f"=COUNTIF(Overall_Portfolio!$P$2:$P$1000, \"<0\")", 
         f"=COUNTIF(Current_Portfolio!$U$2:$U$1000, \"<0\")", '0'),
        ('Win / Loss Ratio', 
         f"=IF({col_e}{row+1}>0, {col_e}{row}/{col_e}{row+1}, IF({col_e}{row}>0, \"No Loss\", 0))", 
         f"=IF({col_f}{row+1}>0, {col_f}{row}/{col_f}{row+1}, IF({col_f}{row}>0, \"No Loss\", 0))", '0.00'),
        ('Avg Win (₹)', 
         f"=IF({col_e}{row}>0, SUMIF(Overall_Portfolio!$P$2:$P$1000, \">0\")/{col_e}{row}, 0)", 
         f"=IF({col_f}{row}>0, SUMIF(Current_Portfolio!$U$2:$U$1000, \">0\")/{col_f}{row}, 0)", INR_FMT),
        ('Avg Loss (₹)', 
         f"=IF({col_e}{row+1}>0, SUMIF(Overall_Portfolio!$P$2:$P$1000, \"<0\")/{col_e}{row+1}, 0)", 
         f"=IF({col_f}{row+1}>0, SUMIF(Current_Portfolio!$U$2:$U$1000, \"<0\")/{col_f}{row+1}, 0)", INR_FMT),
        ('Risk / Reward Ratio', 
         f"=IF({col_e}{row+4}<0, ABS({col_e}{row+3}/{col_e}{row+4}), 0)", 
         f"=IF({col_f}{row+4}<0, ABS({col_f}{row+3}/{col_f}{row+4}), 0)", '0.00')
    ]
    
    for label, form_r, form_u, fmt in metrics:
        _data_cell(ws, row, col_start, label, font=LABEL_FONT)
        _data_cell(ws, row, col_start + 1, form_r, fmt=fmt)
        _data_cell(ws, row, col_start + 2, form_u, fmt=fmt)
        row += 1

    # Leave a blank row
    row += 1

    # Sub-headers for Advancing / Declining
    row = _styled_header(ws, row, col_start, ['TF Classfication', 'Advancing', 'Declining', 'Performance'])

    # Advancing / Declining sub-metrics using dynamic SUMPRODUCT and LET formulas
    sub_metrics = [
        ('Core (Previous month close)',
         '=SUMPRODUCT((Current_Portfolio!$D$2:$D$1000="Core")*(Current_Portfolio!$M$2:$M$1000>Current_Portfolio!$AA$2:$AA$1000)*(Current_Portfolio!$M$2:$M$1000>0))',
         '=SUMPRODUCT((Current_Portfolio!$D$2:$D$1000="Core")*(Current_Portfolio!$M$2:$M$1000<=Current_Portfolio!$AA$2:$AA$1000)*(Current_Portfolio!$M$2:$M$1000>0))',
         '=LET(curr, SUMIF(Current_Portfolio!$D$2:$D$1000, "Core", Current_Portfolio!$T$2:$T$1000), pmc, SUMPRODUCT((Current_Portfolio!$D$2:$D$1000="Core")*(Current_Portfolio!$F$2:$F$1000)*(Current_Portfolio!$AA$2:$AA$1000)), IF(pmc>0, (curr-pmc)/pmc, 0))'),
        ('Satellite (Previous week close)',
         '=SUMPRODUCT((Current_Portfolio!$D$2:$D$1000="Satellite")*(Current_Portfolio!$M$2:$M$1000>Current_Portfolio!$O$2:$O$1000)*(Current_Portfolio!$M$2:$M$1000>0))',
         '=SUMPRODUCT((Current_Portfolio!$D$2:$D$1000="Satellite")*(Current_Portfolio!$M$2:$M$1000<=Current_Portfolio!$O$2:$O$1000)*(Current_Portfolio!$M$2:$M$1000>0))',
         '=LET(curr, SUMIF(Current_Portfolio!$D$2:$D$1000, "Satellite", Current_Portfolio!$T$2:$T$1000), pwc, SUMPRODUCT((Current_Portfolio!$D$2:$D$1000="Satellite")*(Current_Portfolio!$F$2:$F$1000)*(Current_Portfolio!$O$2:$O$1000)), IF(pwc>0, (curr-pwc)/pwc, 0))'),
        ('Core (Previous Close)',
         '=SUMPRODUCT((Current_Portfolio!$D$2:$D$1000="Core")*(Current_Portfolio!$M$2:$M$1000>Current_Portfolio!$N$2:$N$1000)*(Current_Portfolio!$M$2:$M$1000>0))',
         '=SUMPRODUCT((Current_Portfolio!$D$2:$D$1000="Core")*(Current_Portfolio!$M$2:$M$1000<=Current_Portfolio!$N$2:$N$1000)*(Current_Portfolio!$M$2:$M$1000>0))',
         '=LET(curr, SUMIF(Current_Portfolio!$D$2:$D$1000, "Core", Current_Portfolio!$T$2:$T$1000), pdc, SUMPRODUCT((Current_Portfolio!$D$2:$D$1000="Core")*(Current_Portfolio!$F$2:$F$1000)*(Current_Portfolio!$N$2:$N$1000)), IF(pdc>0, (curr-pdc)/pdc, 0))'),
        ('Satellite (Previous Close)',
         '=SUMPRODUCT((Current_Portfolio!$D$2:$D$1000="Satellite")*(Current_Portfolio!$M$2:$M$1000>Current_Portfolio!$N$2:$N$1000)*(Current_Portfolio!$M$2:$M$1000>0))',
         '=SUMPRODUCT((Current_Portfolio!$D$2:$D$1000="Satellite")*(Current_Portfolio!$M$2:$M$1000<=Current_Portfolio!$N$2:$N$1000)*(Current_Portfolio!$M$2:$M$1000>0))',
         '=LET(curr, SUMIF(Current_Portfolio!$D$2:$D$1000, "Satellite", Current_Portfolio!$T$2:$T$1000), pdc, SUMPRODUCT((Current_Portfolio!$D$2:$D$1000="Satellite")*(Current_Portfolio!$F$2:$F$1000)*(Current_Portfolio!$N$2:$N$1000)), IF(pdc>0, (curr-pdc)/pdc, 0))')
    ]

    for label, adv_form, dec_form, perf_form in sub_metrics:
        _data_cell(ws, row, col_start, label, font=LABEL_FONT)
        _data_cell(ws, row, col_start + 1, adv_form, fmt='0')
        _data_cell(ws, row, col_start + 2, dec_form, fmt='0')
        _data_cell(ws, row, col_start + 3, perf_form, fmt='0.00%;[Red]-0.00%')
        row += 1

    return row


# ═══════════════════════════════════════════════════════════════════
#  Table 2: Top 5 Gainers / Bottom 5 Losers
# ═══════════════════════════════════════════════════════════════════

def _write_top_bottom_table(ws, df, portfolio_styles, row, classification_filter='Satellite'):
    title = f'Top 5 Gainers / Bottom 5 Losers ({classification_filter})'
    row = _section_title(ws, row, 1, title)
    row = _styled_header(ws, row, 1, ['Symbol', 'Classification', 'Invested (₹)', 'Current (₹)', 'PnL (₹)', 'PnL %'])
    
    # Write a simple placeholder formula that openpyxl can serialize cleanly
    ws.cell(row=row, column=1, value=f'="Gainers/Losers ({classification_filter}) (requires Excel 365)"')
    
    # Style and format the 10x6 spill area range while leaving cell values empty (None) to prevent SPILL errors
    for r in range(row, row + 10):
        for c_idx in range(1, 7):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = VALUE_FONT
            if c_idx in [3, 4, 5]:
                cell.number_format = INR_FMT
            elif c_idx == 6:
                cell.number_format = PCT_FMT

    # Apply positive/negative color highlighting to PnL (E) and PnL % (F) columns (columns 5 and 6)
    _apply_dynamic_pnl_color(ws, f"E{row}:E{row+9}")
    _apply_dynamic_pnl_color(ws, f"F{row}:F{row+9}")

    return row + 10


# ═══════════════════════════════════════════════════════════════════
#  Table 3: Stocks Nearest to Stop Loss
# ═══════════════════════════════════════════════════════════════════

def _write_nearest_sl_table(ws, df, portfolio_styles, row):
    """Writes stocks closest to SL. Returns next free row."""
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment

    row = _section_title(ws, row, 1, '⚠️ Stocks Nearest to Stop Loss')
    row = _styled_header(ws, row, 1, ['Symbol', 'Classification', 'LTP', 'SL', 'Diff (₹)', 'Diff (%)', 'Tranche'])

    # Write a simple placeholder formula that openpyxl can serialize cleanly
    ws.cell(row=row, column=1, value='="Stop Loss Table (requires Excel 365)"')

    # Style and format the 10x7 spill area range while leaving cell values empty (None) to prevent SPILL errors
    for r in range(row, row + 10):
        for c_idx in range(1, 8):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = VALUE_FONT
            if c_idx in [3, 4, 5]:
                cell.number_format = INR_FMT
            elif c_idx == 6:
                cell.number_format = PCT_FMT

    # Apply stop loss alerts to Diff (%) column (column 6 / F)
    ws.conditional_formatting.add(f"F{row}:F{row+9}", CellIsRule(operator='lessThan', formula=['0.05'], font=RED_FONT))
    ws.conditional_formatting.add(f"F{row}:F{row+9}", CellIsRule(operator='greaterThan', formula=['0.15'], font=GREEN_FONT))

    return row + 10


#  Table 4: Benchmark Returns
# ═══════════════════════════════════════════════════════════════════

def _write_benchmark_returns_table(ws, benchmark_returns: dict, row: int, price_cols: tuple, col_start=10):
    """Writes the Nifty Benchmark Returns table. Returns next free row."""
    if not benchmark_returns:
        return row
        
    row = _section_title(ws, row, col_start, '📈 Benchmark Returns (Since Invest Start)')
    row = _styled_header(ws, row, col_start, ['Index Tracker (ETF)', 'Start Price (₹)', 'LTP (₹)', 'Return %'])
    
    sym_col_let, ltp_col_let = price_cols or ('A', 'F')
    sym_col_let = sym_col_let or 'A'
    ltp_col_let = ltp_col_let or 'F'

    c_tracker = get_column_letter(col_start)
    c_start = get_column_letter(col_start + 1)
    c_ltp = get_column_letter(col_start + 2)
    c_ret = get_column_letter(col_start + 3)

    for name, data in benchmark_returns.items():
        _data_cell(ws, row, col_start, name, font=LABEL_FONT)
        _data_cell(ws, row, col_start + 1, data.get('Start_Price', 0), fmt=INR_FMT)
        
        # Dynamic index lookup from Price_Update sheet
        _data_cell(ws, row, col_start + 2, f"=IFERROR(INDEX(Price_Update!${ltp_col_let}:${ltp_col_let}, MATCH({c_tracker}{row}, Price_Update!${sym_col_let}:${sym_col_let}, 0)), {data.get('LTP', 0)})", fmt=INR_FMT)
        
        _data_cell(ws, row, col_start + 3, f"=({c_ltp}{row}-{c_start}{row})/{c_start}{row}", fmt=PCT_FMT)
        _apply_dynamic_pnl_color(ws, f"{c_ret}{row}")
        row += 1
        
    return row


# ═══════════════════════════════════════════════════════════════════
#  Table 4: Cap-wise Allocation (RIGHT SIDE)
# ═══════════════════════════════════════════════════════════════════

def _write_cap_allocation(ws, df, overall_df, row, col_start=1):
    """Cap Allocation & PnL Breakdown. Returns next free row."""
    row = _section_title(ws, row, col_start, 'Cap Allocation & PnL Breakdown')
    row = _styled_header(ws, row, col_start, ['Cap', 'Invested (₹)', '% of Total', 'Returns', 'Realized (₹)', 'Un-Realized (₹)', 'Total PnL (₹)'])

    if 'Cap' not in overall_df.columns:
        _data_cell(ws, row, col_start, 'No data')
        return row + 1

    # Group by Cap in overall_df to include all traded categories (active and closed)
    grouped = overall_df.groupby('Cap').agg(
        Realized=('Realized_PnL', 'sum')
    ).sort_index()
    
    start_row = row
    num_caps = len(grouped)
    total_row_idx = start_row + num_caps

    c_cap = get_column_letter(col_start)
    c_inv = get_column_letter(col_start + 1)
    c_pct = get_column_letter(col_start + 2)
    c_ret = get_column_letter(col_start + 3)
    c_real = get_column_letter(col_start + 4)
    c_unreal = get_column_letter(col_start + 5)
    c_tot = get_column_letter(col_start + 6)

    for idx, (label, row_data) in enumerate(grouped.iterrows()):
        r = start_row + idx
        _data_cell(ws, row, col_start, str(label), font=LABEL_FONT)
        _data_cell(ws, row, col_start + 1, f"=SUMIF(Current_Portfolio!$B$2:$B$1000, \"{label}\", Current_Portfolio!$L$2:$L$1000)", fmt=INR_FMT)
        _data_cell(ws, row, col_start + 2, f"={c_inv}{r}/${c_inv}${total_row_idx}", fmt=PCT_FMT)
        _data_cell(ws, row, col_start + 3, f"=IF({c_inv}{r}>0, (SUMIF(Current_Portfolio!$B$2:$B$1000, \"{label}\", Current_Portfolio!$T$2:$T$1000) - {c_inv}{r}) / {c_inv}{r}, 0)", fmt=PCT_FMT)
        _data_cell(ws, row, col_start + 4, f"=SUMIF(Overall_Portfolio!$B$2:$B$1000, \"{label}\", Overall_Portfolio!$P$2:$P$1000)", fmt=INR_FMT)
        _data_cell(ws, row, col_start + 5, f"=SUMIF(Current_Portfolio!$B$2:$B$1000, \"{label}\", Current_Portfolio!$U$2:$U$1000)", fmt=INR_FMT)
        _data_cell(ws, row, col_start + 6, f"={c_real}{r}+{c_unreal}{r}", fmt=INR_FMT)
        _apply_dynamic_pnl_color(ws, f"{c_ret}{row}:{c_tot}{row}")
        row += 1

    # Total row
    _data_cell(ws, row, col_start, 'TOTAL', font=LABEL_FONT)
    _data_cell(ws, row, col_start + 1, f"=SUM({c_inv}{start_row}:{c_inv}{row-1})", fmt=INR_FMT, font=LABEL_FONT)
    _data_cell(ws, row, col_start + 2, 1.0, fmt=PCT_FMT, font=LABEL_FONT)
    _data_cell(ws, row, col_start + 3, f"=IF({c_inv}{row}>0, (SUM(Current_Portfolio!$T$2:$T$1000) - {c_inv}{row}) / {c_inv}{row}, 0)", fmt=PCT_FMT)
    _data_cell(ws, row, col_start + 4, f"=SUM({c_real}{start_row}:{c_real}{row-1})", fmt=INR_FMT, font=LABEL_FONT)
    _data_cell(ws, row, col_start + 5, f"=SUM({c_unreal}{start_row}:{c_unreal}{row-1})", fmt=INR_FMT, font=LABEL_FONT)
    _data_cell(ws, row, col_start + 6, f"=SUM({c_tot}{start_row}:{c_tot}{row-1})", fmt=INR_FMT, font=LABEL_FONT)
    _apply_dynamic_pnl_color(ws, f"{c_ret}{row}:{c_tot}{row}")
    row += 1

    return row


# ═══════════════════════════════════════════════════════════════════
#  Table 5: Core & Satellite Distribution (RIGHT SIDE)
# ═══════════════════════════════════════════════════════════════════

def _write_classification_allocation(ws, df, row, col_start=1):
    """Core & Satellite distribution. Returns next free row."""
    row = _section_title(ws, row, col_start, 'Core & Satellite Distribution')
    row = _styled_header(ws, row, col_start, ['Classification', 'Invested (₹)', '% of Total', 'Returns', 'Realized (₹)', 'Un-Realized (₹)'])

    if 'TF_Classification' not in df.columns or 'Invested_Value' not in df.columns or 'Current_Value' not in df.columns:
        _data_cell(ws, row, col_start, 'No data')
        return row + 1

    grouped = df.groupby('TF_Classification')[['Invested_Value', 'Current_Value']].sum()
    grouped = grouped[grouped.index != '']
    grouped = grouped.sort_values(by='Invested_Value', ascending=False)
    
    start_row = row
    num_items = len(grouped)
    total_row_idx = start_row + num_items

    c_class = get_column_letter(col_start)
    c_inv = get_column_letter(col_start + 1)
    c_pct = get_column_letter(col_start + 2)
    c_ret = get_column_letter(col_start + 3)
    c_real = get_column_letter(col_start + 4)
    c_unreal = get_column_letter(col_start + 5)

    for idx, (label, row_data) in enumerate(grouped.iterrows()):
        r = start_row + idx
        _data_cell(ws, row, col_start, str(label), font=LABEL_FONT)
        _data_cell(ws, row, col_start + 1, f"=SUMIF(Current_Portfolio!$D$2:$D$1000, \"{label}\", Current_Portfolio!$L$2:$L$1000)", fmt=INR_FMT)
        _data_cell(ws, row, col_start + 2, f"={c_inv}{r}/${c_inv}${total_row_idx}", fmt=PCT_FMT)
        _data_cell(ws, row, col_start + 3, f"=(SUMIF(Current_Portfolio!$D$2:$D$1000, \"{label}\", Current_Portfolio!$T$2:$T$1000) - {c_inv}{r}) / {c_inv}{r}", fmt=PCT_FMT)
        _data_cell(ws, row, col_start + 4, f"=SUMIF(Overall_Portfolio!$D$2:$D$1000, \"{label}\", Overall_Portfolio!$P$2:$P$1000)", fmt=INR_FMT)
        _data_cell(ws, row, col_start + 5, f"=SUMIF(Current_Portfolio!$D$2:$D$1000, \"{label}\", Current_Portfolio!$U$2:$U$1000)", fmt=INR_FMT)
        _apply_dynamic_pnl_color(ws, f"{c_ret}{row}:{c_unreal}{row}")
        row += 1

    _data_cell(ws, row, col_start, 'TOTAL', font=LABEL_FONT)
    _data_cell(ws, row, col_start + 1, f"=SUM({c_inv}{start_row}:{c_inv}{row-1})", fmt=INR_FMT, font=LABEL_FONT)
    _data_cell(ws, row, col_start + 2, 1.0, fmt=PCT_FMT, font=LABEL_FONT)
    _data_cell(ws, row, col_start + 3, f"=(SUM(Current_Portfolio!$T$2:$T$1000) - {c_inv}{row}) / {c_inv}{row}", fmt=PCT_FMT)
    _data_cell(ws, row, col_start + 4, f"=SUM({c_real}{start_row}:{c_real}{row-1})", fmt=INR_FMT, font=LABEL_FONT)
    _data_cell(ws, row, col_start + 5, f"=SUM({c_unreal}{start_row}:{c_unreal}{row-1})", fmt=INR_FMT, font=LABEL_FONT)
    _apply_dynamic_pnl_color(ws, f"{c_ret}{row}:{c_unreal}{row}")
    row += 1

    return row


# ═══════════════════════════════════════════════════════════════════
#  Table 5: Sector Allocation & PnL Breakdown (RIGHT SIDE)
# ═══════════════════════════════════════════════════════════════════

def _write_sector_allocation(ws, df, overall_df, row, col_start=1):
    """Sector Allocation & PnL Breakdown. Returns next free row."""
    row = _section_title(ws, row, col_start, 'Sector Allocation & PnL Breakdown')
    row = _styled_header(ws, row, col_start, ['Sector', 'Core (₹)', 'Satellite (₹)', 'Total Invested (₹)', '% of Total', 'Realized (₹)', 'Un-Realized (₹)', 'Total PnL (₹)'])

    if 'TF_Sector' not in overall_df.columns:
        _data_cell(ws, row, col_start, 'No data')
        return row + 1

    # Extract unique sectors from overall_df (excluding blanks/NaNs)
    sectors = overall_df['TF_Sector'].unique()
    sectors = [s for s in sectors if s and str(s).strip() != '']
    
    # Compute active invested value for sorting
    active_invested = {}
    for s in sectors:
        val = 0
        if 'TF_Sector' in df.columns and 'Invested_Value' in df.columns:
            val = df[df['TF_Sector'] == s]['Invested_Value'].sum()
        active_invested[s] = val
        
    # Sort sectors: active ones first (descending by invested value), then others alphabetically
    sectors.sort(key=lambda s: (-active_invested[s], str(s).lower()))
    
    start_row = row
    num_sectors = len(sectors)
    total_row_idx = start_row + num_sectors

    c_sec = get_column_letter(col_start)
    c_core = get_column_letter(col_start + 1)
    c_sat = get_column_letter(col_start + 2)
    c_tot_inv = get_column_letter(col_start + 3)
    c_pct = get_column_letter(col_start + 4)
    c_real = get_column_letter(col_start + 5)
    c_unreal = get_column_letter(col_start + 6)
    c_tot_pnl = get_column_letter(col_start + 7)

    for idx, sector in enumerate(sectors):
        r = start_row + idx
        _data_cell(ws, row, col_start, str(sector), font=LABEL_FONT)
        _data_cell(ws, row, col_start + 1, f"=SUMIFS(Current_Portfolio!$L$2:$L$1000, Current_Portfolio!$C$2:$C$1000, \"{sector}\", Current_Portfolio!$D$2:$D$1000, \"*Core*\")", fmt=INR_FMT)
        _data_cell(ws, row, col_start + 2, f"=SUMIFS(Current_Portfolio!$L$2:$L$1000, Current_Portfolio!$C$2:$C$1000, \"{sector}\", Current_Portfolio!$D$2:$D$1000, \"*Satellite*\")", fmt=INR_FMT)
        _data_cell(ws, row, col_start + 3, f"={c_core}{r}+{c_sat}{r}", fmt=INR_FMT)
        _data_cell(ws, row, col_start + 4, f"={c_tot_inv}{r}/${c_tot_inv}${total_row_idx}", fmt=PCT_FMT)
        _data_cell(ws, row, col_start + 5, f"=SUMIF(Overall_Portfolio!$C$2:$C$1000, \"{sector}\", Overall_Portfolio!$P$2:$P$1000)", fmt=INR_FMT)
        _data_cell(ws, row, col_start + 6, f"=SUMIF(Current_Portfolio!$C$2:$C$1000, \"{sector}\", Current_Portfolio!$U$2:$U$1000)", fmt=INR_FMT)
        _data_cell(ws, row, col_start + 7, f"={c_real}{r}+{c_unreal}{r}", fmt=INR_FMT)
        _apply_dynamic_pnl_color(ws, f"{c_real}{row}:{c_tot_pnl}{row}")
        row += 1

    # Total row
    _data_cell(ws, row, col_start, 'TOTAL', font=LABEL_FONT)
    _data_cell(ws, row, col_start + 1, f"=SUM({c_core}{start_row}:{c_core}{row-1})", fmt=INR_FMT, font=LABEL_FONT)
    _data_cell(ws, row, col_start + 2, f"=SUM({c_sat}{start_row}:{c_sat}{row-1})", fmt=INR_FMT, font=LABEL_FONT)
    _data_cell(ws, row, col_start + 3, f"=SUM({c_tot_inv}{start_row}:{c_tot_inv}{row-1})", fmt=INR_FMT, font=LABEL_FONT)
    _data_cell(ws, row, col_start + 4, 1.0, fmt=PCT_FMT, font=LABEL_FONT)
    _data_cell(ws, row, col_start + 5, f"=SUM({c_real}{start_row}:{c_real}{row-1})", fmt=INR_FMT, font=LABEL_FONT)
    _data_cell(ws, row, col_start + 6, f"=SUM({c_unreal}{start_row}:{c_unreal}{row-1})", fmt=INR_FMT, font=LABEL_FONT)
    _data_cell(ws, row, col_start + 7, f"=SUM({c_tot_pnl}{start_row}:{c_tot_pnl}{row-1})", fmt=INR_FMT, font=LABEL_FONT)
    _apply_dynamic_pnl_color(ws, f"{c_real}{row}:{c_tot_pnl}{row}")
    row += 1

    return row


# ═══════════════════════════════════════════════════════════════════
#  Table 7: Tranche Distribution (RIGHT SIDE)
# ═══════════════════════════════════════════════════════════════════

def _write_tranche_distribution(ws, df, row, col_start=10):
    """Tranche distribution counts split by Core vs Satellite. Returns next free row."""
    row = _section_title(ws, row, col_start, 'Tranche Distribution')
    row = _styled_header(ws, row, col_start, ['Tranche', 'Core Count', 'Satellite Count', 'Total Count', '% of Holdings'])

    if 'Latest_Tranche' not in df.columns or len(df) == 0:
        _data_cell(ws, row, col_start, 'No data')
        return row + 1

    # Filter out empty or null tranche values
    valid_df = df[df['Latest_Tranche'].notna() & df['Latest_Tranche'].str.strip().ne('')].copy()
    if valid_df.empty:
        _data_cell(ws, row, col_start, 'No tranche data available')
        return row + 1

    # Group by Tranche and Classification
    tranche_groups = valid_df.groupby(['Latest_Tranche', 'TF_Classification']).size().unstack(fill_value=0)
    
    # Ensure both Core and Satellite columns exist
    if 'Core' not in tranche_groups.columns:
        tranche_groups['Core'] = 0
    if 'Satellite' not in tranche_groups.columns:
        tranche_groups['Satellite'] = 0
        
    tranche_groups['Total'] = tranche_groups['Core'] + tranche_groups['Satellite']
    tranche_groups = tranche_groups.sort_index()
    
    total_holdings = tranche_groups['Total'].sum()

    for label, r_data in tranche_groups.iterrows():
        core_cnt = int(r_data['Core'])
        sat_cnt = int(r_data['Satellite'])
        tot_cnt = int(r_data['Total'])
        
        _data_cell(ws, row, col_start, str(label), font=LABEL_FONT)
        _data_cell(ws, row, col_start + 1, core_cnt)
        _data_cell(ws, row, col_start + 2, sat_cnt)
        _data_cell(ws, row, col_start + 3, tot_cnt)
        _data_cell(ws, row, col_start + 4, tot_cnt / total_holdings if total_holdings > 0 else 0, fmt=PCT_FMT)
        row += 1

    # Add TOTAL row
    _data_cell(ws, row, col_start, 'TOTAL', font=LABEL_FONT)
    _data_cell(ws, row, col_start + 1, int(tranche_groups['Core'].sum()))
    _data_cell(ws, row, col_start + 2, int(tranche_groups['Satellite'].sum()))
    _data_cell(ws, row, col_start + 3, int(total_holdings))
    _data_cell(ws, row, col_start + 4, 1.0, fmt=PCT_FMT)
    row += 1

    return row


# ═══════════════════════════════════════════════════════════════════
#  Table 8: Holding Period Distribution (RIGHT SIDE)
# ═══════════════════════════════════════════════════════════════════

def _write_holding_distribution(ws, df, row, col_start=10):
    """Holding period bucket counts. Returns next free row."""
    row = _section_title(ws, row, col_start, 'Holding Period Distribution')
    row = _styled_header(ws, row, col_start, ['Period', 'Count', '% of Holdings'])

    total = len(df)
    if 'Holding_Period' not in df.columns or total == 0:
        _data_cell(ws, row, col_start, 'No data')
        return row + 1

    hp = df['Holding_Period']
    buckets = [
        ('0 - 30 days', int(((hp >= 0) & (hp <= 30)).sum())),
        ('31 - 90 days', int(((hp > 30) & (hp <= 90)).sum())),
        ('91 - 180 days', int(((hp > 90) & (hp <= 180)).sum())),
        ('180+ days', int((hp > 180).sum())),
    ]

    for label, cnt in buckets:
        _data_cell(ws, row, col_start, label, font=LABEL_FONT)
        _data_cell(ws, row, col_start + 1, cnt)
        _data_cell(ws, row, col_start + 2, cnt / total if total > 0 else 0, fmt=PCT_FMT)
        row += 1

    return row


# ═══════════════════════════════════════════════════════════════════
#  Table 10: Corporate Actions (LEFT SIDE)
# ═══════════════════════════════════════════════════════════════════

def _write_corporate_actions(ws, portfolio_df, overall_df, portfolio_styles, row):
    """Writes stocks with splits/bonuses since purchase. Returns next free row."""
    row = _section_title(ws, row, 1, '🔔 Corporate Actions (Splits / Bonus)')
    row = _styled_header(ws, row, 1, ['Symbol', 'Split / Bonus Details', 'Adj Required'])

    # Filter for rows with split info in current portfolio only
    all_stocks = portfolio_df.copy()

    if 'Split_Info' not in all_stocks.columns or len(all_stocks) == 0:
        _data_cell(ws, row, 1, 'No corporate actions detected')
        return row + 1

    actions = all_stocks[all_stocks['Split_Info'] != ''].sort_values('Symbol')

    if len(actions) == 0:
        _data_cell(ws, row, 1, 'No splits/bonuses detected since purchase')
        return row + 1

    for _, s in actions.iterrows():
        symbol = s.get('Symbol', '')
        classification = s.get('TF_Classification', '')
        font, fill = _get_portfolio_style(symbol, portfolio_styles)
        
        c = _data_cell(ws, row, 1, symbol, font=font)
        if fill: c.fill = fill
        
        _data_cell(ws, row, 2, s.get('Split_Info', ''))
        adj = s.get('Adj_Required', 'No')
        _data_cell(ws, row, 3, adj,
                   font=RED_FONT if adj == 'Yes' else None)
        row += 1

    return row


# ═══════════════════════════════════════════════════════════════════
#  Table 11: Top 5 Cheat Stocks (LEFT SIDE)
# ═══════════════════════════════════════════════════════════════════

def _write_top_cheats_table(ws, df, portfolio_styles, row, classification_filter='Satellite'):
    """Writes Top 5 Cheat Stocks by holding period in descending order. Returns next free row."""
    title = f'Top 5 Cheat Stocks ({classification_filter}) by Holding Period'
    row = _section_title(ws, row, 1, title)
    row = _styled_header(ws, row, 1, ['Stock Name', 'Cheat', 'Holding Period (Days)', 'Invested Value', 'Current Value', 'Return %', 'XIRR'])

    if 'Latest_Tranche' not in df.columns or len(df) == 0:
        _data_cell(ws, row, 1, 'No cheat stocks found')
        return row + 1

    # Filter for active positions that are Cheat stocks of the specified classification
    mask = df['Latest_Tranche'].str.startswith('Cheat', na=False)
    if 'TF_Classification' in df.columns:
        mask = mask & (df['TF_Classification'] == classification_filter)
    cheat_df = df[mask].copy()

    if cheat_df.empty:
        _data_cell(ws, row, 1, f'No active {classification_filter} cheat stocks in portfolio')
        return row + 1

    # Sort by holding period desc and take top 5
    top_cheats = cheat_df.sort_values(by='Holding_Period', ascending=False).head(5)

    col_keys = ['Symbol', 'Latest_Tranche', 'Holding_Period', 'Invested_Value', 'Current_Value', 'Return_Pct', 'XIRR']
    formats = [None, None, '0', INR_FMT, INR_FMT, '0.00%', '0.00%']

    for _, s in top_cheats.iterrows():
        for ci, (key, fmt) in enumerate(zip(col_keys, formats), 1):
            val = s.get(key, '')
            font = VALUE_FONT
            fill = None
            
            if key == 'Symbol':
                classification = s.get('TF_Classification', 'Satellite')
                font, fill = _get_portfolio_style(val, portfolio_styles)
            
            # Apply dynamic color formatting for Return_Pct and XIRR
            if key in ['Return_Pct', 'XIRR']:
                try:
                    fval = float(val)
                    if fval >= 0:
                        font = GREEN_FONT
                    else:
                        font = RED_FONT
                except Exception:
                    pass
            
            c = _data_cell(ws, row, ci, val, fmt=fmt, font=font)
            if fill and key == 'Symbol':
                c.fill = fill
        row += 1

    return row


def _write_daily_losers_table(ws, df, portfolio_styles, row, classification_filter='Satellite'):
    """Writes Top 10 Daily Losers for the given classification. Returns next free row."""
    from openpyxl.styles import Alignment
    
    title = f'Top 10 underperforming {classification_filter} Stocks (LTP < Prev Day Close)'
    row = _section_title(ws, row, 1, title)
    row = _styled_header(ws, row, 1, [
        'Symbol', 'Classification', 'Invested Value (₹)', 'Prev Day Close (₹)',
        'LTP (₹)', 'Change (₹)', 'Change (%)', 'Current Return %'
    ])
    
    # Write placeholder formula
    ws.cell(row=row, column=1, value=f'="Daily Losers ({classification_filter}) (requires Excel 365)"')
    
    # Style and format the 10x8 spill area range while leaving cell values empty (None) to prevent SPILL errors
    for r in range(row, row + 10):
        for c_idx in range(1, 9):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = VALUE_FONT
            if c_idx in [3, 4, 5, 6]:
                cell.number_format = INR_FMT
            elif c_idx in [7, 8]:
                cell.number_format = PCT_FMT

    # Apply red/green dynamic highlighting to Change (₹) (F), Change (%) (G) and Current Return % (H) (columns 6, 7, 8)
    _apply_dynamic_pnl_color(ws, f"F{row}:F{row+9}")
    _apply_dynamic_pnl_color(ws, f"G{row}:G{row+9}")
    _apply_dynamic_pnl_color(ws, f"H{row}:H{row+9}")

    return row + 10


# ═══════════════════════════════════════════════════════════════════
#  Table 12: Top 10 Movers (Current Portfolio) & Table 13: Watchlist
# ═══════════════════════════════════════════════════════════════════

def _write_portfolio_movers_table(ws, df, watchlist_df, portfolio_styles, row, col_start=10):
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Font, PatternFill
    
    start_row = row
    row = _section_title(ws, row, col_start, 'Top 10 Movers (Current Portfolio)')
    row = _styled_header(ws, row, col_start, ['Symbol', 'Prev Week Close', 'LTP', '% Change'])
    
    # Write a simple placeholder formula that openpyxl can serialize cleanly
    ws.cell(row=row, column=col_start, value='="Movers (requires Excel 365)"')
    
    # Style and format the 10x4 spill area range while leaving cell values empty (None) to prevent SPILL errors
    for r in range(row, row + 10):
        for c_idx in range(col_start, col_start + 4):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = VALUE_FONT
            if c_idx in [col_start + 1, col_start + 2]:
                cell.number_format = INR_FMT
            elif c_idx == col_start + 3:
                cell.number_format = PCT_FMT

    # Apply positive/negative color highlighting to % Change column
    c_pct = get_column_letter(col_start + 3)
    _apply_dynamic_pnl_color(ws, f"{c_pct}{row}:{c_pct}{row+9}")

    return row + 10

def _write_watchlist_movers_table(ws, watchlist_df, portfolio_df, latest_colors, row, col_start=10):
    from openpyxl.styles import Alignment, Font
    
    row = _section_title(ws, row, col_start, 'Top 10 Movers (Watchlist Only)')
    row = _styled_header(ws, row, col_start, ['Symbol', 'Prev Week Close', 'Price', '% Change'])
    
    # Write a simple placeholder formula that openpyxl can serialize cleanly
    ws.cell(row=row, column=col_start, value='="Watchlist Movers (requires Excel 365)"')
    
    # Style and format the 10x4 spill area range while leaving cell values empty (None) to prevent SPILL errors
    for r in range(row, row + 10):
        for c_idx in range(col_start, col_start + 4):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = VALUE_FONT
            if c_idx in [col_start + 1, col_start + 2]:
                cell.number_format = INR_FMT
            elif c_idx == col_start + 3:
                cell.number_format = PCT_FMT

    # Apply positive/negative color highlighting to % Change column
    c_pct = get_column_letter(col_start + 3)
    _apply_dynamic_pnl_color(ws, f"{c_pct}{row}:{c_pct}{row+9}")

    return row + 10
