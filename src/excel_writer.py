"""
Excel Writer Module
===================

Handles saving DataFrames to the master Excel workbook with proper formatting.

Features:
    - Writes/replaces specific sheets (Raw_Tradebook, Transaction,
      Current_Portfolio, Overall_Portfolio) without destroying user-created
      custom sheets or charts.
    - Applies INR currency formatting (₹) to monetary columns.
    - Applies percentage formatting to PnL percentage columns.
    - Converts large integer ID columns from string back to native Excel
      numbers with '0' format to prevent scientific notation display.

Uses openpyxl engine with mode='a' (append) and if_sheet_exists='replace'
when the workbook already exists.
"""

import os

import pandas as pd

from .dashboard import create_dashboard


def _find_price_update_columns_from_file(output_file: str) -> tuple:
    """
    Locates the Symbol and LTP columns in the existing Price_Update sheet.
    Returns (symbol_col_letter, ltp_col_letter) or (None, None) if not found.
    """
    import os
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    if not os.path.exists(output_file):
        return None, None, None
    try:
        wb = load_workbook(output_file, read_only=True)
        if 'Price_Update' not in wb.sheetnames:
            wb.close()
            return None, None, None
        ws = wb['Price_Update']
        symbol_col_letter = None
        ltp_col_letter = None
        prev_close_col_letter = None
        
        # Check the first 10 rows and 100 columns
        for row_idx in range(1, 11):
            row_vals = []
            for col_idx in range(1, 100):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    header = str(cell_val).strip().upper()
                    if header in ['SYMBOL', 'STOCK SYMBOL', 'STOCK_SYMBOL']:
                        symbol_col_letter = get_column_letter(col_idx)
                    elif header in ['LTP', 'LAST TRADED PRICE', 'LAST_TRADED_PRICE', 'PRICE']:
                        ltp_col_letter = get_column_letter(col_idx)
                    elif header in ['PREVIOUS DAY CLOSE', 'PREV DAY CLOSE', 'PREV_DAY_CLOSE']:
                        prev_close_col_letter = get_column_letter(col_idx)
            if symbol_col_letter and ltp_col_letter and prev_close_col_letter:
                break
                
        wb.close()
        return symbol_col_letter, ltp_col_letter, prev_close_col_letter
    except Exception:
        return None, None, None

def _convert_portfolios_to_formulas(portfolio_df: pd.DataFrame, overall_df: pd.DataFrame, output_file: str) -> tuple:
    """
    Converts LTP and its dependent calculations to standard Excel formulas.
    Returns (portfolio_write_df, overall_write_df).
    """
    sym_col_let, ltp_col_let, prev_close_col_let = _find_price_update_columns_from_file(output_file)
    sym_col_let = sym_col_let or 'A'
    ltp_col_let = ltp_col_let or 'F'
    prev_close_col_let = prev_close_col_let or 'I'
    
    port_f = portfolio_df.copy()
    over_f = overall_df.copy()
    
    # 1. Update Current_Portfolio
    if len(port_f) > 0:
        # Convert columns to object type to support string formulas
        for col in ['LTP', 'Prev_Day_Close', 'LTP_SL_Diff', 'LTP_SL_Diff_Pct', 'Current_Value', 'Unrealized_PnL', 'Return_Pct', 'Trend']:
            if col in port_f.columns:
                port_f[col] = port_f[col].astype(object)
                
        for idx, (p_idx, row) in enumerate(portfolio_df.iterrows()):
            r = idx + 2 # row number in Excel (1-based, plus 1 header row)
            fallback_ltp = row['LTP']
            if pd.isna(fallback_ltp):
                fallback_ltp = 0.0
            fallback_prev_close = row.get('Prev_Day_Close', 0.0)
            if pd.isna(fallback_prev_close):
                fallback_prev_close = 0.0
                
            port_f.at[p_idx, 'LTP'] = f"=IFERROR(INDEX(Price_Update!${ltp_col_let}:${ltp_col_let}, MATCH(A{r}, Price_Update!${sym_col_let}:${sym_col_let}, 0)), {fallback_ltp})"
            if 'Prev_Day_Close' in port_f.columns:
                port_f.at[p_idx, 'Prev_Day_Close'] = f"=IFERROR(INDEX(Price_Update!${prev_close_col_let}:${prev_close_col_let}, MATCH(A{r}, Price_Update!${sym_col_let}:${sym_col_let}, 0)), {fallback_prev_close})"
            port_f.at[p_idx, 'Trend'] = f'=IF(M{r}>N{r}, "▲", IF(M{r}<N{r}, "▼", "─"))'
            port_f.at[p_idx, 'LTP_SL_Diff'] = f"=M{r}-I{r}"
            port_f.at[p_idx, 'LTP_SL_Diff_Pct'] = f"=IF(M{r}>0, (M{r}-I{r})/M{r}, 0)"
            port_f.at[p_idx, 'Current_Value'] = f"=F{r}*M{r}"
            port_f.at[p_idx, 'Unrealized_PnL'] = f"=T{r}-L{r}"
            port_f.at[p_idx, 'Return_Pct'] = f"=IF(L{r}>0, U{r}/L{r}, 0)"
            
    # 2. Update Overall_Portfolio
    if len(over_f) > 0:
        for col in ['LTP', 'Current_Value', 'Unrealized_PnL', 'Total_PnL', 'Total_PnL_Percentage']:
            if col in over_f.columns:
                over_f[col] = over_f[col].astype(object)
                
        for idx, (o_idx, row) in enumerate(overall_df.iterrows()):
            r = idx + 2 # row number in Excel
            fallback_ltp = row['LTP']
            if pd.isna(fallback_ltp):
                fallback_ltp = 0.0
                
            over_f.at[o_idx, 'LTP'] = f"=IFERROR(INDEX(Price_Update!${ltp_col_let}:${ltp_col_let}, MATCH(A{r}, Price_Update!${sym_col_let}:${sym_col_let}, 0)), {fallback_ltp})"
            over_f.at[o_idx, 'Current_Value'] = f"=L{r}*N{r}"
            over_f.at[o_idx, 'Unrealized_PnL'] = f"=O{r}-M{r}"
            over_f.at[o_idx, 'Total_PnL'] = f"=P{r}+Q{r}"
            over_f.at[o_idx, 'Total_PnL_Percentage'] = f"=IF(G{r}>0, R{r}/G{r}, 0)"
            
    return port_f, over_f


def save_workbook(
    df: pd.DataFrame,
    grouped_df: pd.DataFrame,
    portfolio_df: pd.DataFrame,
    overall_df: pd.DataFrame,
    output_file: str,
    benchmark_returns: dict = None
) -> None:
    """
    Saves all computed DataFrames to the master Excel workbook.

    Writes four sheets:
        - Raw_Tradebook:     The full, merged raw trade data.
        - Transaction:       Grouped and labeled trade summary.
        - Current_Portfolio:  Active holdings with LTP, EMAs, SL.
        - Overall_Portfolio:  Complete PnL report for all symbols.

    If the workbook already exists, sheets are replaced in-place (mode='a')
    so that any user-created custom sheets are preserved.

    Args:
        df:            The merged raw trade DataFrame.
        grouped_df:    The grouped transaction DataFrame.
        portfolio_df:  The current portfolio DataFrame.
        overall_df:    The overall portfolio DataFrame.
        output_file:   Path to the output Excel file.
    """
    print(f"Saving transformed data to {output_file}...")
    
    # Convert portfolios to dynamic Excel formulas for writing
    portfolio_write_df, overall_write_df = _convert_portfolios_to_formulas(portfolio_df, overall_df, output_file)
    
    # 1. Parse latest core watchlist trends and load satellite watchlist strictly from output_file
    latest_core_trends = {}
    watchlist_df = None
    has_watchlist = False
    if os.path.exists(output_file):
        try:
            import openpyxl as op
            wb_test = op.load_workbook(output_file, read_only=True)
            has_core_sheet = any(name in wb_test.sheetnames for name in ['Core_Watchlist', 'Core_Portfolio'])
            has_watchlist = 'Satellite_Watchlist' in wb_test.sheetnames
            wb_test.close()
            
            if has_core_sheet or has_watchlist:
                with pd.ExcelFile(output_file) as xls:
                    if has_core_sheet:
                        # Find sheet name
                        sheet_name = None
                        for name in ['Core_Watchlist', 'Core_Portfolio']:
                            if name in xls.sheet_names:
                                sheet_name = name
                                break
                        if sheet_name:
                            core_df = pd.read_excel(xls, sheet_name=sheet_name)
                            if 'Company' in core_df.columns and 'Trend Status' in core_df.columns:
                                # Enforce strict latest month filtering
                                if 'Month' in core_df.columns:
                                    core_df['Month'] = pd.to_datetime(core_df['Month'], errors='coerce')
                                    latest_month = core_df['Month'].max()
                                    core_df = core_df[core_df['Month'] == latest_month]
                                
                                core_df = core_df.dropna(subset=['Company', 'Trend Status'])
                                core_df['Company'] = core_df['Company'].astype(str).str.strip().str.upper()
                                core_df['Trend Status'] = core_df['Trend Status'].astype(str).str.strip()
                                latest_core_trends = core_df.drop_duplicates(subset=['Company']).set_index('Company')['Trend Status'].to_dict()
                    
                    if has_watchlist:
                        watchlist_df = pd.read_excel(xls, sheet_name='Satellite_Watchlist')
        except Exception as e:
            print(f"Note: Watchlists could not be loaded in save_workbook: {e}")
    
    # Determine if we should use the hybrid xlwings approach to preserve Excel STOCKS rich data types
    use_xlwings = False
    if os.path.exists(output_file):

        try:
            import xlwings as xw
            use_xlwings = True
        except ImportError:
            print("xlwings is not installed. Falling back to openpyxl (which converts Excel Stocks types to plain text).")

    # Disable xlwings under unit tests or when explicitly requested to prevent headless background hangs
    import sys
    is_testing = 'unittest' in sys.modules
    no_xlwings_flag = '--no-xlwings' in sys.argv or os.environ.get('NO_XLWINGS') == '1'
    if is_testing or no_xlwings_flag:
        use_xlwings = False

    if use_xlwings:
        print("Using xlwings hybrid approach to preserve STOCKS data types and live formulas...")
        temp_file = "temp_transformed.xlsx"
        if os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except Exception:
                pass

        try:
            # 1. Generate updated sheets in a temporary workbook using openpyxl
            with pd.ExcelWriter(temp_file, engine='openpyxl', mode='w') as writer:
                df.to_excel(writer, sheet_name='Raw_Tradebook', index=False)
                grouped_df.to_excel(writer, sheet_name='Transaction', index=False)
                portfolio_write_df.to_excel(writer, sheet_name='Current_Portfolio', index=False)
                overall_write_df.to_excel(writer, sheet_name='Overall_Portfolio', index=False)

                # Create a dummy Price_Update sheet so local formulas don't break during copy
                writer.book.create_sheet('Price_Update')

                # --- Format ID columns as flat integers ---
                _format_id_columns(writer, df, 'Raw_Tradebook')

                # --- Apply INR currency formatting ---
                inr_format = '[$₹-en-IN] #,##0.00'

                _apply_number_format(
                    writer, grouped_df, 'Transaction',
                    columns=['Average_Price', 'Total_Value'],
                    number_format=inr_format
                )

                _apply_number_format(
                    writer, portfolio_df, 'Current_Portfolio',
                    columns=['Average_Buy_Price', 'SL', 'LTP_SL_Diff', 'Invested_Value', 'LTP',
                             'Prev_Day_Close', 'Prev_Week_Close', 'EMA9', 'EMA10', 'EMA11', 'EMA21',
                             'Current_Value', 'Unrealized_PnL'],
                    number_format=inr_format
                )

                _apply_number_format(
                    writer, overall_df, 'Overall_Portfolio',
                    columns=['Total_Buy_Value', 'Average_Buy_Price', 'Total_Sell_Value',
                             'Average_Sell_Price', 'Invested_Value', 'LTP',
                             'Current_Value', 'Realized_PnL', 'Unrealized_PnL', 'Total_PnL'],
                    number_format=inr_format
                )

                _apply_number_format(
                    writer, overall_df, 'Overall_Portfolio',
                    columns=['Total_PnL_Percentage'],
                    number_format='0.00%'
                )

                _apply_number_format(
                    writer, portfolio_df, 'Current_Portfolio',
                    columns=['LTP_SL_Diff_Pct', 'Return_Pct', 'XIRR'],
                    number_format='0.00%'
                )

                # --- Apply conditional formatting based on LTP relationship ---
                _apply_ltp_comparison_formatting(writer, portfolio_df, 'Current_Portfolio')

                # --- Apply custom color formatting for Satellite stocks based on Satellite_Watchlist ---
                # This reads color configurations from the original master file (which is intact)
                _apply_satellite_watchlist_formatting(writer, portfolio_df, output_file)

                # --- Apply custom color formatting for Core stocks based on Core_Watchlist ---
                _apply_core_watchlist_formatting(writer, portfolio_df, latest_core_trends)

                # --- Auto-fit columns and freeze panes for readability ---
                for sheet in ['Raw_Tradebook', 'Transaction', 'Current_Portfolio', 'Overall_Portfolio']:
                    if sheet in writer.sheets:
                        _auto_fit_columns_and_freeze(writer, sheet)

                # --- Create Dashboard with charts ---
                create_dashboard(writer.book, portfolio_df, overall_df, df, benchmark_returns, watchlist_df, latest_core_trends=latest_core_trends)

            # 2. Use xlwings to copy updated sheets and update watchlist columns in master workbook
            import xlwings as xw
            
            # Start Excel in visible mode (much more stable under COM for STOCKS types)
            app = xw.App(visible=True)
            app.display_alerts = False
            
            import time
            retries = 3
            for attempt in range(retries):
                try:
                    # Open both workbooks
                    wb_master = app.books.open(os.path.abspath(output_file), update_links=False)
                    
                    # Safety Check: If the file is open in another Excel instance, Excel opens it as Read-Only.
                    # Saving to a Read-Only workbook will trigger a hidden "Save As" modal and hang execution.
                    if wb_master.api.ReadOnly:
                        print(f"\nERROR: '{output_file}' is currently open in another Excel window.")
                        print("Please close the Excel file on your desktop and run the script again.\n")
                        wb_master.close()
                        raise PermissionError(f"Workbook '{output_file}' is locked/open in another Excel window.")
                    
                    # Give Excel a brief moment to settle down (especially for STOCKS data queries)
                    time.sleep(1)
                    
                    wb_temp = app.books.open(os.path.abspath(temp_file), update_links=False)
                    
                    # Get sheet names robustly to avoid COM "This object does not support enumeration" bug
                    temp_sheet_names = [wb_temp.sheets[i].name for i in range(len(wb_temp.sheets))]
                    master_sheet_names = [wb_master.sheets[i].name for i in range(len(wb_master.sheets))]
                    
                    # Copy sheet data for raw data/portfolios to preserve sheet order and formula references
                    for name in ['Raw_Tradebook', 'Transaction', 'Current_Portfolio', 'Overall_Portfolio']:
                        if name in temp_sheet_names:
                            if name not in master_sheet_names:
                                wb_master.sheets.add(name)
                                master_sheet_names.append(name)
                            ws_temp = wb_temp.sheets[name]
                            ws_master = wb_master.sheets[name]
                            
                            if name in ['Raw_Tradebook', 'Transaction']:
                                ws_master.clear()
                                ws_temp.used_range.copy(ws_master.range('A1'))
                            else:
                                # For Current_Portfolio and Overall_Portfolio, copy Column A's values
                                # and Column B onwards directly, to preserve Excel STOCKS rich data types.
                                last_row_temp = ws_temp.used_range.last_cell.row
                                last_col_temp = ws_temp.used_range.last_cell.column
                                last_row_master = ws_master.used_range.last_cell.row
                                
                                # Clear columns B onwards in master
                                ws_master.range((1, 2), (max(last_row_master, last_row_temp) + 10, last_col_temp + 10)).clear()
                                
                                # Copy Column A values directly (bypasses clipboard, preserves STOCKS formatting)
                                ws_master.range((1, 1), (last_row_temp, 1)).options(transpose=True).value = ws_temp.range((1, 1), (last_row_temp, 1)).value
                                
                                # Copy Column A formats (colors, fonts) to apply Watchlist colors without breaking STOCKS types
                                ws_temp.range((1, 1), (last_row_temp, 1)).copy()
                                ws_master.range((1, 1)).paste(paste='formats')
                                
                                # Copy Columns B onwards (direct copy bypasses system clipboard, OLE-safe)
                                ws_temp.range((1, 2), (last_row_temp, last_col_temp)).copy(ws_master.range((1, 2)))
                                
                                # Clear extra rows in Column A
                                if last_row_master > last_row_temp:
                                    ws_master.range((last_row_temp + 1, 1), (last_row_master, 1)).clear()
                    
                    # Dashboard has floating charts, so we delete and re-copy the entire sheet object
                    if 'Dashboard' in master_sheet_names:
                        wb_master.sheets['Dashboard'].delete()
                        master_sheet_names.remove('Dashboard')
                    if 'Dashboard' in temp_sheet_names:
                        wb_temp.sheets['Dashboard'].copy(before=wb_master.sheets[0])
                        wb_master.sheets['Dashboard'].name = 'Dashboard'
                        master_sheet_names.insert(0, 'Dashboard')
                        
                    # Clean up corrupted external links caused by cross-workbook sheet copying
                    for sheet_name in ['Dashboard', 'Raw_Tradebook', 'Transaction', 'Current_Portfolio', 'Overall_Portfolio']:
                        if sheet_name in master_sheet_names:
                            try:
                                # Find and replace all instances of the temporary workbook name in formulas
                                wb_master.sheets[sheet_name].api.Cells.Replace(What="[temp_transformed.xlsx]", Replacement="", LookAt=2)
                            except Exception:
                                pass
                    
                    # 3. Update Satellite_Watchlist Columns G, H, I, J using xlwings
                    if 'Satellite_Watchlist' in master_sheet_names:
                        ws_watchlist = wb_master.sheets['Satellite_Watchlist']
                        
                        # Ensure headers are set
                        ws_watchlist.range('G1').value = "Previous week Close"
                        ws_watchlist.range('H1').value = "EMA 9 (weekly)"
                        ws_watchlist.range('I1').value = "EMA 11 (weekly)"
                        ws_watchlist.range('J1').value = "EMA 21 (weekly)"
                        
                        # Read unique symbols from Column B (B2 to B<last_row>) in a single bulk COM call
                        last_row = ws_watchlist.range('B' + str(ws_watchlist.cells.last_cell.row)).end('up').row
                        if last_row >= 2:
                            symbols_values = ws_watchlist.range((2, 2), (last_row, 2)).value
                            if not isinstance(symbols_values, list):
                                symbols_values = [symbols_values]
                                
                            # Extract unique symbols
                            symbols = []
                            for sym_val in symbols_values:
                                if sym_val:
                                    sym_str = str(sym_val).strip().upper()
                                    if sym_str and sym_str not in symbols:
                                        symbols.append(sym_str)
                                        
                            if symbols:
                                from src.market_api import fetch_market_data_from_yahoo
                                classifications = {sym: 'Satellite' for sym in symbols}
                                market_data = fetch_market_data_from_yahoo(symbols, classifications=classifications, fetch_info=False)
                                
                                # Construct 2D list for bulk write to save COM roundtrips
                                rows_to_write = []
                                for sym_val in symbols_values:
                                    if sym_val:
                                        sym_str = str(sym_val).strip().upper()
                                        if sym_str in market_data:
                                            data = market_data[sym_str]
                                            rows_to_write.append([
                                                data.get('Prev_Week_Close', 0.0),
                                                data.get('EMA9', 0.0),
                                                data.get('EMA11', 0.0),
                                                data.get('EMA21', 0.0)
                                            ])
                                        else:
                                            rows_to_write.append([0.0, 0.0, 0.0, 0.0])
                                    else:
                                        rows_to_write.append([0.0, 0.0, 0.0, 0.0])
                                
                                # Bulk write Columns G, H, I, J in a single operation
                                ws_watchlist.range((2, 7), (last_row, 10)).value = rows_to_write
                                
                                # Apply INR currency formatting (₹) to the entire range in bulk
                                inr_excel_format = '[$₹-en-IN] #,##0.00'
                                ws_watchlist.range((2, 7), (last_row, 10)).number_format = inr_excel_format
                                            
                        # Auto-fit watchlist columns
                        ws_watchlist.autofit()
                    
                    # Save and close the master workbook
                    wb_master.save()
                    wb_master.close()
                    wb_temp.close()
                    print("Workbook saved successfully via xlwings (STOCKS data types preserved)!")
                    break
                except PermissionError:
                    raise
                except Exception as e:
                    if attempt == retries - 1:
                        raise
                    print(f"\nExcel was busy (COM call rejected: {e}). Retrying in 2 seconds... (Attempt {attempt+1}/{retries})")
                    try:
                        for book in list(app.books):
                            book.close()
                    except Exception:
                        pass
                    time.sleep(2)
            try:
                app.quit()
            except Exception:
                pass
                
            # Clean up the temporary file
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except Exception:
                    pass
            return
        except PermissionError as pe:
            print(f"\nERROR: Excel file lock detected. Exiting process to avoid corruption: {pe}\n")
            try:
                app.quit()
            except Exception:
                pass
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except Exception:
                    pass
            import sys
            sys.exit(1)
        except Exception as e:
            try:
                app.quit()
            except Exception:
                pass
            print(f"\nWARNING: FAILED USING XLWINGS HYBRID APPROACH: {e}")
            print("WARNING: FALLING BACK TO STANDARD OPENPYXL...")
            print("WARNING: openpyxl will strip native Excel STOCKS data types from all sheets (including Watchlists).")
            print("WARNING: To preserve STOCKS formats, close the Excel file before running main.py!\n")
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except Exception:
                    pass

    # --- FALLBACK: Standard openpyxl Writer Block ---
    try:
        mode = 'a' if os.path.exists(output_file) else 'w'
        if_sheet_exists = 'replace' if mode == 'a' else None

        with pd.ExcelWriter(output_file, engine='openpyxl', mode=mode, if_sheet_exists=if_sheet_exists) as writer:
            df.to_excel(writer, sheet_name='Raw_Tradebook', index=False)
            grouped_df.to_excel(writer, sheet_name='Transaction', index=False)
            portfolio_write_df.to_excel(writer, sheet_name='Current_Portfolio', index=False)
            overall_write_df.to_excel(writer, sheet_name='Overall_Portfolio', index=False)

            # --- Format ID columns as flat integers (no scientific notation) ---
            _format_id_columns(writer, df, 'Raw_Tradebook')

            # --- Apply INR currency formatting ---
            inr_format = '[$₹-en-IN] #,##0.00'

            _apply_number_format(
                writer, grouped_df, 'Transaction',
                columns=['Average_Price', 'Total_Value'],
                number_format=inr_format
            )

            _apply_number_format(
                writer, portfolio_df, 'Current_Portfolio',
                columns=['Average_Buy_Price', 'SL', 'LTP_SL_Diff', 'Invested_Value', 'LTP',
                         'Prev_Day_Close', 'Prev_Week_Close', 'EMA9', 'EMA10', 'EMA11', 'EMA21',
                         'Current_Value', 'Unrealized_PnL'],
                number_format=inr_format
            )

            _apply_number_format(
                writer, overall_df, 'Overall_Portfolio',
                columns=['Total_Buy_Value', 'Average_Buy_Price', 'Total_Sell_Value',
                         'Average_Sell_Price', 'Invested_Value', 'LTP',
                         'Current_Value', 'Realized_PnL', 'Unrealized_PnL', 'Total_PnL'],
                number_format=inr_format
            )

            _apply_number_format(
                writer, overall_df, 'Overall_Portfolio',
                columns=['Total_PnL_Percentage'],
                number_format='0.00%'
            )

            _apply_number_format(
                writer, portfolio_df, 'Current_Portfolio',
                columns=['LTP_SL_Diff_Pct', 'Return_Pct', 'XIRR'],
                number_format='0.00%'
            )

            # --- Apply conditional formatting based on LTP relationship ---
            _apply_ltp_comparison_formatting(writer, portfolio_df, 'Current_Portfolio')

            # --- Apply custom color formatting for Satellite stocks based on Satellite_Watchlist ---
            _apply_satellite_watchlist_formatting(writer, portfolio_df, output_file)

            # --- Apply custom color formatting for Core stocks based on Core_Watchlist ---
            _apply_core_watchlist_formatting(writer, portfolio_df, latest_core_trends)

            # --- Update columns and dynamic conditional formatting for Satellite_Watchlist ---
            print("Updating Satellite_Watchlist columns...")
            _update_satellite_watchlist_columns(writer, output_file)
            print("Applying Satellite_Watchlist formatting...")
            _apply_satellite_watchlist_conditional_formatting(writer)

            # --- Auto-fit columns and freeze panes for readability ---
            print("Auto-fitting columns...")
            for sheet in ['Raw_Tradebook', 'Transaction', 'Current_Portfolio', 'Overall_Portfolio', 'Satellite_Watchlist']:
                if sheet in writer.sheets or sheet in writer.book.sheetnames:
                    _auto_fit_columns_and_freeze(writer, sheet)

            # --- Create Dashboard with charts ---
            print("Creating Dashboard...")
            create_dashboard(writer.book, portfolio_df, overall_df, df, benchmark_returns, watchlist_df, latest_core_trends=latest_core_trends)
            
            print("Saving Excel file (this might take a moment)...")

        print("Done!")
    except PermissionError:
        print(f"Error saving {output_file}: Permission denied. Please close the Excel file if it is open and try again.")
    except Exception as e:
        print(f"Error saving {output_file}: {e}")


def _format_id_columns(writer, df: pd.DataFrame, sheet_name: str) -> None:
    """
    Converts ID columns from string values to native Excel integers
    and applies '0' number format to prevent scientific notation.

    Args:
        writer:     The active pd.ExcelWriter instance.
        df:         The DataFrame whose columns to scan.
        sheet_name: The target worksheet name.
    """
    worksheet = writer.sheets[sheet_name]
    for col_idx, col_name in enumerate(df.columns, 1):
        if 'ID' in str(col_name).upper():
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=col_idx)
                try:
                    if pd.notna(cell.value):
                        cell.value = int(float(cell.value))
                except Exception:
                    pass
                cell.number_format = '0'


def _apply_number_format(
    writer,
    df: pd.DataFrame,
    sheet_name: str,
    columns: list,
    number_format: str
) -> None:
    """
    Applies a specified Excel number format to selected columns in a worksheet.

    Args:
        writer:        The active pd.ExcelWriter instance.
        df:            The DataFrame being written to the sheet.
        sheet_name:    The target worksheet name.
        columns:       List of column names to format.
        number_format: The Excel number format string (e.g., '[$₹-en-IN] #,##0.00').
    """
    worksheet = writer.sheets[sheet_name]
    for col_idx, col_name in enumerate(df.columns, 1):
        if col_name in columns:
            for row in range(2, len(df) + 2):
                worksheet.cell(row=row, column=col_idx).number_format = number_format


def _auto_fit_columns_and_freeze(writer, sheet_name: str) -> None:
    """
    Adjusts the column widths dynamically based on the maximum string length
    in each column (including the header) and freezes the top row.
    Also formats the header row with bold text, borders, and centering.

    Args:
        writer:     The active pd.ExcelWriter instance.
        sheet_name: The target worksheet name.
    """
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    worksheet = writer.sheets[sheet_name]
    
    # Freeze the top header row
    worksheet.freeze_panes = 'A2'

    # Define Header Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name (e.g., 'A')
        
        # Apply style to the first cell (header) in the column
        header_cell = col[0]
        header_cell.font = header_font
        header_cell.alignment = header_alignment
        header_cell.border = thin_border
        header_cell.fill = header_fill

        for cell in col:
            # Apply border to all data cells in the populated range
            if 1 < cell.row <= worksheet.max_row:
                cell.border = thin_border
                
            try:
                # Calculate length of the string representation
                if cell.value is not None:
                    if cell.number_format and ('%' in cell.number_format or '₹' in cell.number_format):
                        # Give extra padding for formatted numbers (currency symbols, commas, decimals)
                        length = len(str(cell.value)) + 5
                    else:
                        length = len(str(cell.value))
                    
                    if length > max_length:
                        max_length = length
            except:
                pass
        
        # Set a slightly padded width, bound between a min and max width
        adjusted_width = min(max_length + 2, 40)
        # Ensure a minimum width so headers aren't totally squished if empty
        adjusted_width = max(adjusted_width, 10)
        
        worksheet.column_dimensions[column].width = adjusted_width


def _apply_ltp_comparison_formatting(writer, df: pd.DataFrame, sheet_name: str) -> None:
    """
    Applies native, dynamic conditional formatting to 'Prev_Day_Close' and 'Prev_Week_Close'
    based on their relationship with 'LTP'.
    If LTP > Close: Light Green fill ('E2EFDA')
    Else: Light Pink fill ('FCE4D6')
    """
    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.utils import get_column_letter
    
    if sheet_name not in writer.sheets:
        return
        
    worksheet = writer.sheets[sheet_name]
    
    # Map column headers to 1-based indices
    col_map = {col_name: col_idx for col_idx, col_name in enumerate(df.columns, 1)}
    
    if 'LTP' not in col_map:
        return
        
    ltp_idx = col_map['LTP']
    ltp_letter = get_column_letter(ltp_idx)
    
    green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    pink_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    
    for target_col in ['Prev_Day_Close', 'Prev_Week_Close']:
        if target_col in col_map:
            target_idx = col_map[target_col]
            target_letter = get_column_letter(target_idx)
            
            # Target range from row 2 to end of data
            range_str = f"{target_letter}2:{target_letter}{len(df) + 1}"
            
            # Formulas adapted for row 2 (relative rows so Excel propagates them)
            green_formula = f"=${ltp_letter}2>${target_letter}2"
            pink_formula = f"=${ltp_letter}2<=${target_letter}2"
            
            green_rule = FormulaRule(formula=[green_formula], fill=green_fill)
            pink_rule = FormulaRule(formula=[pink_formula], fill=pink_fill)
            
            worksheet.conditional_formatting.add(range_str, green_rule)
            worksheet.conditional_formatting.add(range_str, pink_rule)

    # 2. Apply Trend column color formatting
    if 'Trend' in col_map:
        trend_idx = col_map['Trend']
        trend_letter = get_column_letter(trend_idx)
        range_str = f"{trend_letter}2:{trend_letter}{len(df) + 1}"
        
        # Center align Trend cells
        from openpyxl.styles import Alignment
        center_align = Alignment(horizontal='center')
        for r_idx in range(2, len(df) + 2):
            worksheet.cell(row=r_idx, column=trend_idx).alignment = center_align

        # Red and Green fonts for up and down arrows
        from openpyxl.styles import Font
        GREEN_FONT = Font(name='Calibri', bold=True, color='147A1E')
        RED_FONT = Font(name='Calibri', bold=True, color='C00000')
        
        from openpyxl.formatting.rule import CellIsRule
        worksheet.conditional_formatting.add(
            range_str,
            CellIsRule(operator='equal', formula=['"▲"'], font=GREEN_FONT)
        )
        worksheet.conditional_formatting.add(
            range_str,
            CellIsRule(operator='equal', formula=['"▼"'], font=RED_FONT)
        )


def _apply_satellite_watchlist_formatting(writer, df: pd.DataFrame, output_file: str) -> None:
    """
    Reads the Satellite_Watchlist sheet from the existing workbook and styles
    the Symbol cell in the Current_Portfolio sheet for any matching 'Satellite'
    stocks with their corresponding watchlist color on the latest watchlist date.
    Clears background fill and resets font for stocks not in the latest date's watchlist.
    """
    import os
    import pandas as pd
    from openpyxl.styles import PatternFill, Font
    
    # Premium color palettes (light background fill, dark font text)
    COLOR_MAP = {
        'BLUE':   {'bg': 'DDEBF7', 'font': '1F4E78'},
        'ORANGE': {'bg': 'FCE4D6', 'font': 'C65911'},
        'GREEN':  {'bg': 'E2EFDA', 'font': '375623'},
        'RED':    {'bg': 'FADBD8', 'font': 'A93226'},
        'PINK':   {'bg': 'FADBD8', 'font': 'A93226'},
        'YELLOW': {'bg': 'FFF2CC', 'font': '7F6000'},
        'PURPLE': {'bg': 'E1D5E7', 'font': '60497A'}
    }

    if not os.path.exists(output_file):
        return

    try:
        with pd.ExcelFile(output_file) as xls:
            watchlist_df = pd.read_excel(xls, sheet_name='Satellite_Watchlist')
    except Exception as e:
        print(f"Note: Satellite_Watchlist sheet could not be loaded from existing workbook: {e}")
        return

    try:
        # Filter out empty entries and sanitize dates
        watchlist_df = watchlist_df.dropna(subset=['Stock', 'Color'])
        watchlist_df['Stock'] = watchlist_df['Stock'].astype(str).str.strip().str.upper()
        watchlist_df['Color'] = watchlist_df['Color'].astype(str).str.strip().str.upper()
        
        # Sort chronologically by Date (newest first) to find the latest
        watchlist_df['Date'] = pd.to_datetime(watchlist_df['Date'], dayfirst=True, errors='coerce')
        watchlist_df = watchlist_df.dropna(subset=['Date'])
        
        # Enforce strict latest date filtering
        latest_colors = {}
        if not watchlist_df.empty:
            latest_date = watchlist_df['Date'].max()
            watchlist_df = watchlist_df[watchlist_df['Date'] == latest_date]
            
            # Deduplicate to keep only the latest color code per stock symbol on the latest date
            latest_colors = watchlist_df.drop_duplicates(subset=['Stock']).set_index('Stock')['Color'].to_dict()
    except Exception as e:
        print(f"Error parsing Satellite_Watchlist data: {e}")
        return

    if 'Current_Portfolio' not in writer.sheets:
        return
        
    worksheet = writer.sheets['Current_Portfolio']
    col_map = {col_name: col_idx for col_idx, col_name in enumerate(df.columns, 1)}

    if 'Symbol' not in col_map or 'TF_Classification' not in col_map:
        return

    symbol_idx = col_map['Symbol']
    class_idx = col_map['TF_Classification']

    # Apply styling row-by-row
    for row in range(2, len(df) + 2):
        class_val = str(worksheet.cell(row=row, column=class_idx).value).strip()
        if class_val == 'Satellite':
            symbol_cell = worksheet.cell(row=row, column=symbol_idx)
            symbol = str(symbol_cell.value).strip().upper()
            
            if symbol in latest_colors:
                color_name = latest_colors[symbol]
                if color_name in COLOR_MAP:
                    bg_color = COLOR_MAP[color_name]['bg']
                    font_color = COLOR_MAP[color_name]['font']
                    
                    symbol_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                    symbol_cell.font = Font(color=font_color, bold=True)
                else:
                    symbol_cell.fill = PatternFill(fill_type=None)
                    symbol_cell.font = Font(color='000000', bold=False)
            else:
                # Stock not in the latest date's watchlist: Clear fill and restore standard style
                symbol_cell.fill = PatternFill(fill_type=None)
                symbol_cell.font = Font(color='000000', bold=False)


def _apply_core_watchlist_formatting(writer, df: pd.DataFrame, latest_core_trends: dict) -> None:
    """
    Reads the Core_Watchlist sheet from the existing workbook and styles
    the Symbol cell in the Current_Portfolio sheet for any matching 'Core'
    stocks with their corresponding watchlist color.
    """
    if 'Current_Portfolio' not in writer.sheets:
        return
        
    worksheet = writer.sheets['Current_Portfolio']
    col_map = {col_name: col_idx for col_idx, col_name in enumerate(df.columns, 1)}

    if 'Symbol' not in col_map or 'TF_Classification' not in col_map:
        return

    symbol_idx = col_map['Symbol']
    class_idx = col_map['TF_Classification']

    from openpyxl.styles import PatternFill, Font
    
    COLOR_MAP = {
        'Strong Trend- Green': '00B050',
        'Medium Trend':        'FFC000',
        'Weak Trend- Red':    'DC3939',
        'Core-Weekly':         'B1A0C7'
    }

    # Apply styling row-by-row
    for row in range(2, len(df) + 2):
        class_val = str(worksheet.cell(row=row, column=class_idx).value).strip()
        if class_val == 'Core':
            symbol_cell = worksheet.cell(row=row, column=symbol_idx)
            symbol = str(symbol_cell.value).strip().upper()
            
            if symbol in latest_core_trends:
                trend = latest_core_trends[symbol]
                if trend in COLOR_MAP:
                    bg_color = COLOR_MAP[trend]
                    symbol_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                    # Make font bold and black for high readability on these backgrounds
                    symbol_cell.font = Font(color='000000', bold=True)
                else:
                    symbol_cell.fill = PatternFill(fill_type=None)
                    symbol_cell.font = Font(color='000000', bold=False)
            else:
                symbol_cell.fill = PatternFill(fill_type=None)
                symbol_cell.font = Font(color='000000', bold=False)


def _update_satellite_watchlist_columns(writer, output_file) -> None:
    """
    Reads unique stocks from the Satellite_Watchlist sheet in writer.book,
    downloads weekly data (Prev_Week_Close, EMA 9, EMA 11, EMA 21) from Yahoo Finance,
    and populates Columns G, H, I, J in the Satellite_Watchlist sheet.
    Preserves Columns D, E, F exactly.
    """
    from src.market_api import fetch_market_data_from_yahoo
    
    if 'Satellite_Watchlist' not in writer.book.sheetnames:
        return

    ws = writer.book['Satellite_Watchlist']
    
    # Ensure headers are set
    ws.cell(row=1, column=7, value="Previous week Close")
    ws.cell(row=1, column=8, value="EMA 9 (weekly)")
    ws.cell(row=1, column=9, value="EMA 11 (weekly)")
    ws.cell(row=1, column=10, value="EMA 21 (weekly)")
    
    # Read unique symbols from Column B (row 2 to max_row)
    symbols = []
    for r in range(2, ws.max_row + 1):
        sym_val = ws.cell(row=r, column=2).value
        if sym_val:
            sym_str = str(sym_val).strip().upper()
            if sym_str and sym_str not in symbols:
                symbols.append(sym_str)
                
    if not symbols:
        return
        
    # Map symbols to 'Satellite' classification so fetch_market_data_from_yahoo Resamples to weekly
    classifications = {sym: 'Satellite' for sym in symbols}
    
    # Fetch market data
    market_data = fetch_market_data_from_yahoo(symbols, classifications=classifications, fetch_info=False)
    
    # Formatting for currency columns
    inr_format = '[$₹-en-IN] #,##0.00'
    
    # Write values row-by-row
    for r in range(2, ws.max_row + 1):
        sym_val = ws.cell(row=r, column=2).value
        if sym_val:
            sym_str = str(sym_val).strip().upper()
            if sym_str in market_data:
                data = market_data[sym_str]
                
                # Column G: Previous week Close
                cell_g = ws.cell(row=r, column=7, value=data.get('Prev_Week_Close', 0.0))
                cell_g.number_format = inr_format
                
                # Column H: EMA 9 (weekly)
                cell_h = ws.cell(row=r, column=8, value=data.get('EMA9', 0.0))
                cell_h.number_format = inr_format
                
                # Column I: EMA 11 (weekly)
                cell_i = ws.cell(row=r, column=9, value=data.get('EMA11', 0.0))
                cell_i.number_format = inr_format
                
                # Column J: EMA 21 (weekly)
                cell_j = ws.cell(row=r, column=10, value=data.get('EMA21', 0.0))
                cell_j.number_format = inr_format


def _apply_satellite_watchlist_conditional_formatting(writer) -> None:
    """
    Applies dynamic conditional formatting to the Satellite_Watchlist worksheet.
    - If Price (Column E) > Previous Close (Column F) -> Column F becomes green.
    - If Previous Close (Column F) > Previous week Close (Column G) -> Column G becomes green.
    """
    from openpyxl.styles import PatternFill, Font
    from openpyxl.formatting.rule import FormulaRule
    
    if 'Satellite_Watchlist' not in writer.book.sheetnames:
        return
        
    ws = writer.book['Satellite_Watchlist']
    max_row = ws.max_row
    if max_row < 2:
        return
        
    green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    green_font = Font(color='375623', bold=True)
    
    # Clear old conditional formatting rules for the watchlist if any exist to avoid compounding
    if hasattr(ws, 'conditional_formatting') and hasattr(ws.conditional_formatting, '_cf_rules'):
        ws.conditional_formatting._cf_rules.clear()
    
    # Range for Column F: F2:F<max_row>
    range_f = f"F2:F{max_row}"
    # Range for Column G: G2:G<max_row>
    range_g = f"G2:G{max_row}"
    
    # Formulas relative to row 2
    formula_f = "=$E2>$F2"
    formula_g = "=$F2>$G2"
    
    rule_f = FormulaRule(formula=[formula_f], fill=green_fill, font=green_font)
    rule_g = FormulaRule(formula=[formula_g], fill=green_fill, font=green_font)
    
    ws.conditional_formatting.add(range_f, rule_f)
    ws.conditional_formatting.add(range_g, rule_g)
